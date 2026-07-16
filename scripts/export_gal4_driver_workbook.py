#!/usr/bin/env python3
"""
Export stock and phenotype evidence for a list of GAL4 driver symbols.

Reads a CSV with a ``GAL4 symbol`` column, resolves orderable GAL4 driver stocks
(including mixed-genotype Bloomington lines) via FlyBase derived stock components,
builds curated phenotype rows using the same pipeline logic as the All
Phenotypic Stocks Sheet, and writes an Excel workbook with:

- One sheet per stock collection (BDSC, VDRC, Custom Lab Reagent, etc.), with
  one row per unique GAL4 stock/reagent and phenotype/reference metadata
  aggregated onto that stock row.
- ``References``: one row per unique PMID/PMCID with expanded metadata and the
  associated GAL4 reagents/stocks.

Example::

    python scripts/export_gal4_driver_workbook.py gal4_list.csv -o gal4_stocks.xlsx
"""

from __future__ import annotations

import argparse
import math
import re
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Set, Tuple

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from fl_ai_reagent_stocker.config import Settings  # noqa: E402
from fl_ai_reagent_stocker.utils import (  # noqa: E402
    REAGENT_BUCKET_COLUMNS,
    canonical_stock_candidate_label,
    clean_id,
    find_latest_tsv,
    format_pmcid_display,
    format_pmid_display,
    format_reference_id_display,
    load_flybase_tsv,
    normalize_collection_name,
    unique_join,
)

GAL4_SYMBOL_COLUMN = "GAL4 symbol"
INPUT_SYMBOL_COLUMN = "GAL4 symbol (input)"
SOURCE_STOCK_COLUMN = "Source/ Stock #"
STOCK_GENOTYPE_COLUMN = "Stock genotype (output)"
DRIVER_GENOTYPE_COLUMN = "Resolved driver genotype (output)"
PHENOTYPE_GENOTYPES_COLUMN = "Phenotype record genotypes"
INPUT_SYMBOLS_COLUMN = "GAL4 symbols (input)"
STOCK_ID_COLUMN = "Stock ID"
REFERENCES_SHEET_NAME = "References"
COVERAGE_SHEET_NAME = "Coverage"
EXCEL_SHEET_NAME_MAX = 31
INVALID_SHEET_CHARS = re.compile(r"[:\\/?*\[\]]")
FLYBASE_ID_RE = re.compile(r"FB[a-zA-Z]{2,4}[0-9]+")
CUSTOM_LAB_REAGENT_COLLECTION = "Custom Lab Reagent"
BDSC_STOCKS_FILENAME = "bloomington.csv"
BDSC_STOCK_COMPONENTS_FILENAME = "stockgenes.csv"
BDSC_INTERNAL_FBST_PREFIX = "BDSC:"


@dataclass
class Gal4StockMatch:
    """Resolved GAL4 driver stock for an input symbol."""

    input_symbol: str
    fbst: str
    stock_number: str
    collection: str
    source_stock: str
    stock_genotype: str = ""
    component_ids: Set[str] = field(default_factory=set)
    component_symbols: Set[str] = field(default_factory=set)
    driver_genotype: str = ""
    resolution_method: str = ""
    match_confidence: str = ""


@dataclass
class StockResolutionIndex:
    """Precomputed derived-stock lookups to avoid repeated full-table scans."""

    label_to_fbst: Dict[str, str]
    label_to_gal4_component_ids: Dict[str, Set[str]]
    fbst_to_stock_meta: Dict[str, Dict[str, str]]
    fbst_to_component_symbols: Dict[str, Dict[str, str]]
    gal4_component_records: List[Tuple[str, str, str, str]]
    gal4_component_records_by_key: Dict[str, List[Tuple[str, str, str, str]]]


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Export stock, phenotype, and reference metadata for GAL4 drivers "
            "listed in a CSV."
        )
    )
    parser.add_argument(
        "input_csv",
        type=Path,
        help=f"CSV file containing a '{GAL4_SYMBOL_COLUMN}' column.",
    )
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=None,
        help="Output .xlsx path (default: <input_stem>_gal4_stocks.xlsx).",
    )
    parser.add_argument(
        "--no-fetch-pubmed",
        action="store_true",
        help="Do not fetch missing PubMed metadata for phenotype references.",
    )
    parser.add_argument(
        "--quiet",
        action="store_true",
        help="Suppress progress messages.",
    )
    return parser


def _read_gal4_symbols(csv_path: Path) -> List[str]:
    header = pd.read_csv(csv_path, nrows=0)
    if GAL4_SYMBOL_COLUMN not in header.columns:
        raise ValueError(
            f"Input CSV must contain a '{GAL4_SYMBOL_COLUMN}' column; "
            f"found: {', '.join(header.columns)}"
        )
    df = pd.read_csv(csv_path, dtype=str).fillna("")
    symbols: List[str] = []
    seen: Set[str] = set()
    for raw in df[GAL4_SYMBOL_COLUMN].astype(str):
        symbol = raw.strip()
        if not symbol or symbol in seen:
            continue
        seen.add(symbol)
        symbols.append(symbol)
    if not symbols:
        raise ValueError(f"No GAL4 symbols found in '{GAL4_SYMBOL_COLUMN}' column.")
    return symbols


def _load_derived_components(settings: Settings) -> pd.DataFrame:
    path = settings.flybase_derived_components_path
    if not path.exists():
        raise FileNotFoundError(f"Derived stock component CSV not found: {path}")
    df = pd.read_csv(path, dtype=str).fillna("")
    if "collection_short_name" in df.columns and "collection" not in df.columns:
        df = df.rename(columns={"collection_short_name": "collection"})
    for column in (
        "FBst",
        "stock_number",
        "collection",
        "derived_stock_component",
        "object_symbol",
        "GeneSymbol",
        "FB_genotype",
    ):
        if column not in df.columns:
            df[column] = ""
    df["FBst"] = df["FBst"].apply(clean_id)
    df["stock_number"] = df["stock_number"].apply(clean_id)
    df["derived_stock_component"] = df["derived_stock_component"].apply(clean_id)
    df["collection"] = df["collection"].fillna("").astype(str).str.strip()
    return df


def _merge_native_bdsc_stock_data(
    derived_df: pd.DataFrame,
    bdsc_data_path: Path,
) -> pd.DataFrame:
    """Overlay current BDSC availability and add native GAL4 component names."""
    stocks_path = bdsc_data_path / BDSC_STOCKS_FILENAME
    components_path = bdsc_data_path / BDSC_STOCK_COMPONENTS_FILENAME
    if not stocks_path.exists() or not components_path.exists():
        return derived_df

    stocks_df = pd.read_csv(
        stocks_path,
        dtype=str,
        usecols=["Stk #", "Genotype", "A.K.A"],
    ).fillna("")
    stocks_df["Stk #"] = stocks_df["Stk #"].apply(clean_id)
    stocks_df = stocks_df[stocks_df["Stk #"] != ""].drop_duplicates(
        subset=["Stk #"],
        keep="first",
    )
    available_stock_numbers = set(stocks_df["Stk #"])
    genotype_by_stock = dict(zip(stocks_df["Stk #"], stocks_df["Genotype"]))

    merged = derived_df.copy()
    merged["stock_number"] = merged["stock_number"].apply(clean_id)
    bloomington_mask = merged["collection"].apply(
        lambda value: normalize_collection_name(value) == "Bloomington"
    )
    merged = merged[
        ~bloomington_mask
        | merged["stock_number"].isin(available_stock_numbers)
    ].copy()
    bloomington_mask = merged["collection"].apply(
        lambda value: normalize_collection_name(value) == "Bloomington"
    )
    native_genotypes = merged["stock_number"].map(genotype_by_stock).fillna("")
    merged.loc[
        bloomington_mask & (native_genotypes != ""),
        "FB_genotype",
    ] = native_genotypes

    stock_to_fbst: Dict[str, str] = {}
    existing_symbol_keys: Set[Tuple[str, str]] = set()
    for stock_number, collection, fbst, symbol in zip(
        merged["stock_number"],
        merged["collection"],
        merged["FBst"],
        merged["object_symbol"],
    ):
        if normalize_collection_name(collection) != "Bloomington":
            continue
        stock_number = clean_id(stock_number)
        fbst = clean_id(fbst)
        if stock_number and fbst:
            stock_to_fbst.setdefault(stock_number, fbst)
        symbol_key = _compact_symbol_key(symbol)
        if stock_number and symbol_key:
            existing_symbol_keys.add((stock_number, symbol_key))

    native_rows: List[Dict[str, str]] = []

    def add_native_component(
        stock_number: Any,
        genotype: Any,
        symbol: Any,
        gene_symbol: Any,
        native_component_id: Any,
    ) -> None:
        stock_number_text = clean_id(stock_number)
        symbol_text = str(symbol or "").strip()
        gene_symbol_text = str(gene_symbol or "").strip()
        symbol_key = _compact_symbol_key(symbol_text)
        if (
            not stock_number_text
            or stock_number_text not in available_stock_numbers
            or not symbol_key
            or (stock_number_text, symbol_key) in existing_symbol_keys
            or not _looks_like_gal4_symbol(symbol_text, gene_symbol_text)
        ):
            return
        existing_symbol_keys.add((stock_number_text, symbol_key))
        component_suffix = clean_id(native_component_id) or symbol_key
        native_rows.append(
            {
                "FBst": stock_to_fbst.get(
                    stock_number_text,
                    f"{BDSC_INTERNAL_FBST_PREFIX}{stock_number_text}",
                ),
                "stock_number": stock_number_text,
                "collection": "Bloomington",
                "FB_genotype": str(genotype or "").strip()
                or genotype_by_stock.get(stock_number_text, ""),
                "derived_stock_component": (
                    f"BDSCcomp:{stock_number_text}:{component_suffix}"
                ),
                "embedded_type": "BDSC",
                "object_symbol": symbol_text,
                "GeneSymbol": gene_symbol_text,
            }
        )

    for chunk in pd.read_csv(
        components_path,
        dtype=str,
        usecols=[
            "stknum",
            "genotype",
            "component_symbol",
            "gene_symbol",
            "bdsc_symbol_id",
        ],
        chunksize=50_000,
    ):
        chunk = chunk.fillna("")
        chunk["stknum"] = chunk["stknum"].apply(clean_id)
        chunk = chunk[chunk["stknum"].isin(available_stock_numbers)]
        gal4_mask = (
            chunk["component_symbol"].astype(str)
            + " "
            + chunk["gene_symbol"].astype(str)
        ).str.contains(r"gal4|gawb", case=False, na=False, regex=True)
        for row in chunk[gal4_mask].itertuples(index=False):
            add_native_component(
                row.stknum,
                row.genotype,
                row.component_symbol,
                row.gene_symbol,
                row.bdsc_symbol_id,
            )

    for stock_number, genotype, aka in stocks_df[
        ["Stk #", "Genotype", "A.K.A"]
    ].itertuples(index=False, name=None):
        add_native_component(
            stock_number,
            genotype,
            aka,
            "",
            "AKA",
        )

    if native_rows:
        merged = pd.concat(
            [merged, pd.DataFrame(native_rows)],
            ignore_index=True,
            sort=False,
        ).fillna("")
    return merged


def _extract_flybase_ids(value: Any) -> List[str]:
    """Extract canonical FlyBase IDs from genotype ID fields."""
    deduped: List[str] = []
    seen: Set[str] = set()
    for match in FLYBASE_ID_RE.findall(str(value or "")):
        match_clean = clean_id(match)
        if match_clean and match_clean not in seen:
            seen.add(match_clean)
            deduped.append(match_clean)
    return deduped


def _normalize_symbol_token(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).lower()


def _compact_symbol_key(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").lower())


def _split_genotype_symbol_tokens(value: Any) -> List[str]:
    tokens: List[str] = []
    seen: Set[str] = set()
    raw_parts = re.split(r"[;|,\n]+", str(value or ""))
    for raw_part in raw_parts:
        for slash_part in str(raw_part).split("/"):
            for candidate in re.split(r"\s+", str(slash_part).strip()):
                token = candidate.strip()
                if not token or token in {"+", "-"}:
                    continue
                normalized = _normalize_symbol_token(token)
                if normalized in seen:
                    continue
                seen.add(normalized)
                tokens.append(token)
    return tokens


def _format_stock_candidate_label(
    fbst: Any,
    collection: Any,
    stock_number: Any,
) -> str:
    """Format a stock candidate label for best-effort partner resolution."""
    stock_number_str = clean_id(stock_number)
    collection_str = normalize_collection_name(collection)
    if stock_number_str and collection_str:
        return f"({stock_number_str}, {collection_str})"
    if stock_number_str:
        return f"({stock_number_str})"
    if collection_str:
        return f"({collection_str})"
    return clean_id(fbst)


def _format_source_stock_label(collection: Any, stock_number: Any) -> str:
    """Format the source-stock label used by phenotype evidence sheets."""
    collection_str = normalize_collection_name(collection)
    stock_number_str = str(stock_number or "").strip()
    if collection_str and collection_str.lower() == "nan":
        collection_str = ""
    if stock_number_str and stock_number_str.lower() == "nan":
        stock_number_str = ""
    if collection_str and stock_number_str:
        return f"{collection_str} ({stock_number_str})"
    return stock_number_str or collection_str


def _might_be_gal4_component(symbol: str, gene_symbol: str) -> bool:
    blob = f"{symbol} {gene_symbol}".lower()
    return (
        "gal4" in blob
        or "gawb" in blob
        or "-gal4" in blob
        or "lexa" in blob
        or blob.startswith("p{")
    )


def _normalized_text_looks_like_gal4_driver(normalized: str) -> bool:
    """Return True when a normalized symbol/gene fragment is a GAL4 driver."""
    if not normalized:
        return False
    if "gal4" in normalized:
        return True
    if "gawb" in normalized:
        return True
    if re.search(r"p\{[^}]*gal4", normalized):
        return True
    if re.search(r"gal4::", normalized):
        return True
    if "aygal4" in normalized:
        return True
    if re.search(r"-gal4\b", normalized):
        return True
    if re.search(r"\bvt\d*-gal4\b", normalized):
        return True
    if re.search(r"\bgmr\d*-gal4\b", normalized):
        return True
    if re.search(r"p\{gawb", normalized):
        return True
    return False


def _looks_like_gal4_symbol(symbol: Any, gene_symbol: Any = "") -> bool:
    """Return True when the symbol or linked gene clearly indicates GAL4."""
    for raw_value in (symbol, gene_symbol):
        normalized = _normalize_symbol_token(raw_value)
        if normalized and _normalized_text_looks_like_gal4_driver(normalized):
            return True
    return False


def _gal4_search_keys(symbol: Any) -> Set[str]:
    """Return compact lookup keys for GAL4 driver names across FlyBase formats."""
    text = str(symbol or "").strip()
    lowered = text.lower()
    keys: Set[str] = set()

    def add(value: Any) -> None:
        key = _compact_symbol_key(value)
        if len(key) >= 3:
            keys.add(key)
        if key.endswith("gal4") and len(key) > 4:
            keys.add(key[:-4])
        if "gal4" in key:
            prefix = key.split("gal4", 1)[0]
            if len(prefix) >= 3:
                keys.add(prefix)
        if key.startswith("gmr") and len(key) > 6 and re.match(r"gmr[0-9]", key):
            keys.add(key[3:])
            keys.add(f"r{key[3:]}")
        if key.startswith("ey") and len(key) > 5 and re.match(r"ey[a-z]*[0-9]", key):
            keys.add(key[2:])

    add(text)
    for inner in re.findall(r"\{([^}]*)\}", text):
        cleaned_inner = re.sub(r"^[^=]+=", "", inner)
        cleaned_inner = re.sub(r"^(it\.)", "", cleaned_inner, flags=re.IGNORECASE)
        add(cleaned_inner)
    for inner in re.findall(r"\[([^\]]+)\]", text):
        add(inner)
    for prefix, suffix in re.findall(r"\}([A-Za-z][A-Za-z0-9_.\\-]*)\[([^\]]+)\]", text):
        add(f"{prefix}{suffix}")
    for token in re.split(r"[^A-Za-z0-9_.()\\-]+", text):
        add(token)

    # Common FlyBase encodings keep the driver name outside the construct braces.
    without_wrappers = re.sub(r"[A-Za-z]*\{[^}]*\}", " ", text)
    for token in re.split(r"[^A-Za-z0-9_.()\\-]+", without_wrappers):
        add(token)

    if "scer\\gal4" in lowered:
        for inner in re.findall(r"scer\\gal4\[([^\]]+)\]", lowered):
            add(inner)
    return keys


def _input_gal4_lookup_keys(input_symbol: Any) -> Set[str]:
    keys = set(_gal4_search_keys(input_symbol))
    for token in _shorthand_search_tokens(str(input_symbol or "")):
        keys.update(_gal4_search_keys(token))
    return keys


def _informative_gal4_alias_keys(symbol: Any) -> Set[str]:
    """Return driver-specific keys, excluding generic GAL4 species/tool labels."""
    generic_keys = {
        "driver",
        "gal4",
        "gawb",
        "p",
        "scer",
        "scergal4",
    }
    return {
        key
        for key in _gal4_search_keys(symbol)
        if key not in generic_keys and len(key) >= 3
    }


def _build_gal4_construct_alias_fbti_index(
    construct_df: pd.DataFrame,
    fbtp_to_fbti_df: pd.DataFrame,
) -> Dict[str, Set[str]]:
    """Map FlyBase GAL4 allele/construct aliases to their insertion IDs."""
    required_construct_columns = {
        "Component Allele (symbol)",
        "Transgenic Construct (symbol)",
        "Transgenic Construct (id)",
    }
    if not required_construct_columns.issubset(construct_df.columns):
        return {}
    if not {"FBtp", "FBti"}.issubset(fbtp_to_fbti_df.columns):
        return {}

    fbtp_to_fbtis: Dict[str, Set[str]] = defaultdict(set)
    for fbtp_raw, fbti_raw in zip(
        fbtp_to_fbti_df["FBtp"].tolist(),
        fbtp_to_fbti_df["FBti"].tolist(),
    ):
        fbtp_id = clean_id(fbtp_raw)
        fbti_id = clean_id(fbti_raw)
        if fbtp_id and fbti_id:
            fbtp_to_fbtis[fbtp_id].add(fbti_id)

    alias_index: Dict[str, Set[str]] = defaultdict(set)
    for allele_symbol, construct_symbol, fbtp_raw in zip(
        construct_df["Component Allele (symbol)"].fillna("").astype(str),
        construct_df["Transgenic Construct (symbol)"].fillna("").astype(str),
        construct_df["Transgenic Construct (id)"].fillna("").astype(str),
    ):
        fbtp_ids = {
            clean_id(value)
            for value in str(fbtp_raw or "").split("|")
            if clean_id(value)
        }
        linked_fbtis = {
            fbti_id
            for fbtp_id in fbtp_ids
            for fbti_id in fbtp_to_fbtis.get(fbtp_id, set())
        }
        if not linked_fbtis or not _looks_like_gal4_symbol(
            allele_symbol,
            construct_symbol,
        ):
            continue
        for alias in (allele_symbol, construct_symbol):
            for key in _informative_gal4_alias_keys(alias):
                alias_index[key].update(linked_fbtis)
    return dict(alias_index)


def _load_gal4_construct_alias_fbti_index(
    flybase_data_path: Path,
) -> Dict[str, Set[str]]:
    """Load the FlyBase FBal/FBtp alias chain used to reach stock FBti IDs."""
    constructs_dir = flybase_data_path / "transgenic_constructs"
    mapping_path = flybase_data_path / "transgenic_insertions" / "fbtp_to_fbti.csv"
    if not mapping_path.exists():
        return {}
    try:
        construct_df = load_flybase_tsv(
            find_latest_tsv(constructs_dir, "transgenic_construct_descriptions")
        )
    except FileNotFoundError:
        return {}
    fbtp_to_fbti_df = pd.read_csv(mapping_path, dtype=str).fillna("")
    return _build_gal4_construct_alias_fbti_index(
        construct_df,
        fbtp_to_fbti_df,
    )


def _is_split_gal4_component(symbol: Any) -> bool:
    lowered = str(symbol or "").lower()
    return bool(
        ".dbd" in lowered
        or "-dbd" in lowered
        or ".ad" in lowered
        or "-ad" in lowered
        or "p65.ad" in lowered
        or "split-gal4" in lowered
    )


def _input_requests_split_gal4(input_symbol: Any) -> bool:
    lowered = str(input_symbol or "").lower()
    return bool(
        ".dbd" in lowered
        or "-dbd" in lowered
        or ".ad" in lowered
        or "-ad" in lowered
        or "p65" in lowered
        or "split-gal4" in lowered
    )


def _gal4_driver_fingerprint(symbol: Any) -> str:
    keys = _gal4_search_keys(symbol)
    if not keys:
        return ""
    return sorted(keys, key=lambda value: (-len(value), value))[0]


def _gal4_driver_symbols_match(left: Any, right: Any) -> bool:
    """Return True when two labels refer to the same GAL4 driver."""
    left_keys = _gal4_search_keys(left)
    right_keys = _gal4_search_keys(right)
    if not left_keys or not right_keys:
        return False
    if left_keys & right_keys:
        return True
    for left_key in left_keys:
        for right_key in right_keys:
            if len(left_key) >= 5 and len(right_key) >= 5 and (
                left_key in right_key or right_key in left_key
            ):
                return True
    return False


def _primary_gal4_component_symbol_for_fbst(
    fbst: str,
    component_rows_by_fbst: Dict[str, List[Tuple[str, str]]],
    fbst_to_fb_genotype: Dict[str, str],
) -> str:
    """Return the best FlyBase component symbol for a GAL4 stock."""
    gal4_symbols: List[str] = []
    for symbol, gene_symbol in component_rows_by_fbst.get(fbst, []):
        symbol_text = str(symbol or "").strip()
        if symbol_text and _looks_like_gal4_symbol(symbol_text, gene_symbol):
            gal4_symbols.append(symbol_text)
    if gal4_symbols:
        gal4_symbols.sort(
            key=lambda value: (
                _is_split_gal4_component(value),
                not value.startswith("P{"),
                not value.startswith("M{"),
                -len(value),
                value.lower(),
            )
        )
        return gal4_symbols[0]

    fb_genotype = str(fbst_to_fb_genotype.get(fbst, "") or "").strip()
    for token in _split_genotype_symbol_tokens(fb_genotype):
        if _looks_like_gal4_symbol(token):
            return token
    return ""


def _build_stock_candidate_label_to_driver_genotype(
    gal4_only_fbst: Set[str],
    component_id_to_stock_candidate_details: Dict[str, Set[Tuple[str, str]]],
    component_rows_by_fbst: Dict[str, List[Tuple[str, str]]],
    fbst_to_fb_genotype: Dict[str, str],
) -> Dict[str, str]:
    """Map stock labels to orderable driver component symbols."""
    fbst_to_symbol = {
        fbst: _primary_gal4_component_symbol_for_fbst(
            fbst,
            component_rows_by_fbst,
            fbst_to_fb_genotype,
        )
        for fbst in gal4_only_fbst
    }
    label_to_genotype: Dict[str, str] = {}
    for candidate_details in component_id_to_stock_candidate_details.values():
        for fbst, label in candidate_details:
            if not label or fbst not in gal4_only_fbst:
                continue
            genotype = str(fbst_to_symbol.get(fbst, "") or "").strip()
            if genotype:
                label_to_genotype[label] = genotype
    return label_to_genotype


def _build_gal4_driver_stock_candidate_index(
    component_id_to_symbol: Dict[str, str],
    component_id_to_gene_symbol: Dict[str, str],
    component_id_to_stock_candidate_details: Dict[str, Set[Tuple[str, str]]],
) -> Dict[str, Set[Tuple[str, str]]]:
    """Index GAL4 stock candidates by compact driver lookup key."""
    index: Dict[str, Set[Tuple[str, str]]] = defaultdict(set)
    for component_id, candidate_details in component_id_to_stock_candidate_details.items():
        symbol = component_id_to_symbol.get(component_id, "")
        gene_symbol = component_id_to_gene_symbol.get(component_id, "")
        if not _looks_like_gal4_symbol(symbol, gene_symbol):
            continue
        for key in _gal4_search_keys(symbol):
            index[key].update(candidate_details)
    return index


def _resolve_gal4_stock_candidates_by_driver_symbol(
    partner_symbol: Any,
    driver_stock_index: Dict[str, Set[Tuple[str, str]]],
    gal4_only_fbst: Set[str],
) -> Set[str]:
    """Resolve GAL4-only stock labels from an input driver symbol."""
    candidates: Set[str] = set()
    for key in _input_gal4_lookup_keys(partner_symbol):
        for fbst, label in driver_stock_index.get(key, set()):
            if fbst in gal4_only_fbst and label:
                candidates.add(label)
    return candidates


def _build_gal4_lookup_context(
    derived_df: pd.DataFrame,
) -> Tuple[
    Dict[str, str],
    Dict[str, str],
    Dict[str, Set[Tuple[str, str]]],
    Dict[str, List[Tuple[str, str]]],
    Set[str],
    Dict[str, str],
    Dict[str, Set[Tuple[str, str]]],
    Dict[str, str],
    StockResolutionIndex,
]:
    component_id_to_symbol: Dict[str, str] = {}
    component_id_to_gene_symbol: Dict[str, str] = {}
    component_id_to_stock_candidate_details: Dict[str, Set[Tuple[str, str]]] = defaultdict(set)
    component_rows_by_fbst: Dict[str, List[Tuple[str, str]]] = defaultdict(list)
    fbst_to_fb_genotype: Dict[str, str] = {}
    label_to_gal4_component_ids: Dict[str, Set[str]] = defaultdict(set)
    label_to_fbst: Dict[str, str] = {}
    fbst_to_stock_meta: Dict[str, Dict[str, str]] = {}
    fbst_to_component_symbols: Dict[str, Dict[str, str]] = defaultdict(dict)
    gal4_component_records: List[Tuple[str, str, str, str]] = []
    gal4_component_records_by_key: Dict[str, List[Tuple[str, str, str, str]]] = defaultdict(list)

    fbst_col = derived_df["FBst"].tolist()
    stock_number_col = derived_df["stock_number"].tolist()
    collection_col = derived_df["collection"].tolist()
    cid_col = derived_df["derived_stock_component"].tolist()
    sym_col = derived_df["object_symbol"].tolist()
    gene_col = derived_df["GeneSymbol"].tolist()
    genotype_col = derived_df["FB_genotype"].tolist()

    for fbst, stock_number, collection_raw, cid, sym, gene_symbol, fallback_symbol in zip(
        fbst_col,
        stock_number_col,
        collection_col,
        cid_col,
        sym_col,
        gene_col,
        genotype_col,
    ):
        cid = clean_id(cid)
        sym = str(sym or "").strip()
        gene_symbol = str(gene_symbol or "").strip()
        fallback_symbol = str(fallback_symbol or "").strip()
        fbst = clean_id(fbst)
        stock_number = clean_id(stock_number)
        collection = normalize_collection_name(collection_raw)
        candidate_label = _format_stock_candidate_label(
            fbst,
            collection,
            stock_number,
        )
        if fbst and fallback_symbol and fbst not in fbst_to_fb_genotype:
            fbst_to_fb_genotype[fbst] = fallback_symbol
        elif fbst and fallback_symbol:
            existing = fbst_to_fb_genotype.get(fbst, "")
            if len(fallback_symbol) > len(existing):
                fbst_to_fb_genotype[fbst] = fallback_symbol
        if cid and sym:
            component_id_to_symbol[cid] = sym
        if cid and gene_symbol:
            component_id_to_gene_symbol[cid] = gene_symbol
        if cid and fbst and candidate_label:
            component_id_to_stock_candidate_details[cid].add((fbst, candidate_label))
            if candidate_label not in label_to_fbst:
                label_to_fbst[candidate_label] = fbst
            if _might_be_gal4_component(sym, gene_symbol) and _looks_like_gal4_symbol(sym, gene_symbol):
                label_to_gal4_component_ids[candidate_label].add(cid)
        if fbst:
            component_rows_by_fbst[fbst].append((sym or fallback_symbol, gene_symbol))
            if cid and sym:
                fbst_to_component_symbols[fbst][cid] = sym
            if fbst not in fbst_to_stock_meta:
                fbst_to_stock_meta[fbst] = {
                    "stock_number": stock_number,
                    "collection": collection,
                    "source_stock": _format_source_stock_label(collection, stock_number),
                    "stock_genotype": fallback_symbol,
                }
            elif fallback_symbol and len(fallback_symbol) > len(
                fbst_to_stock_meta[fbst].get("stock_genotype", "")
            ):
                fbst_to_stock_meta[fbst]["stock_genotype"] = fallback_symbol
            if cid and _might_be_gal4_component(sym, gene_symbol) and _looks_like_gal4_symbol(sym, gene_symbol):
                record = (fbst, cid, sym, fallback_symbol)
                gal4_component_records.append(record)
                for key in _gal4_search_keys(sym):
                    gal4_component_records_by_key[key].append(record)

    gal4_only_fbst = {
        fbst
        for fbst, component_rows in component_rows_by_fbst.items()
        if component_rows
        and all(
            _looks_like_gal4_symbol(symbol, gene_symbol)
            for symbol, gene_symbol in component_rows
        )
    }
    driver_stock_index = _build_gal4_driver_stock_candidate_index(
        component_id_to_symbol,
        component_id_to_gene_symbol,
        component_id_to_stock_candidate_details,
    )
    stock_candidate_label_to_driver_genotype = _build_stock_candidate_label_to_driver_genotype(
        gal4_only_fbst,
        component_id_to_stock_candidate_details,
        component_rows_by_fbst,
        fbst_to_fb_genotype,
    )
    resolution_index = StockResolutionIndex(
        label_to_fbst=label_to_fbst,
        label_to_gal4_component_ids=dict(label_to_gal4_component_ids),
        fbst_to_stock_meta=fbst_to_stock_meta,
        fbst_to_component_symbols=dict(fbst_to_component_symbols),
        gal4_component_records=gal4_component_records,
        gal4_component_records_by_key=dict(gal4_component_records_by_key),
    )
    return (
        component_id_to_symbol,
        component_id_to_gene_symbol,
        component_id_to_stock_candidate_details,
        component_rows_by_fbst,
        gal4_only_fbst,
        stock_candidate_label_to_driver_genotype,
        driver_stock_index,
        fbst_to_fb_genotype,
        resolution_index,
    )


def _resolve_gal4_stock_candidates_mixed(
    partner_symbol: str,
    driver_stock_index: Dict[str, Set[Tuple[str, str]]],
) -> Set[str]:
    """Resolve stock labels from a driver symbol without gal4-only filtering."""
    matched_details: Set[Tuple[str, str]] = set()
    for key in _input_gal4_lookup_keys(partner_symbol):
        matched_details.update(driver_stock_index.get(key, set()))

    return {label for _fbst, label in matched_details if label}


def _shorthand_search_tokens(input_symbol: str) -> List[str]:
    """Return search tokens for BDSC-style shorthand driver names."""
    sym = str(input_symbol or "").strip()
    if not sym:
        return []

    tokens: Set[str] = {sym}
    tokens.add(sym.replace(".", "-"))
    tokens.add(sym.replace("-", "."))

    for suffix in ("PU", "PE", "PK", "PR", "PD", "PO"):
        dotted = f".{suffix}"
        if sym.endswith(dotted):
            base = sym[: -len(dotted)]
            if base and len(base) >= 5:
                tokens.add(base)
                tokens.add(f"{base}-GAL4")

    if "gal4" not in sym.lower() and not sym.upper().endswith("-GAL4"):
        tokens.add(f"{sym}-GAL4")

    return sorted({token for token in tokens if len(token) >= 3}, key=len, reverse=True)


def _generic_shorthand_driver_base(input_symbol: Any) -> str:
    """Extract the promoter name from FlyBase's generic .P? driver shorthand."""
    text = str(input_symbol or "").strip()
    match = re.fullmatch(r"(.+)\.(PU|PE|PK|PR|PD|PO)", text, flags=re.IGNORECASE)
    if not match:
        return ""
    base = match.group(1).strip()
    return base if len(_compact_symbol_key(base)) >= 3 else ""


def _gal4_promoter_keys(symbol: Any) -> Set[str]:
    """Extract promoter/founder names from common GAL4 construct encodings."""
    text = str(symbol or "").strip()
    keys: Set[str] = set()

    def add(value: Any) -> None:
        token = re.split(r"[.\[]", str(value or "").strip(), maxsplit=1)[0]
        key = _compact_symbol_key(token)
        if len(key) >= 3 and key not in {"gal4", "gawb", "scer"}:
            keys.add(key)

    for value in re.findall(
        r"([A-Za-z][A-Za-z0-9]*)[-_.]GAL4(?:[-_.:{\[]|$)",
        text,
        flags=re.IGNORECASE,
    ):
        add(value)
    for value in re.findall(
        r"GAL4[-_.]([A-Za-z][A-Za-z0-9]*)",
        text,
        flags=re.IGNORECASE,
    ):
        add(value)
    for value in re.findall(
        r"\}(?:[A-Za-z]*\\)?([A-Za-z][A-Za-z0-9]*)\[",
        text,
    ):
        add(value)
    for value in re.findall(
        r"(?:Scer\\)?GAL4\[([A-Za-z][A-Za-z0-9]*)",
        text,
        flags=re.IGNORECASE,
    ):
        add(value)
    return keys


def _stock_has_non_driver_transgene(
    fbst: str,
    component_rows_by_fbst: Dict[str, List[Tuple[str, str]]],
) -> bool:
    """Return True when a candidate driver stock also carries another transgene."""
    for symbol, gene_symbol in component_rows_by_fbst.get(fbst, []):
        symbol_text = str(symbol or "").strip()
        if not symbol_text or _looks_like_gal4_symbol(symbol_text, gene_symbol):
            continue
        lowered = symbol_text.lower()
        if any(
            marker in lowered
            for marker in (
                "uas-",
                "uas_",
                "uas::",
                "lexaop",
                "quas",
                "gal80",
                "qf",
                "lexa",
            )
        ):
            return True
    return False


def _symbol_matches_input(
    input_symbol: str,
    candidate_text: str,
    search_tokens: List[str],
    *,
    allow_substring: bool = True,
) -> bool:
    text = str(candidate_text or "").strip()
    if not text:
        return False
    if _gal4_driver_symbols_match(input_symbol, text):
        return True
    if not allow_substring:
        return False
    lowered = text.lower()
    min_token_len = max(4, len(str(input_symbol or "").strip()) - 1)
    for token in search_tokens:
        if len(token) < min_token_len:
            continue
        if token.lower() in lowered:
            return True
    return False


def _resolve_gal4_via_derived_text_search(
    input_symbol: str,
    resolution_index: StockResolutionIndex,
    component_id_to_symbol: Dict[str, str],
    component_id_to_gene_symbol: Dict[str, str],
) -> Dict[str, Set[str]]:
    """Return FBst -> matched GAL4 component IDs via shorthand/text matching."""
    search_tokens = _shorthand_search_tokens(input_symbol)
    if not search_tokens:
        return {}

    matched_by_fbst: Dict[str, Set[str]] = defaultdict(set)
    candidate_records: Dict[Tuple[str, str], Tuple[str, str, str, str]] = {}
    for key in _input_gal4_lookup_keys(input_symbol):
        for record in resolution_index.gal4_component_records_by_key.get(key, []):
            candidate_records[(record[0], record[1])] = record

    allow_split = _input_requests_split_gal4(input_symbol)
    for fbst, cid, symbol, _genotype_text in candidate_records.values():
        if _is_split_gal4_component(symbol) and not allow_split:
            continue
        if not _symbol_matches_input(
            input_symbol,
            symbol,
            search_tokens,
            allow_substring=True,
        ):
            continue
        matched_by_fbst[fbst].add(cid)
        component_id_to_symbol.setdefault(cid, symbol)
        if cid in resolution_index.fbst_to_component_symbols.get(fbst, {}):
            gene_symbol = component_id_to_gene_symbol.get(cid, "")
        else:
            gene_symbol = ""
        if gene_symbol:
            component_id_to_gene_symbol.setdefault(cid, gene_symbol)
    return matched_by_fbst


def _matches_from_fbst_map(
    input_symbol: str,
    fbst_to_component_ids: Dict[str, Set[str]],
    resolution_index: StockResolutionIndex,
    component_rows_by_fbst: Dict[str, List[Tuple[str, str]]],
    fbst_to_fb_genotype: Dict[str, str],
    stock_candidate_label_to_driver_genotype: Dict[str, str],
    *,
    resolution_method: str = "Exact name or FlyBase alias",
    match_confidence: str = "Exact/authoritative",
) -> List[Gal4StockMatch]:
    matches: List[Gal4StockMatch] = []
    seen_source_stocks: Set[str] = set()

    for fbst, component_ids in sorted(fbst_to_component_ids.items()):
        if not fbst or not component_ids:
            continue
        stock_meta = resolution_index.fbst_to_stock_meta.get(fbst)
        if not stock_meta:
            continue
        stock_number = stock_meta.get("stock_number", "")
        collection = stock_meta.get("collection", "")
        source_stock = stock_meta.get("source_stock", "")
        if not source_stock or source_stock in seen_source_stocks:
            continue
        seen_source_stocks.add(source_stock)

        component_symbols = {
            resolution_index.fbst_to_component_symbols.get(fbst, {}).get(cid, "")
            for cid in component_ids
        }
        component_symbols = {sym for sym in component_symbols if sym}

        label = _format_stock_candidate_label(fbst, collection, stock_number)
        stock_genotype = (
            fbst_to_fb_genotype.get(fbst, "")
            or stock_meta.get("stock_genotype", "")
        )
        driver_genotype = stock_candidate_label_to_driver_genotype.get(label, "")
        if not driver_genotype:
            driver_genotype = _primary_gal4_component_symbol_for_fbst(
                fbst,
                component_rows_by_fbst,
                fbst_to_fb_genotype,
            )

        matches.append(
            Gal4StockMatch(
                input_symbol=input_symbol,
                fbst=(
                    ""
                    if fbst.startswith(BDSC_INTERNAL_FBST_PREFIX)
                    else fbst
                ),
                stock_number=stock_number,
                collection=collection,
                source_stock=source_stock,
                stock_genotype=stock_genotype,
                component_ids=set(component_ids),
                component_symbols=component_symbols,
                driver_genotype=driver_genotype,
                resolution_method=resolution_method,
                match_confidence=match_confidence,
            )
        )
    return matches


def _fbst_map_from_candidate_labels(
    candidate_labels: Iterable[str],
    resolution_index: StockResolutionIndex,
) -> Dict[str, Set[str]]:
    fbst_to_component_ids: Dict[str, Set[str]] = defaultdict(set)
    for label in candidate_labels:
        fbst = resolution_index.label_to_fbst.get(label, "")
        if not fbst:
            continue
        for cid in resolution_index.label_to_gal4_component_ids.get(label, set()):
            fbst_to_component_ids[fbst].add(cid)
    return fbst_to_component_ids


def _fbst_map_from_construct_aliases(
    input_symbol: str,
    construct_alias_fbti_index: Dict[str, Set[str]],
    component_id_to_stock_candidate_details: Dict[str, Set[Tuple[str, str]]],
    *,
    allowed_fbst: Optional[Set[str]] = None,
) -> Dict[str, Set[str]]:
    """Resolve input aliases through FlyBase construct IDs to stock insertions."""
    linked_fbtis: Set[str] = set()
    for key in _informative_gal4_alias_keys(input_symbol):
        linked_fbtis.update(construct_alias_fbti_index.get(key, set()))

    fbst_to_component_ids: Dict[str, Set[str]] = defaultdict(set)
    for fbti_id in linked_fbtis:
        for fbst, _label in component_id_to_stock_candidate_details.get(
            fbti_id,
            set(),
        ):
            if allowed_fbst is None or fbst in allowed_fbst:
                fbst_to_component_ids[fbst].add(fbti_id)
    return fbst_to_component_ids


def _merge_fbst_component_maps(
    target: Dict[str, Set[str]],
    source: Dict[str, Set[str]],
) -> None:
    for fbst, component_ids in source.items():
        target[fbst].update(component_ids)


def _resolve_generic_promoter_driver_fallback(
    input_symbol: str,
    resolution_index: StockResolutionIndex,
    component_rows_by_fbst: Dict[str, List[Tuple[str, str]]],
) -> Dict[str, Set[str]]:
    """Choose the simplest current stock for a generic FlyBase .P? driver."""
    base = _generic_shorthand_driver_base(input_symbol)
    base_key = _compact_symbol_key(base)
    if not base_key:
        return {}

    candidates: List[Tuple[Tuple[int, int, int, int], str, str]] = []
    allow_split = _input_requests_split_gal4(input_symbol)
    for fbst, component_id, symbol, _genotype in resolution_index.gal4_component_records:
        if base_key not in _gal4_promoter_keys(symbol):
            continue
        if _is_split_gal4_component(symbol) and not allow_split:
            continue
        lowered = str(symbol or "").lower()
        explicit_promoter_construct = bool(
            re.search(
                rf"(?:{re.escape(base.lower())}[-_.]gal4|gal4[-_.]{re.escape(base.lower())})",
                lowered,
            )
        )
        stock_meta = resolution_index.fbst_to_stock_meta.get(fbst, {})
        stock_number = clean_id(stock_meta.get("stock_number", ""))
        numeric_stock_number = (
            int(stock_number)
            if stock_number.isdigit()
            else 10**12
        )
        score = (
            int(_stock_has_non_driver_transgene(fbst, component_rows_by_fbst)),
            int(not explicit_promoter_construct),
            numeric_stock_number,
            len(component_rows_by_fbst.get(fbst, [])),
        )
        candidates.append((score, fbst, component_id))

    if not candidates:
        return {}
    best_score, best_fbst, best_component_id = min(candidates)
    _ = best_score
    return {best_fbst: {best_component_id}}


def _should_merge_gmr_alias_text_match(
    input_symbol: str,
    fbst: str,
    component_ids: Set[str],
    resolution_index: StockResolutionIndex,
) -> bool:
    """Allow GMRxx-GAL4 stock symbols to complement Rxx-GAL4 inputs."""
    input_keys = _input_gal4_lookup_keys(input_symbol)
    if not any(re.fullmatch(r"r\d+[a-z]\d+(gal4)?", key) for key in input_keys):
        return False
    component_symbols = resolution_index.fbst_to_component_symbols.get(fbst, {})
    for cid in component_ids:
        symbol = component_symbols.get(cid, "")
        if _is_split_gal4_component(symbol):
            continue
        symbol_keys = _gal4_search_keys(symbol)
        if "gmr" in str(symbol or "").lower() and input_keys & symbol_keys:
            return True
    return False


def _resolve_gal4_symbol_to_stocks(
    input_symbol: str,
    gal4_only_fbst: Set[str],
    driver_stock_index: Dict[str, Set[Tuple[str, str]]],
    component_id_to_symbol: Dict[str, str],
    component_id_to_gene_symbol: Dict[str, str],
    component_id_to_stock_candidate_details: Dict[str, Set[Tuple[str, str]]],
    component_rows_by_fbst: Dict[str, List[Tuple[str, str]]],
    stock_candidate_label_to_driver_genotype: Dict[str, str],
    fbst_to_fb_genotype: Dict[str, str],
    resolution_index: StockResolutionIndex,
    construct_alias_fbti_index: Dict[str, Set[str]],
) -> List[Gal4StockMatch]:
    candidate_labels = sorted(
        _resolve_gal4_stock_candidates_by_driver_symbol(
            input_symbol,
            driver_stock_index,
            gal4_only_fbst,
        )
    )
    fbst_to_component_ids = _fbst_map_from_candidate_labels(
        candidate_labels,
        resolution_index,
    )
    _merge_fbst_component_maps(
        fbst_to_component_ids,
        _fbst_map_from_construct_aliases(
            input_symbol,
            construct_alias_fbti_index,
            component_id_to_stock_candidate_details,
            allowed_fbst=gal4_only_fbst,
        ),
    )
    if fbst_to_component_ids:
        for fbst, component_ids in _resolve_gal4_via_derived_text_search(
            input_symbol,
            resolution_index,
            component_id_to_symbol,
            component_id_to_gene_symbol,
        ).items():
            if _should_merge_gmr_alias_text_match(
                input_symbol,
                fbst,
                component_ids,
                resolution_index,
            ):
                fbst_to_component_ids[fbst].update(component_ids)
        matches = _matches_from_fbst_map(
            input_symbol,
            fbst_to_component_ids,
            resolution_index,
            component_rows_by_fbst,
            fbst_to_fb_genotype,
            stock_candidate_label_to_driver_genotype,
        )
        if matches:
            return matches

    mixed_candidate_labels = sorted(
        _resolve_gal4_stock_candidates_mixed(input_symbol, driver_stock_index)
    )
    mixed_fbst_to_component_ids = _fbst_map_from_candidate_labels(
        mixed_candidate_labels,
        resolution_index,
    )
    _merge_fbst_component_maps(
        mixed_fbst_to_component_ids,
        _fbst_map_from_construct_aliases(
            input_symbol,
            construct_alias_fbti_index,
            component_id_to_stock_candidate_details,
        ),
    )
    if len(mixed_fbst_to_component_ids) <= 5:
        if mixed_fbst_to_component_ids:
            matches = _matches_from_fbst_map(
                input_symbol,
                mixed_fbst_to_component_ids,
                resolution_index,
                component_rows_by_fbst,
                fbst_to_fb_genotype,
                stock_candidate_label_to_driver_genotype,
            )
            if matches:
                return matches

    text_matches = _resolve_gal4_via_derived_text_search(
        input_symbol,
        resolution_index,
        component_id_to_symbol,
        component_id_to_gene_symbol,
    )
    matches = _matches_from_fbst_map(
        input_symbol,
        text_matches,
        resolution_index,
        component_rows_by_fbst,
        fbst_to_fb_genotype,
        stock_candidate_label_to_driver_genotype,
    )
    if matches:
        return matches

    promoter_fallback = _resolve_generic_promoter_driver_fallback(
        input_symbol,
        resolution_index,
        component_rows_by_fbst,
    )
    matches = _matches_from_fbst_map(
        input_symbol,
        promoter_fallback,
        resolution_index,
        component_rows_by_fbst,
        fbst_to_fb_genotype,
        stock_candidate_label_to_driver_genotype,
        resolution_method="BDSC promoter-family substitution",
        match_confidence="Candidate substitution—not the exact published construct",
    )
    if matches:
        return matches

    return [
        Gal4StockMatch(
            input_symbol=input_symbol,
            fbst="",
            stock_number="",
            collection="",
            source_stock="",
            resolution_method="Unresolved",
            match_confidence="No orderable stock candidate found",
        )
    ]


def _load_component_links_for_seeds(
    seed_component_ids: Set[str],
    flybase_data_path: Path,
) -> Tuple[Dict[str, Set[str]], Dict[str, Set[str]], Dict[str, Set[str]]]:
    """Load only the FBal↔FBtp↔FBti links needed for the resolved seed components."""
    fbal_to_fbtps: Dict[str, Set[str]] = defaultdict(set)
    fbal_to_fbtis: Dict[str, Set[str]] = defaultdict(set)
    fbtp_to_fbtis: Dict[str, Set[str]] = defaultdict(set)
    if not seed_component_ids:
        return fbal_to_fbtps, fbal_to_fbtis, fbtp_to_fbtis

    alleles_dir = flybase_data_path / "alleles_and_stocks"
    constructs_dir = flybase_data_path / "transgenic_constructs"
    insertions_dir = flybase_data_path / "transgenic_insertions"

    seed_fbtis = {cid for cid in seed_component_ids if cid.startswith("FBti")}
    seed_fbtps = {cid for cid in seed_component_ids if cid.startswith("FBtp")}
    seed_fbals = {cid for cid in seed_component_ids if cid.startswith("FBal")}

    fbtp_to_fbti_path = insertions_dir / "fbtp_to_fbti.csv"
    if fbtp_to_fbti_path.exists():
        mapping_df = pd.read_csv(fbtp_to_fbti_path, dtype=str).fillna("")
        if seed_fbtps or seed_fbtis:
            mapping_df = mapping_df[
                mapping_df["FBtp"].astype(str).isin(seed_fbtps)
                | mapping_df["FBti"].astype(str).isin(seed_fbtis)
            ]
        for fbtp_raw, fbti_raw in zip(mapping_df["FBtp"], mapping_df["FBti"]):
            fbtp_id = clean_id(fbtp_raw)
            fbti_id = clean_id(fbti_raw)
            if fbtp_id and fbti_id:
                fbtp_to_fbtis[fbtp_id].add(fbti_id)
                seed_fbtps.add(fbtp_id)
                seed_fbtis.add(fbti_id)

    if seed_fbtis:
        try:
            insertion_df = load_flybase_tsv(
                find_latest_tsv(alleles_dir, "dmel_classical_and_insertion_allele_descriptions")
            )
            allele_col = "Allele (id)" if "Allele (id)" in insertion_df.columns else None
            fbti_col = "Insertion (id)" if "Insertion (id)" in insertion_df.columns else None
            if allele_col and fbti_col:
                insertion_df = insertion_df[insertion_df[fbti_col].astype(str).isin(seed_fbtis)]
                for allele_raw, fbti_raw in zip(insertion_df[allele_col], insertion_df[fbti_col]):
                    fbal_id = clean_id(allele_raw)
                    fbti_id = clean_id(fbti_raw)
                    if fbal_id and fbti_id:
                        fbal_to_fbtis[fbal_id].add(fbti_id)
                        seed_fbals.add(fbal_id)
        except FileNotFoundError:
            pass

    if seed_fbals:
        try:
            construct_df = load_flybase_tsv(
                find_latest_tsv(constructs_dir, "transgenic_construct_descriptions")
            )
            fbal_col = (
                "Component Allele (id)"
                if "Component Allele (id)" in construct_df.columns
                else None
            )
            fbtp_col = (
                "Transgenic Construct (id)"
                if "Transgenic Construct (id)" in construct_df.columns
                else None
            )
            if fbal_col is None or fbtp_col is None:
                raw_cols = list(construct_df.columns)
                if len(raw_cols) >= 4:
                    fbal_col, fbtp_col = raw_cols[1], raw_cols[3]
            if fbal_col and fbtp_col:
                construct_df = construct_df[construct_df[fbal_col].astype(str).isin(seed_fbals)]
                for fbal_raw, fbtp_raw in zip(construct_df[fbal_col], construct_df[fbtp_col]):
                    fbal_id = clean_id(fbal_raw)
                    if not fbal_id:
                        continue
                    for token in str(fbtp_raw or "").split("|"):
                        fbtp_id = clean_id(token)
                        if fbtp_id:
                            fbal_to_fbtps[fbal_id].add(fbtp_id)
                            seed_fbtps.add(fbtp_id)
        except FileNotFoundError:
            pass

    if seed_fbtps and fbtp_to_fbti_path.exists():
        mapping_df = pd.read_csv(fbtp_to_fbti_path, dtype=str).fillna("")
        mapping_df = mapping_df[mapping_df["FBtp"].astype(str).isin(seed_fbtps)]
        for fbtp_raw, fbti_raw in zip(mapping_df["FBtp"], mapping_df["FBti"]):
            fbtp_id = clean_id(fbtp_raw)
            fbti_id = clean_id(fbti_raw)
            if fbtp_id and fbti_id:
                fbtp_to_fbtis[fbtp_id].add(fbti_id)

    return fbal_to_fbtps, fbal_to_fbtis, fbtp_to_fbtis


def _fbals_for_seed_components(
    seed_component_ids: Set[str],
    fbal_to_fbtps: Dict[str, Set[str]],
    fbal_to_fbtis: Dict[str, Set[str]],
    fbtp_to_fbtis: Dict[str, Set[str]],
) -> Set[str]:
    """Return FBal IDs linked to seed FBti/FBtp component IDs."""
    fbti_to_fbal: Dict[str, Set[str]] = defaultdict(set)
    fbtp_to_fbal: Dict[str, Set[str]] = defaultdict(set)
    for fbal_id, fbti_ids in fbal_to_fbtis.items():
        for fbti_id in fbti_ids:
            fbti_to_fbal[fbti_id].add(fbal_id)
    for fbal_id, fbtp_ids in fbal_to_fbtps.items():
        for fbtp_id in fbtp_ids:
            fbtp_to_fbal[fbtp_id].add(fbal_id)

    linked_fbals: Set[str] = set()
    for cid in seed_component_ids:
        if cid.startswith("FBal"):
            linked_fbals.add(cid)
        elif cid.startswith("FBti"):
            linked_fbals.update(fbti_to_fbal.get(cid, set()))
        elif cid.startswith("FBtp"):
            linked_fbals.update(fbtp_to_fbal.get(cid, set()))
    return linked_fbals


def _component_type_for_id(component_id: str) -> str:
    if component_id.startswith("FBal"):
        return "FBal"
    if component_id.startswith("FBtp"):
        return "FBtp"
    if component_id.startswith("FBti"):
        return "FBti"
    return ""


def _build_synthetic_inputs(
    matches: Iterable[Gal4StockMatch],
    *,
    flybase_data_path: Path,
    component_id_to_symbol: Dict[str, str],
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    stock_rows: List[Dict[str, Any]] = []
    reagent_rows: List[Dict[str, str]] = []
    seen_fbst: Set[str] = set()
    seen_components: Set[str] = set()

    seed_component_ids: Set[str] = set()
    for match in matches:
        seed_component_ids.update(match.component_ids)

    fbal_to_fbtps, fbal_to_fbtis, fbtp_to_fbtis = _load_component_links_for_seeds(
        seed_component_ids,
        flybase_data_path,
    )

    for match in matches:
        if match.fbst and match.fbst not in seen_fbst:
            seen_fbst.add(match.fbst)
            stock_rows.append(
                {
                    "FBst": match.fbst,
                    "stock_number": match.stock_number,
                    "collection": match.collection,
                    "relevant_gene_symbols": unique_join(match.component_symbols),
                    "relevant_component_ids": unique_join(sorted(match.component_ids)),
                    "Balancers": "",
                    "matched_component_types": "",
                    "allele_class_terms": "",
                    "transgenic_product_class_terms": "",
                    "RNAi Type": "",
                    **{column: column == "GAL4" for column in REAGENT_BUCKET_COLUMNS},
                }
            )

        linked_fbals = _fbals_for_seed_components(
            match.component_ids,
            fbal_to_fbtps,
            fbal_to_fbtis,
            fbtp_to_fbtis,
        )
        if not linked_fbals and match.component_ids:
            linked_fbals = set(match.component_ids)

        for fbal_id in sorted(linked_fbals):
            component_group: Set[str] = {fbal_id}
            component_group.update(fbal_to_fbtis.get(fbal_id, set()))
            for fbtp_id in fbal_to_fbtps.get(fbal_id, set()):
                component_group.add(fbtp_id)
                component_group.update(fbtp_to_fbtis.get(fbtp_id, set()))
            component_group.update(match.component_ids)

            for cid in sorted(component_group):
                if not cid or cid in seen_components:
                    continue
                seen_components.add(cid)
                symbol = component_id_to_symbol.get(cid, "") or unique_join(match.component_symbols)
                if not symbol and cid in match.component_ids:
                    symbol = unique_join(match.component_symbols)
                reagent_rows.append(
                    {
                        "input_flybase_gene_id": "",
                        "input_gene_symbol": match.input_symbol,
                        "component_id": cid,
                        "component_type": _component_type_for_id(cid),
                        "component_symbol": symbol or match.input_symbol,
                        "source_allele_id": fbal_id,
                        "source_allele_symbol": component_id_to_symbol.get(fbal_id, match.input_symbol),
                        "allele_class_terms": "",
                        "transgenic_product_class_terms": "",
                        "rnai_type": "",
                        "match_provenance": "gal4_driver_export",
                    }
                )

    stocks_df = pd.DataFrame(stock_rows)
    reagent_index_df = pd.DataFrame(reagent_rows)
    return stocks_df, reagent_index_df


def _build_component_to_source_stocks(
    matches: Iterable[Gal4StockMatch],
    reagent_index_df: pd.DataFrame,
) -> Dict[str, Set[str]]:
    """Map FlyBase component IDs to resolved stock labels."""
    symbol_to_source_stocks: Dict[str, Set[str]] = defaultdict(set)
    for match in matches:
        if match.source_stock:
            symbol_to_source_stocks[match.input_symbol].add(match.source_stock)

    component_to_source_stocks: Dict[str, Set[str]] = defaultdict(set)
    for match in matches:
        if not match.source_stock:
            continue
        for cid in match.component_ids:
            component_to_source_stocks[cid].add(match.source_stock)

    if len(reagent_index_df) > 0 and "component_id" in reagent_index_df.columns:
        for component_id, input_symbol in zip(
            reagent_index_df["component_id"].tolist(),
            reagent_index_df["input_gene_symbol"].tolist(),
        ):
            cid = clean_id(component_id)
            sym = str(input_symbol or "").strip()
            if not cid or not sym:
                continue
            component_to_source_stocks[cid].update(symbol_to_source_stocks.get(sym, set()))
    return component_to_source_stocks


def _filter_phenotype_df_for_components(
    phenotype_df: pd.DataFrame,
    component_ids: Set[str],
) -> pd.DataFrame:
    """Keep only phenotype rows whose genotype references at least one component ID."""
    if len(phenotype_df) == 0 or not component_ids:
        return pd.DataFrame()

    genotype_ids = phenotype_df.get("genotype_FBids", pd.Series("", index=phenotype_df.index))
    genotype_ids = genotype_ids.fillna("").astype(str)
    id_list = sorted(component_ids)
    mask = pd.Series(False, index=phenotype_df.index)
    chunk_size = 40
    for start in range(0, len(id_list), chunk_size):
        chunk = id_list[start : start + chunk_size]
        pattern = "|".join(re.escape(component_id) for component_id in chunk)
        mask |= genotype_ids.str.contains(pattern, na=False, regex=True)
    return phenotype_df[mask].copy()


def _load_fbrf_reference_maps(
    flybase_data_path: Path,
) -> Tuple[Dict[str, str], Dict[str, str], Dict[str, str]]:
    refs_dir = flybase_data_path / "references"
    fbrf_to_pmid: Dict[str, str] = {}
    fbrf_to_pmcid: Dict[str, str] = {}
    fbrf_to_miniref: Dict[str, str] = {}
    try:
        refs_tsv = load_flybase_tsv(find_latest_tsv(refs_dir, "fbrf_pmid_pmcid_doi"))
        for fbrf, pmid, pmcid, miniref in zip(
            refs_tsv.get("FBrf", pd.Series("", index=refs_tsv.index)).tolist(),
            refs_tsv.get("PMID", pd.Series("", index=refs_tsv.index)).tolist(),
            refs_tsv.get("PMCID", pd.Series("", index=refs_tsv.index)).tolist(),
            refs_tsv.get("miniref", pd.Series("", index=refs_tsv.index)).tolist(),
        ):
            fbrf = clean_id(fbrf)
            if not fbrf:
                continue
            fbrf_to_pmid[fbrf] = clean_id(pmid)
            fbrf_to_pmcid[fbrf] = clean_id(pmcid)
            fbrf_to_miniref[fbrf] = str(miniref or "").strip()
    except FileNotFoundError:
        pass
    return fbrf_to_pmid, fbrf_to_pmcid, fbrf_to_miniref


def _reference_metadata_for_fbrf(
    fbrf: str,
    fbrf_to_pmid: Dict[str, str],
    fbrf_to_pmcid: Dict[str, str],
    fbrf_to_miniref: Dict[str, str],
    pubmed_cache: Optional[Dict[str, dict]],
) -> Dict[str, str]:
    pmid = fbrf_to_pmid.get(fbrf, "")
    pmcid = fbrf_to_pmcid.get(fbrf, "")
    miniref = fbrf_to_miniref.get(fbrf, "")
    title = ""
    authors = ""
    journal = ""
    year = ""
    if pmid and pubmed_cache and pmid in pubmed_cache:
        meta = pubmed_cache[pmid]
        title = str(meta.get("title", "") or "").strip()
        authors = str(meta.get("authors", "") or "").strip()
        journal = str(meta.get("journal", "") or "").strip()
        year = str(meta.get("year", "") or "").strip()
    display = title or miniref or pmid or pmcid or fbrf
    return {
        "pmid": pmid,
        "pmcid": pmcid,
        "display": display,
        "authors": authors,
        "journal": journal,
        "year": year,
    }


def _build_gal4_phenotype_evidence(
    matches: List[Gal4StockMatch],
    reagent_index_df: pd.DataFrame,
    flybase_data_path: Path,
    pubmed_cache_path: Optional[Path],
    *,
    fetch_pubmed: bool = True,
    verbose: bool = False,
) -> pd.DataFrame:
    """Build a slim phenotype sheet scoped to the resolved GAL4 reagents only."""
    component_to_source_stocks = _build_component_to_source_stocks(matches, reagent_index_df)
    relevant_ids = {cid for cid in component_to_source_stocks if cid}
    if not relevant_ids:
        return pd.DataFrame()

    alleles_dir = flybase_data_path / "alleles_and_stocks"
    try:
        phenotype_df = load_flybase_tsv(find_latest_tsv(alleles_dir, "genotype_phenotype_data"))
    except FileNotFoundError:
        if verbose:
            print(f"    Warning: genotype_phenotype_data not found under {alleles_dir}")
        return pd.DataFrame()

    phenotype_df = _filter_phenotype_df_for_components(phenotype_df, relevant_ids)
    if len(phenotype_df) == 0:
        return pd.DataFrame()

    fbrf_to_pmid, fbrf_to_pmcid, fbrf_to_miniref = _load_fbrf_reference_maps(flybase_data_path)
    pubmed_cache: Optional[Dict[str, dict]] = None
    pubmed_cache_obj = None
    if pubmed_cache_path:
        try:
            from fl_ai_reagent_stocker.integrations.pubmed import PubMedCache

            pubmed_cache_obj = PubMedCache(pubmed_cache_path)
            pubmed_cache = pubmed_cache_obj.load()
        except Exception:
            pubmed_cache_obj = None
            pubmed_cache = None

    fbrfs: Set[str] = set()
    for raw_reference in phenotype_df.get("reference", pd.Series(dtype=str)).fillna("").astype(str):
        for value in str(raw_reference).split("|"):
            fbrf = clean_id(value)
            if fbrf:
                fbrfs.add(fbrf)
    if fetch_pubmed:
        pmids_to_fetch = sorted(
            {
                pmid
                for fbrf in fbrfs
                for pmid in [fbrf_to_pmid.get(fbrf, "")]
                if pmid and (not pubmed_cache or pmid not in pubmed_cache)
            }
        )
        if pmids_to_fetch:
            try:
                from fl_ai_reagent_stocker.integrations.pubmed import PubMedClient

                client = PubMedClient(cache=pubmed_cache_obj)
                fetched = client.fetch_metadata(pmids_to_fetch)
                pubmed_cache = pubmed_cache or {}
                pubmed_cache.update(fetched)
            except Exception as exc:
                if verbose:
                    print(f"    Warning: PubMed fetch failed: {exc}")

    rows: List[Dict[str, str]] = []
    phenotype_names = phenotype_df.get("phenotype_name", pd.Series("", index=phenotype_df.index)).fillna("").astype(str).tolist()
    qualifier_names = phenotype_df.get("qualifier_names", pd.Series("", index=phenotype_df.index)).fillna("").astype(str).tolist()
    genotype_symbols = phenotype_df.get("genotype_symbols", pd.Series("", index=phenotype_df.index)).fillna("").astype(str).tolist()
    genotype_fbids = phenotype_df.get("genotype_FBids", pd.Series("", index=phenotype_df.index)).fillna("").astype(str).tolist()
    references = phenotype_df.get("reference", pd.Series("", index=phenotype_df.index)).fillna("").astype(str).tolist()

    for phenotype_name, qualifier_name, genotype_label, genotype_fbids_text, reference_value in zip(
        phenotype_names,
        qualifier_names,
        genotype_symbols,
        genotype_fbids,
        references,
    ):
        component_ids = _extract_flybase_ids(genotype_fbids_text)
        matched_ids = [cid for cid in component_ids if cid in relevant_ids]
        if not matched_ids:
            continue

        source_stocks: Set[str] = set()
        for cid in matched_ids:
            source_stocks.update(component_to_source_stocks.get(cid, set()))
        if not source_stocks:
            continue

        raw_fbrfs = [clean_id(value) for value in reference_value.split("|") if clean_id(value)]
        refs_to_emit = raw_fbrfs if raw_fbrfs else [""]
        for fbrf in refs_to_emit:
            ref_details = (
                _reference_metadata_for_fbrf(
                    fbrf,
                    fbrf_to_pmid,
                    fbrf_to_pmcid,
                    fbrf_to_miniref,
                    pubmed_cache,
                )
                if fbrf
                else {
                    "pmid": "",
                    "pmcid": "",
                    "display": "",
                    "authors": "",
                    "journal": "",
                    "year": "",
                }
            )
            for source_stock in sorted(source_stocks):
                rows.append(
                    {
                        SOURCE_STOCK_COLUMN: source_stock,
                        "Phenotype": phenotype_name,
                        "Qualifier": qualifier_name,
                        "Genotype": genotype_label,
                        "PMID": ref_details["pmid"],
                        "PMCID": ref_details["pmcid"],
                        "Reference": ref_details["display"],
                        "Authors": ref_details["authors"],
                        "Journal": ref_details["journal"],
                        "Year of Publication": ref_details["year"],
                    }
                )

    if verbose:
        print(f"    Filtered to {len(phenotype_df)} phenotype row(s) for resolved GAL4 components")
    return pd.DataFrame(rows)


def _reference_key(pmid: Any, pmcid: Any) -> str:
    pmid_text = clean_id(pmid)
    pmcid_text = clean_id(pmcid)
    if pmid_text:
        return f"PMID{pmid_text}"
    if pmcid_text:
        normalized = re.sub(r"^PMCID[: ]?", "", pmcid_text, flags=re.IGNORECASE)
        normalized = re.sub(r"^PMC", "", normalized, flags=re.IGNORECASE).strip()
        return f"PMCID{normalized}" if normalized else ""
    return ""


def _aggregate_reference_field(values: Iterable[str]) -> str:
    deduped: List[str] = []
    seen: Set[str] = set()
    for raw in values:
        text = str(raw or "").strip()
        if not text or text == "-" or text in seen:
            continue
        seen.add(text)
        deduped.append(text)
    return "; ".join(deduped) if deduped else "-"


def _format_referenced_phenotype(ref_id: str, phenotype: str) -> str:
    phenotype_text = str(phenotype or "").strip()
    if not phenotype_text:
        return ""
    reference_text = str(ref_id or "").strip() or "-"
    return f"{{{reference_text}, {phenotype_text}}}"


def _aggregate_referenced_phenotypes(values: Iterable[str]) -> str:
    deduped: List[str] = []
    seen: Set[str] = set()
    for raw in values:
        text = str(raw or "").strip()
        if not text or text == "-" or text in seen:
            continue
        seen.add(text)
        deduped.append(text)
    return ",\n".join(deduped) if deduped else "-"


def _collection_output_for_match(match: Gal4StockMatch) -> str:
    """Return stock-center collection, or custom label when no stock is available."""
    collection = str(match.collection or "").strip()
    if collection:
        return collection
    return CUSTOM_LAB_REAGENT_COLLECTION if not match.source_stock else "-"


def _build_main_sheet(
    matches: List[Gal4StockMatch],
    phenotype_df: pd.DataFrame,
) -> pd.DataFrame:
    """Build one row per unique stock/reagent across all collections."""
    source_to_stock_key: Dict[str, Tuple[str, str]] = {}

    rows_by_stock: Dict[Tuple[str, str], Dict[str, Any]] = {}

    def _stock_key(match: Gal4StockMatch) -> Tuple[str, str]:
        collection = _collection_output_for_match(match)
        stock_id = match.fbst or match.source_stock or match.input_symbol
        return (collection, stock_id)

    def _ensure_row(match: Gal4StockMatch) -> Dict[str, Any]:
        key = _stock_key(match)
        if key not in rows_by_stock:
            stock_id = match.fbst or match.source_stock or match.input_symbol or "-"
            rows_by_stock[key] = {
                STOCK_ID_COLUMN: stock_id,
                INPUT_SYMBOLS_COLUMN: [],
                DRIVER_GENOTYPE_COLUMN: [],
                "Resolution method": [],
                "Match confidence": [],
                STOCK_GENOTYPE_COLUMN: match.stock_genotype or "-",
                "Collection (output)": _collection_output_for_match(match),
                "Stock # (output)": match.stock_number or "-",
                "Source/ Stock # (output)": match.source_stock or "-",
                "FBst (output)": match.fbst or "-",
                "Phenotypes": [],
                "Qualifiers": [],
                "Reference IDs": [],
                "Reference titles": [],
                "Authors": [],
                "Journals": [],
                "Years": [],
                PHENOTYPE_GENOTYPES_COLUMN: [],
            }
        if match.input_symbol:
            rows_by_stock[key][INPUT_SYMBOLS_COLUMN].append(match.input_symbol)
        if match.driver_genotype:
            rows_by_stock[key][DRIVER_GENOTYPE_COLUMN].append(match.driver_genotype)
        if match.resolution_method:
            rows_by_stock[key]["Resolution method"].append(match.resolution_method)
        if match.match_confidence:
            rows_by_stock[key]["Match confidence"].append(match.match_confidence)
        if match.source_stock:
            source_to_stock_key[match.source_stock] = key
        return rows_by_stock[key]

    for match in matches:
        _ensure_row(match)

    if len(phenotype_df) > 0 and SOURCE_STOCK_COLUMN in phenotype_df.columns:
        for _, row in phenotype_df.iterrows():
            source_stock = str(row.get(SOURCE_STOCK_COLUMN, "") or "").strip()
            if not source_stock:
                continue
            stock_key = source_to_stock_key.get(source_stock)
            if stock_key is None or stock_key not in rows_by_stock:
                continue
            target = rows_by_stock[stock_key]
            phenotype = str(row.get("Phenotype", "") or "").strip()
            qualifier = str(row.get("Qualifier", "") or "").strip()
            genotype = str(row.get("Genotype", "") or "").strip()
            ref_id = format_reference_id_display(row.get("PMID"), row.get("PMCID"))
            if phenotype:
                target["Phenotypes"].append(_format_referenced_phenotype(ref_id, phenotype))
            if qualifier:
                target["Qualifiers"].append(qualifier)
            if genotype:
                target[PHENOTYPE_GENOTYPES_COLUMN].append(genotype)
            if ref_id != "-":
                target["Reference IDs"].append(ref_id)
            title = str(row.get("Reference", "") or "").strip()
            authors = str(row.get("Authors", "") or "").strip()
            journal = str(row.get("Journal", "") or "").strip()
            year = str(row.get("Year of Publication", "") or "").strip()
            if title:
                target["Reference titles"].append(title)
            if authors:
                target["Authors"].append(authors)
            if journal:
                target["Journals"].append(journal)
            if year:
                target["Years"].append(year)

    output_rows: List[Dict[str, str]] = []
    for key in sorted(rows_by_stock.keys(), key=lambda item: (item[0].lower(), item[1])):
        row = rows_by_stock[key]
        output_rows.append(
            {
                STOCK_ID_COLUMN: row[STOCK_ID_COLUMN],
                INPUT_SYMBOLS_COLUMN: _aggregate_reference_field(row[INPUT_SYMBOLS_COLUMN]),
                DRIVER_GENOTYPE_COLUMN: _aggregate_reference_field(row[DRIVER_GENOTYPE_COLUMN]),
                "Resolution method": _aggregate_reference_field(row["Resolution method"]),
                "Match confidence": _aggregate_reference_field(row["Match confidence"]),
                STOCK_GENOTYPE_COLUMN: row[STOCK_GENOTYPE_COLUMN],
                "Collection (output)": row["Collection (output)"],
                "Stock # (output)": row["Stock # (output)"],
                "Source/ Stock # (output)": row["Source/ Stock # (output)"],
                "FBst (output)": row["FBst (output)"],
                "Phenotypes": _aggregate_referenced_phenotypes(row["Phenotypes"]),
                "Qualifiers": _aggregate_reference_field(row["Qualifiers"]),
                "Reference IDs": _aggregate_reference_field(row["Reference IDs"]),
                "Reference titles": _aggregate_reference_field(row["Reference titles"]),
                "Authors": _aggregate_reference_field(row["Authors"]),
                "Journals": _aggregate_reference_field(row["Journals"]),
                "Years": _aggregate_reference_field(row["Years"]),
                PHENOTYPE_GENOTYPES_COLUMN: _aggregate_reference_field(
                    row[PHENOTYPE_GENOTYPES_COLUMN]
                ),
            }
        )
    return pd.DataFrame(output_rows)


def _build_coverage_sheet(matches: List[Gal4StockMatch]) -> pd.DataFrame:
    """Build one audit row per input so unresolved symbols cannot disappear."""
    by_input: Dict[str, List[Gal4StockMatch]] = defaultdict(list)
    for match in matches:
        by_input[match.input_symbol].append(match)

    rows: List[Dict[str, str]] = []
    for input_symbol in sorted(by_input, key=str.lower):
        input_matches = by_input[input_symbol]
        resolved = [match for match in input_matches if match.source_stock]
        is_substitution = any(
            match.resolution_method == "BDSC promoter-family substitution"
            for match in resolved
        )
        if resolved:
            status = (
                "Resolved candidate substitution"
                if is_substitution
                else "Resolved exact/alias"
            )
        else:
            status = "Unresolved"
        rows.append(
            {
                GAL4_SYMBOL_COLUMN: input_symbol,
                "Coverage status": status,
                "Candidate stock count": str(len(resolved)),
                "Stock numbers": _aggregate_reference_field(
                    match.stock_number for match in resolved
                ),
                "Collections": _aggregate_reference_field(
                    match.collection for match in resolved
                ),
                "Resolved driver genotypes": _aggregate_reference_field(
                    match.driver_genotype for match in resolved
                ),
                "Resolution method": _aggregate_reference_field(
                    match.resolution_method for match in input_matches
                ),
                "Match confidence": _aggregate_reference_field(
                    match.match_confidence for match in input_matches
                ),
            }
        )
    return pd.DataFrame(rows)


def _build_references_sheet(
    matches: List[Gal4StockMatch],
    phenotype_df: pd.DataFrame,
) -> pd.DataFrame:
    """Build one row per unique phenotype reference."""
    source_to_inputs: Dict[str, Set[str]] = defaultdict(set)
    match_by_source: Dict[str, Gal4StockMatch] = {}
    for match in matches:
        if match.source_stock:
            source_to_inputs[match.source_stock].add(match.input_symbol)
            match_by_source[match.source_stock] = match

    rows_by_ref: Dict[str, Dict[str, Any]] = {}

    def _blank_ref_row(ref_key: str, row: pd.Series) -> Dict[str, Any]:
        return {
            "Reference ID": format_reference_id_display(row.get("PMID"), row.get("PMCID"))
            if ref_key
            else "-",
            "PMID": format_pmid_display(row.get("PMID")),
            "PMCID": format_pmcid_display(row.get("PMCID")),
            "Title": str(row.get("Reference", "") or "").strip() or "-",
            "Authors": str(row.get("Authors", "") or "").strip() or "-",
            "Journal": str(row.get("Journal", "") or "").strip() or "-",
            "Year": str(row.get("Year of Publication", "") or "").strip() or "-",
            "Associated GAL4 symbols (input)": [],
            "Associated driver genotypes": [],
            "Associated source stocks": [],
            "Associated stock IDs": [],
            "Associated collections": [],
            "Phenotypes": [],
            "Qualifiers": [],
            PHENOTYPE_GENOTYPES_COLUMN: [],
        }

    if len(phenotype_df) == 0 or SOURCE_STOCK_COLUMN not in phenotype_df.columns:
        return pd.DataFrame(
            columns=[
                "Reference ID",
                "PMID",
                "PMCID",
                "Title",
                "Authors",
                "Journal",
                "Year",
                "Associated GAL4 symbols (input)",
                "Associated driver genotypes",
                "Associated source stocks",
                "Associated stock IDs",
                "Associated collections",
                "Phenotypes",
                "Qualifiers",
                PHENOTYPE_GENOTYPES_COLUMN,
            ]
        )

    for _, row in phenotype_df.iterrows():
        ref_key = _reference_key(row.get("PMID"), row.get("PMCID"))
        if not ref_key:
            continue
        source_stock = str(row.get(SOURCE_STOCK_COLUMN, "") or "").strip()
        input_symbols = sorted(source_to_inputs.get(source_stock, set()))
        if not input_symbols:
            continue
        match = match_by_source.get(source_stock)
        if match is None:
            continue

        target = rows_by_ref.setdefault(ref_key, _blank_ref_row(ref_key, row))
        target["Associated GAL4 symbols (input)"].extend(input_symbols)
        if match.driver_genotype:
            target["Associated driver genotypes"].append(match.driver_genotype)
        target["Associated source stocks"].append(source_stock)
        target["Associated stock IDs"].append(match.fbst or source_stock)
        target["Associated collections"].append(_collection_output_for_match(match))

        phenotype = str(row.get("Phenotype", "") or "").strip()
        qualifier = str(row.get("Qualifier", "") or "").strip()
        genotype = str(row.get("Genotype", "") or "").strip()
        if phenotype:
            target["Phenotypes"].append(
                _format_referenced_phenotype(target["Reference ID"], phenotype)
            )
        if qualifier:
            target["Qualifiers"].append(qualifier)
        if genotype:
            target[PHENOTYPE_GENOTYPES_COLUMN].append(genotype)

    output_rows: List[Dict[str, str]] = []
    for ref_key in sorted(rows_by_ref):
        row = rows_by_ref[ref_key]
        output_rows.append(
            {
                "Reference ID": row["Reference ID"],
                "PMID": row["PMID"],
                "PMCID": row["PMCID"],
                "Title": row["Title"],
                "Authors": row["Authors"],
                "Journal": row["Journal"],
                "Year": row["Year"],
                "Associated GAL4 symbols (input)": _aggregate_reference_field(
                    row["Associated GAL4 symbols (input)"]
                ),
                "Associated driver genotypes": _aggregate_reference_field(
                    row["Associated driver genotypes"]
                ),
                "Associated source stocks": _aggregate_reference_field(
                    row["Associated source stocks"]
                ),
                "Associated stock IDs": _aggregate_reference_field(
                    row["Associated stock IDs"]
                ),
                "Associated collections": _aggregate_reference_field(
                    row["Associated collections"]
                ),
                "Phenotypes": _aggregate_referenced_phenotypes(row["Phenotypes"]),
                "Qualifiers": _aggregate_reference_field(row["Qualifiers"]),
                PHENOTYPE_GENOTYPES_COLUMN: _aggregate_reference_field(
                    row[PHENOTYPE_GENOTYPES_COLUMN]
                ),
            }
        )
    return pd.DataFrame(output_rows)


def _sanitize_sheet_name(name: str, used: Set[str]) -> str:
    cleaned = INVALID_SHEET_CHARS.sub("_", name)
    cleaned = cleaned[:EXCEL_SHEET_NAME_MAX] or "Reference"
    candidate = cleaned
    suffix = 2
    while candidate in used:
        tail = f"_{suffix}"
        candidate = f"{cleaned[: EXCEL_SHEET_NAME_MAX - len(tail)]}{tail}"
        suffix += 1
    used.add(candidate)
    return candidate


def _has_meaningful_collection_sheet_value(row: pd.Series, columns: Iterable[str]) -> bool:
    for column in columns:
        value = str(row.get(column, "") or "").strip()
        if value and value != "-":
            return True
    return False


def _drop_placeholder_custom_lab_rows(collection_df: pd.DataFrame) -> pd.DataFrame:
    """Remove unresolved input-only rows from the Custom Lab Reagent worksheet."""
    if len(collection_df) == 0:
        return collection_df
    evidence_or_resolved_columns = (
        DRIVER_GENOTYPE_COLUMN,
        STOCK_GENOTYPE_COLUMN,
        "Stock # (output)",
        "Source/ Stock # (output)",
        "FBst (output)",
        "Phenotypes",
        "Qualifiers",
        "Reference IDs",
        "Reference titles",
        "Authors",
        "Journals",
        "Years",
        PHENOTYPE_GENOTYPES_COLUMN,
    )
    keep_mask = collection_df.apply(
        lambda row: _has_meaningful_collection_sheet_value(
            row,
            evidence_or_resolved_columns,
        ),
        axis=1,
    )
    return collection_df[keep_mask].copy()


def _build_collection_sheets(main_df: pd.DataFrame) -> Dict[str, pd.DataFrame]:
    if len(main_df) == 0 or "Collection (output)" not in main_df.columns:
        return {CUSTOM_LAB_REAGENT_COLLECTION: main_df.copy()}

    collection_order = sorted(
        {
            str(value or "").strip() or CUSTOM_LAB_REAGENT_COLLECTION
            for value in main_df["Collection (output)"].tolist()
        },
        key=lambda value: (value == CUSTOM_LAB_REAGENT_COLLECTION, value.lower()),
    )
    sheets: Dict[str, pd.DataFrame] = {}
    for collection in collection_order:
        collection_df = main_df[
            main_df["Collection (output)"].fillna("").astype(str).str.strip().replace(
                "", CUSTOM_LAB_REAGENT_COLLECTION
            )
            == collection
        ].copy()
        if collection == CUSTOM_LAB_REAGENT_COLLECTION:
            collection_df = _drop_placeholder_custom_lab_rows(collection_df)
        sheets[collection] = collection_df
    return sheets


def _format_worksheet(ws) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    wrapped_headers = {
        INPUT_SYMBOLS_COLUMN,
        "Stock genotype (output)",
        "Resolved driver genotype (output)",
        "Phenotypes",
        "Qualifiers",
        "Reference IDs",
        "Reference titles",
        "Authors",
        "Journals",
        "Title",
        "Associated GAL4 symbols (input)",
        "Associated driver genotypes",
        "Associated source stocks",
        "Associated stock IDs",
        "Associated collections",
        "Phenotype record genotypes",
    }
    max_width_by_header = {
        "Stock genotype (output)": 55,
        "Resolved driver genotype (output)": 42,
        "Reference titles": 55,
        "Title": 65,
        "Authors": 45,
        "Journals": 35,
        "Phenotypes": 62,
        "Qualifiers": 28,
        "Associated GAL4 symbols (input)": 38,
        "Associated driver genotypes": 42,
        "Associated source stocks": 34,
        "Associated stock IDs": 34,
        "Associated collections": 30,
        "Phenotype record genotypes": 55,
    }
    min_width_by_header = {
        STOCK_ID_COLUMN: 16,
        INPUT_SYMBOLS_COLUMN: 22,
        DRIVER_GENOTYPE_COLUMN: 24,
        STOCK_GENOTYPE_COLUMN: 28,
        "Collection (output)": 18,
        "Stock # (output)": 14,
        "Reference ID": 16,
        "Reference IDs": 18,
        "Phenotypes": 34,
        "PMID": 14,
        "PMCID": 14,
    }
    wrapped_col_widths: Dict[int, float] = {}
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        header = str(ws.cell(row=1, column=col_idx).value or "")
        max_len = len(header)
        for row_idx in range(2, min(ws.max_row, 80) + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            value_lines = str(value).splitlines() or [str(value)]
            longest_line = max(len(line) for line in value_lines)
            max_len = max(max_len, min(longest_line, 80))
            if header in wrapped_headers:
                ws.cell(row=row_idx, column=col_idx).alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )
            else:
                ws.cell(row=row_idx, column=col_idx).alignment = Alignment(vertical="top")
        width = min(max(max_len + 2, min_width_by_header.get(header, 10)), 32)
        if header in max_width_by_header:
            width = min(max(max_len + 2, min_width_by_header.get(header, 16)), max_width_by_header[header])
        ws.column_dimensions[letter].width = width
        if header in wrapped_headers:
            wrapped_col_widths[col_idx] = width

    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = _estimate_wrapped_row_height(
            ws,
            row_idx,
            wrapped_col_widths,
        )


def _estimate_wrapped_row_height(
    ws,
    row_idx: int,
    wrapped_col_widths: Dict[int, float],
) -> float:
    """Estimate a readable row height for wrapped evidence-heavy workbook cells."""
    max_lines = 1
    for col_idx, width in wrapped_col_widths.items():
        value = ws.cell(row=row_idx, column=col_idx).value
        if value is None:
            continue
        chars_per_line = max(10, int(width * 1.05))
        line_count = 0
        for raw_line in str(value).splitlines() or [""]:
            line_count += max(1, math.ceil(len(raw_line) / chars_per_line))
        max_lines = max(max_lines, line_count)

    if max_lines <= 1:
        return 34
    return min(132, 22 + (max_lines * 15))


def write_gal4_workbook(
    output_path: Path,
    main_df: pd.DataFrame,
    references_df: pd.DataFrame,
    coverage_df: Optional[pd.DataFrame] = None,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    used_sheet_names: Set[str] = set()
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for collection, collection_df in _build_collection_sheets(main_df).items():
            sheet_name = _sanitize_sheet_name(collection, used_sheet_names)
            collection_out = collection_df.copy()
            if "Stock # (output)" in collection_out.columns:
                collection_out["Stock # (output)"] = (
                    collection_out["Stock # (output)"].apply(clean_id).replace("", "-")
                )
            collection_out.to_excel(writer, sheet_name=sheet_name, index=False)

        if coverage_df is not None:
            coverage_name = _sanitize_sheet_name(COVERAGE_SHEET_NAME, used_sheet_names)
            coverage_df.to_excel(writer, sheet_name=coverage_name, index=False)

        references_name = _sanitize_sheet_name(REFERENCES_SHEET_NAME, used_sheet_names)
        references_df.to_excel(writer, sheet_name=references_name, index=False)

        for ws in writer.book.worksheets:
            _format_worksheet(ws)


def export_gal4_driver_workbook(
    input_csv: Path,
    output_path: Path,
    *,
    fetch_pubmed: bool = True,
    verbose: bool = True,
) -> Path:
    settings = Settings()
    symbols = _read_gal4_symbols(input_csv)
    if verbose:
        print(f"Loaded {len(symbols)} unique GAL4 symbol(s) from {input_csv}", flush=True)

    derived_df = _load_derived_components(settings)
    bdsc_data_path = Path(settings.flybase_data_path).parent / "bdsc"
    derived_df = _merge_native_bdsc_stock_data(derived_df, bdsc_data_path)
    if verbose and (
        (bdsc_data_path / BDSC_STOCKS_FILENAME).exists()
        and (bdsc_data_path / BDSC_STOCK_COMPONENTS_FILENAME).exists()
    ):
        print(
            f"Loaded current native BDSC availability/components from {bdsc_data_path}",
            flush=True,
        )
    (
        component_id_to_symbol,
        component_id_to_gene_symbol,
        component_id_to_stock_candidate_details,
        component_rows_by_fbst,
        gal4_only_fbst,
        stock_candidate_label_to_driver_genotype,
        driver_stock_index,
        fbst_to_fb_genotype,
        resolution_index,
    ) = _build_gal4_lookup_context(derived_df)
    construct_alias_fbti_index = _load_gal4_construct_alias_fbti_index(
        settings.flybase_data_path
    )

    all_matches: List[Gal4StockMatch] = []
    for symbol in symbols:
        all_matches.extend(
            _resolve_gal4_symbol_to_stocks(
                symbol,
                gal4_only_fbst,
                driver_stock_index,
                component_id_to_symbol,
                component_id_to_gene_symbol,
                component_id_to_stock_candidate_details,
                component_rows_by_fbst,
                stock_candidate_label_to_driver_genotype,
                fbst_to_fb_genotype,
                resolution_index,
                construct_alias_fbti_index,
            )
        )

    _, reagent_index_df = _build_synthetic_inputs(
        all_matches,
        flybase_data_path=settings.flybase_data_path,
        component_id_to_symbol=component_id_to_symbol,
    )
    if verbose:
        resolved = len({match.source_stock for match in all_matches if match.source_stock})
        print(f"Resolved {resolved} stock candidate(s) across input symbols", flush=True)

    phenotype_df = _build_gal4_phenotype_evidence(
        all_matches,
        reagent_index_df,
        settings.flybase_data_path,
        settings.pubmed_cache_path,
        fetch_pubmed=fetch_pubmed,
        verbose=verbose,
    )
    if verbose:
        print(f"Built {len(phenotype_df)} phenotype evidence row(s)", flush=True)

    main_df = _build_main_sheet(all_matches, phenotype_df)
    coverage_df = _build_coverage_sheet(all_matches)
    references_df = _build_references_sheet(all_matches, phenotype_df)
    collection_count = len(_build_collection_sheets(main_df))
    write_gal4_workbook(
        output_path,
        main_df,
        references_df,
        coverage_df=coverage_df,
    )
    if verbose:
        print(
            f"Wrote workbook: {output_path} "
            f"({collection_count} collection sheet(s) + Coverage + References)",
            flush=True,
        )
    return output_path


def main(argv: Optional[List[str]] = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    input_csv = args.input_csv.resolve()
    if not input_csv.exists():
        print(f"ERROR: input CSV not found: {input_csv}", file=sys.stderr)
        return 1
    output_path = args.output
    if output_path is None:
        output_path = input_csv.with_name(f"{input_csv.stem}_gal4_stocks.xlsx")
    else:
        output_path = output_path.resolve()

    try:
        export_gal4_driver_workbook(
            input_csv,
            output_path,
            fetch_pubmed=not args.no_fetch_pubmed,
            verbose=not args.quiet,
        )
    except Exception as exc:  # noqa: BLE001 - surface CLI failures clearly
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
