#!/usr/bin/env python3
"""
Aggregate similarity-workbook Contents tables across a directory tree.

This script scans recursively for Excel workbooks, identifies phenotype
similarity sidecar workbooks from their internal `Contents` sheet layout, and
writes one summary workbook into the input directory.
"""

from __future__ import annotations

import argparse
import math
import re
import sys
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font


REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))


COSINE_WORKBOOK_TITLE = "Tier Workbook Contents"
SIMPLE_WORKBOOK_TITLE = "Simple Bucket Workbook Contents"
COSINE_SUMMARY_SHEET = "Cosine Tier Summary"
SIMPLE_SUMMARY_SHEET = "Simple Bucket Summary"

COSINE_HEADER = ("Sheet", "Meaning", "Rows")
SIMPLE_SECTIONS = {
    "Stocks per bucket": "stock_count",
    "Alleles per bucket": "allele_count",
    "Genes per bucket": "gene_count",
}
SIMPLE_SECTION_LABELS = {
    "stock_count": "Stocks per bucket",
    "allele_count": "Alleles per bucket",
    "gene_count": "Genes per bucket",
}

COSINE_RANGE_RE = re.compile(r"Tier for similarity range (.+?)\.\s*$")


@dataclass
class CosineWorkbookSummary:
    path: Path
    tier_counts: Dict[str, int]


@dataclass
class SimpleBucketWorkbookSummary:
    path: Path
    collections: List[str]
    bucket_headers: List[str]
    metrics: Dict[str, Dict[str, Dict[str, int]]]


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Recursively scan a directory for phenotype similarity workbooks and "
            "write one aggregated summary workbook into the input directory."
        )
    )
    parser.add_argument(
        "input_dir",
        type=Path,
        help="Directory to scan recursively for similarity workbooks.",
    )
    parser.add_argument(
        "--output-name",
        default="similarity_workbook_summary.xlsx",
        help="Filename for the aggregated workbook written into input_dir.",
    )
    return parser


def _normalize_cell(value: object) -> str:
    if value is None:
        return ""
    if pd.isna(value):
        return ""
    text = str(value).strip()
    return "" if text.lower() == "nan" else text


def _iter_rows(contents_df: pd.DataFrame) -> Iterable[List[str]]:
    for _, row in contents_df.iterrows():
        yield [_normalize_cell(value) for value in row.tolist()]


def _find_row_prefix(contents_df: pd.DataFrame, prefix: Sequence[str]) -> Optional[int]:
    normalized_prefix = tuple(prefix)
    for idx, row in enumerate(_iter_rows(contents_df)):
        if tuple(row[: len(normalized_prefix)]) == normalized_prefix:
            return idx
    return None


def _sheet_contains(contents_df: pd.DataFrame, needle: str) -> bool:
    for row in _iter_rows(contents_df):
        if needle in row:
            return True
    return False


def _coerce_non_negative_int(value: object, context: str) -> int:
    if value is None or pd.isna(value):
        raise ValueError(f"{context} is missing")
    numeric = pd.to_numeric(pd.Series([value]), errors="coerce").iloc[0]
    if pd.isna(numeric):
        raise ValueError(f"{context} is not numeric: {value!r}")
    numeric_float = float(numeric)
    if not math.isfinite(numeric_float) or numeric_float < 0:
        raise ValueError(f"{context} must be a non-negative integer: {value!r}")
    numeric_int = int(round(numeric_float))
    if not math.isclose(numeric_float, numeric_int, rel_tol=0.0, abs_tol=1e-9):
        raise ValueError(f"{context} must be a whole number: {value!r}")
    return numeric_int


def _extract_cosine_range_label(sheet_name: str, meaning: str) -> str:
    match = COSINE_RANGE_RE.search(meaning or "")
    if match:
        return match.group(1).strip()
    return sheet_name.strip()


def _sort_cosine_range_labels(labels: Iterable[str]) -> List[str]:
    def _sort_key(label: str) -> tuple[float, float, str]:
        normalized = label.strip()
        if normalized.startswith("<"):
            try:
                upper = float(normalized[1:])
            except ValueError:
                return (-1.0, -1.0, normalized)
            return (-1.0, upper, normalized)
        if "-" in normalized:
            left, _, right = normalized.partition("-")
            try:
                lower = float(left)
                upper = float(right)
            except ValueError:
                return (-2.0, -2.0, normalized)
            return (lower, upper, normalized)
        try:
            value = float(normalized)
        except ValueError:
            return (-3.0, -3.0, normalized)
        return (value, value, normalized)

    return sorted(labels, key=_sort_key, reverse=True)


def _read_contents_sheet(path: Path) -> Optional[pd.DataFrame]:
    try:
        return pd.read_excel(path, sheet_name="Contents", header=None)
    except ValueError:
        return None
    except Exception as exc:
        raise ValueError(f"could not read Contents sheet: {exc}") from exc


def parse_cosine_contents(contents_df: pd.DataFrame) -> Dict[str, int]:
    if not _sheet_contains(contents_df, COSINE_WORKBOOK_TITLE):
        raise ValueError("missing cosine contents title")

    header_idx = _find_row_prefix(contents_df, COSINE_HEADER)
    if header_idx is None:
        raise ValueError("missing cosine contents table header")

    tier_counts: Dict[str, int] = {}
    for row_offset, row in enumerate(_iter_rows(contents_df.iloc[header_idx + 1 :]), start=header_idx + 2):
        sheet_name = row[0] if len(row) > 0 else ""
        meaning = row[1] if len(row) > 1 else ""
        row_count = row[2] if len(row) > 2 else ""
        if not sheet_name:
            continue
        if sheet_name in {"Gene Set", "Stock Phenotype Sheet"}:
            continue
        tier_label = _extract_cosine_range_label(sheet_name, meaning)
        tier_counts[tier_label] = _coerce_non_negative_int(
            row_count,
            context=f"{tier_label} row count in Contents row {row_offset}",
        )

    if not tier_counts:
        raise ValueError("cosine contents workbook contained no tier rows")
    return tier_counts


def _parse_simple_matrix(
    contents_df: pd.DataFrame,
    section_label: str,
) -> tuple[List[str], Dict[str, Dict[str, int]]]:
    section_idx = _find_row_prefix(contents_df, (section_label,))
    if section_idx is None:
        raise ValueError(f"missing simple-bucket section: {section_label}")

    header_idx = section_idx + 1
    if header_idx >= len(contents_df):
        raise ValueError(f"missing header after section: {section_label}")

    header_row = [_normalize_cell(value) for value in contents_df.iloc[header_idx].tolist()]
    if not header_row or header_row[0] != "Collection":
        raise ValueError(f"section {section_label} is missing a Collection header row")

    total_indices = [idx for idx, value in enumerate(header_row) if value == "Total"]
    if not total_indices:
        raise ValueError(f"section {section_label} is missing a Total column")
    total_idx = total_indices[-1]

    bucket_columns = [
        (column_idx, header_row[column_idx])
        for column_idx in range(1, total_idx)
        if header_row[column_idx]
    ]
    if not bucket_columns:
        raise ValueError(f"section {section_label} has no bucket columns")

    collection_rows: Dict[str, Dict[str, int]] = {}
    current_idx = header_idx + 1
    while current_idx < len(contents_df):
        row = [_normalize_cell(value) for value in contents_df.iloc[current_idx].tolist()]
        label = row[0] if row else ""
        if not label:
            current_idx += 1
            continue
        if label == "Total":
            break

        counts_by_bucket: Dict[str, int] = {}
        for column_idx, bucket_label in bucket_columns:
            value = row[column_idx] if column_idx < len(row) else ""
            counts_by_bucket[bucket_label] = _coerce_non_negative_int(
                value,
                context=f"{section_label} count for collection {label!r}, bucket {bucket_label!r}",
            )
        collection_rows[label] = counts_by_bucket
        current_idx += 1
    else:
        raise ValueError(f"section {section_label} is missing its footer row")

    bucket_headers = [bucket_label for _, bucket_label in bucket_columns]
    return bucket_headers, collection_rows


def parse_simple_bucket_contents(contents_df: pd.DataFrame) -> SimpleBucketWorkbookSummary:
    raise NotImplementedError("use parse_simple_bucket_contents_from_dataframe instead")


def parse_simple_bucket_contents_from_dataframe(
    path: Path,
    contents_df: pd.DataFrame,
) -> SimpleBucketWorkbookSummary:
    if not _sheet_contains(contents_df, SIMPLE_WORKBOOK_TITLE):
        raise ValueError("missing simple-bucket contents title")

    metric_rows: Dict[str, Dict[str, Dict[str, int]]] = {}
    collection_order: List[str] = []
    bucket_header_order: List[str] = []

    for section_label, metric_key in SIMPLE_SECTIONS.items():
        bucket_headers, rows_by_collection = _parse_simple_matrix(contents_df, section_label)
        metric_rows[metric_key] = rows_by_collection
        for collection in rows_by_collection:
            if collection not in collection_order:
                collection_order.append(collection)
        for bucket_header in bucket_headers:
            if bucket_header not in bucket_header_order:
                bucket_header_order.append(bucket_header)

    if _find_row_prefix(contents_df, ("Sheet legend",)) is None:
        raise ValueError("missing simple-bucket sheet legend section")

    return SimpleBucketWorkbookSummary(
        path=path,
        collections=collection_order,
        bucket_headers=bucket_header_order,
        metrics=metric_rows,
    )


def _classify_contents(contents_df: pd.DataFrame) -> Optional[str]:
    if _sheet_contains(contents_df, COSINE_WORKBOOK_TITLE):
        return "cosine"
    if _sheet_contains(contents_df, SIMPLE_WORKBOOK_TITLE):
        return "simple"
    return None


def discover_similarity_workbooks(
    input_dir: Path,
    output_path: Path,
) -> tuple[List[CosineWorkbookSummary], List[SimpleBucketWorkbookSummary], List[str]]:
    cosine_workbooks: List[CosineWorkbookSummary] = []
    simple_workbooks: List[SimpleBucketWorkbookSummary] = []
    warnings: List[str] = []

    for path in sorted(input_dir.rglob("*.xlsx")):
        if path.resolve() == output_path.resolve():
            continue

        contents_df = _read_contents_sheet(path)
        if contents_df is None:
            continue

        kind = _classify_contents(contents_df)
        if kind is None:
            continue

        try:
            if kind == "cosine":
                cosine_workbooks.append(
                    CosineWorkbookSummary(
                        path=path,
                        tier_counts=parse_cosine_contents(contents_df),
                    )
                )
            else:
                simple_workbooks.append(
                    parse_simple_bucket_contents_from_dataframe(path, contents_df)
                )
        except ValueError as exc:
            warnings.append(f"Skipping {path}: {exc}")

    return cosine_workbooks, simple_workbooks, warnings


def aggregate_cosine_counts(
    workbooks: Sequence[CosineWorkbookSummary],
) -> Dict[str, int]:
    totals: Dict[str, int] = defaultdict(int)
    for workbook in workbooks:
        for tier_label, count in workbook.tier_counts.items():
            totals[tier_label] += int(count)
    ordered_totals = {
        tier_label: totals[tier_label]
        for tier_label in _sort_cosine_range_labels(totals.keys())
    }
    return ordered_totals


def aggregate_simple_bucket_counts(
    workbooks: Sequence[SimpleBucketWorkbookSummary],
) -> tuple[List[str], List[str], Dict[str, Dict[str, Dict[str, int]]]]:
    collection_order: List[str] = []
    bucket_order: List[str] = []
    totals: Dict[str, Dict[str, Dict[str, int]]] = {
        metric_key: defaultdict(lambda: defaultdict(int))
        for metric_key in SIMPLE_SECTION_LABELS
    }

    for workbook in workbooks:
        for collection in workbook.collections:
            if collection not in collection_order:
                collection_order.append(collection)
        for bucket_header in workbook.bucket_headers:
            if bucket_header not in bucket_order:
                bucket_order.append(bucket_header)

        for metric_key, rows_by_collection in workbook.metrics.items():
            for collection, counts_by_bucket in rows_by_collection.items():
                for bucket_header, count in counts_by_bucket.items():
                    totals[metric_key][collection][bucket_header] += int(count)

    materialized: Dict[str, Dict[str, Dict[str, int]]] = {}
    for metric_key, rows_by_collection in totals.items():
        materialized[metric_key] = {
            collection: dict(counts_by_bucket)
            for collection, counts_by_bucket in rows_by_collection.items()
        }
    return collection_order, bucket_order, materialized


def _autosize_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        values = ["" if cell.value is None else str(cell.value) for cell in column_cells]
        width = max((len(value) for value in values), default=10) + 2
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(width, 12), 48)


def write_summary_workbook(
    output_path: Path,
    cosine_workbooks: Sequence[CosineWorkbookSummary],
    simple_workbooks: Sequence[SimpleBucketWorkbookSummary],
) -> Optional[Path]:
    if not cosine_workbooks and not simple_workbooks:
        return None

    workbook = Workbook()
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True)
    sheet_created = False

    if cosine_workbooks:
        worksheet = workbook.active if not sheet_created else workbook.create_sheet(COSINE_SUMMARY_SHEET)
        worksheet.title = COSINE_SUMMARY_SHEET
        sheet_created = True

        cosine_totals = aggregate_cosine_counts(cosine_workbooks)
        worksheet["A1"] = COSINE_SUMMARY_SHEET
        worksheet["A1"].font = title_font
        worksheet["A2"] = "Source workbooks aggregated"
        worksheet["B2"] = len(cosine_workbooks)
        worksheet["A4"] = "Similarity Range"
        worksheet["B4"] = "Total Rows"
        worksheet["A4"].font = header_font
        worksheet["B4"].font = header_font

        row_idx = 5
        for tier_label, total_rows in cosine_totals.items():
            worksheet.cell(row=row_idx, column=1, value=tier_label)
            worksheet.cell(row=row_idx, column=2, value=total_rows)
            row_idx += 1

        worksheet.cell(row=row_idx, column=1, value="Total")
        worksheet.cell(row=row_idx, column=2, value=sum(cosine_totals.values()))
        worksheet.cell(row=row_idx, column=1).font = header_font
        worksheet.cell(row=row_idx, column=2).font = header_font
        _autosize_columns(worksheet)

    if simple_workbooks:
        worksheet = workbook.active if not sheet_created else workbook.create_sheet(SIMPLE_SUMMARY_SHEET)
        worksheet.title = SIMPLE_SUMMARY_SHEET
        sheet_created = True

        collection_order, bucket_order, simple_totals = aggregate_simple_bucket_counts(simple_workbooks)
        worksheet["A1"] = SIMPLE_SUMMARY_SHEET
        worksheet["A1"].font = title_font
        worksheet["A2"] = "Source workbooks aggregated"
        worksheet["B2"] = len(simple_workbooks)

        defn_font = Font(italic=True)
        definitions = [
            ("UAS / non-UAS",
             "Whether the stock genotype contains a UAS (Upstream Activation Sequence) construct."),
            ("slp/ circ / Non slp/ circ",
             "Whether any phenotype row linked to the stock mentions 'sleep' or 'circadian'."),
            ("No bal / Has bal",
             "Whether the stock carries at least one balancer chromosome (e.g. CyO, TM3, TM6B, FM7)."),
        ]
        row_idx = 4
        for defn_label, defn_text in definitions:
            worksheet.cell(row=row_idx, column=1, value=defn_label)
            worksheet.cell(row=row_idx, column=1).font = header_font
            worksheet.cell(row=row_idx, column=2, value=defn_text)
            worksheet.cell(row=row_idx, column=2).font = defn_font
            row_idx += 1
        row_idx += 1
        for metric_key in ("stock_count", "allele_count", "gene_count"):
            worksheet.cell(row=row_idx, column=1, value=SIMPLE_SECTION_LABELS[metric_key])
            worksheet.cell(row=row_idx, column=1).font = header_font
            row_idx += 1

            worksheet.cell(row=row_idx, column=1, value="Collection")
            worksheet.cell(row=row_idx, column=1).font = header_font
            for bucket_idx, bucket_label in enumerate(bucket_order, start=2):
                worksheet.cell(row=row_idx, column=bucket_idx, value=bucket_label)
                worksheet.cell(row=row_idx, column=bucket_idx).font = header_font
            total_column = len(bucket_order) + 2
            worksheet.cell(row=row_idx, column=total_column, value="Total")
            worksheet.cell(row=row_idx, column=total_column).font = header_font
            row_idx += 1

            column_totals = [0] * len(bucket_order)
            grand_total = 0
            for collection in collection_order:
                worksheet.cell(row=row_idx, column=1, value=collection)
                row_total = 0
                metric_counts = simple_totals.get(metric_key, {}).get(collection, {})
                for bucket_offset, bucket_label in enumerate(bucket_order, start=0):
                    value = int(metric_counts.get(bucket_label, 0))
                    worksheet.cell(row=row_idx, column=2 + bucket_offset, value=value)
                    column_totals[bucket_offset] += value
                    row_total += value
                worksheet.cell(row=row_idx, column=total_column, value=row_total)
                grand_total += row_total
                row_idx += 1

            worksheet.cell(row=row_idx, column=1, value="Total")
            worksheet.cell(row=row_idx, column=1).font = header_font
            for bucket_offset, value in enumerate(column_totals, start=0):
                worksheet.cell(row=row_idx, column=2 + bucket_offset, value=value)
                worksheet.cell(row=row_idx, column=2 + bucket_offset).font = header_font
            worksheet.cell(row=row_idx, column=total_column, value=grand_total)
            worksheet.cell(row=row_idx, column=total_column).font = header_font
            row_idx += 2

        _autosize_columns(worksheet)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def main(argv: Optional[Sequence[str]] = None) -> int:
    args = build_parser().parse_args(argv)
    input_dir = args.input_dir.resolve()

    if not input_dir.exists():
        raise SystemExit(f"Input directory does not exist: {input_dir}")
    if not input_dir.is_dir():
        raise SystemExit(f"Input path is not a directory: {input_dir}")

    output_path = input_dir / args.output_name
    cosine_workbooks, simple_workbooks, warnings = discover_similarity_workbooks(
        input_dir=input_dir,
        output_path=output_path,
    )

    for warning in warnings:
        print(warning)

    if not cosine_workbooks and not simple_workbooks:
        print(f"No supported similarity workbooks found under {input_dir}")
        return 0

    saved_path = write_summary_workbook(
        output_path=output_path,
        cosine_workbooks=cosine_workbooks,
        simple_workbooks=simple_workbooks,
    )
    if saved_path is None:
        print(f"No supported similarity workbooks found under {input_dir}")
        return 0

    print(f"Saved summary workbook: {saved_path}")
    print(f"  Cosine-tier workbooks: {len(cosine_workbooks)}")
    print(f"  Simple-bucket workbooks: {len(simple_workbooks)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
