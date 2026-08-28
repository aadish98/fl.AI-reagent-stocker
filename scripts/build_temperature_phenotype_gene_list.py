#!/usr/bin/env python3
"""Build a genome-wide temperature phenotypic-reagent Excel workbook.

The output contains:

* ``Reagents``: unique reagent -> phenotype -> reference rows, grouped by gene
  with genes having the most distinct reagents first.
* ``References``: unique reference -> reagent rows in first-appearance order.
* ``Genes`` and ``Summary``: compact supporting indexes.

Example:

    python scripts/build_temperature_phenotype_gene_list.py
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, Optional, Tuple

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from fl_ai_reagent_stocker.config import DEFAULT_CONTACT_EMAIL, Settings  # noqa: E402
from fl_ai_reagent_stocker.integrations.pubmed import (  # noqa: E402
    PubMedCache,
    PubMedClient,
)
from fl_ai_reagent_stocker.pipelines.stock_splitting import (  # noqa: E402
    _extract_flybase_ids,
)
from fl_ai_reagent_stocker.utils import (  # noqa: E402
    clean_id,
    find_latest_tsv,
    load_flybase_tsv,
    unique_join,
)


DEFAULT_OUTPUT = (
    REPO_ROOT
    / "data"
    / "gene_sets"
    / "temperature_phenotypes"
    / "temperature_phenotype_reagents.xlsx"
)
EXCLUDED_GENES = {"FBgn0003996": "w"}
TEMPERATURE_WORD_RE = re.compile(r"\b(?:temperature|heat|cold)\b|\bthermo", re.I)
FBRF_RE = re.compile(r"FBrf\d+", re.I)
UAS_RNAI_WORD_RE = re.compile(r"\b(?:uas|rnai)\b", re.I)
VDRC_KNOCKDOWN_ALLELE_RE = re.compile(r"[\[{](?:gd|kk|vsh|hms|jf)\d+", re.I)
REFERENCE_COLUMNS = [
    "PMID",
    "Title",
    "Authors",
    "Date of publication",
    "Journal",
    "Associated gene",
    "Associated reagent",
    "FBrf",
    "Stock #",
    "Collection",
    "Phenotype",
    "custom_stock",
]
REAGENT_COLUMNS = [
    "Gene",
    "flybase_gene_id",
    "Collection",
    "Stock Center",
    "Stock #",
    "FBst",
    "custom_stock",
    "Genotype",
    "Allele symbol",
    "Allele ID",
    "Phenotype",
    "Qualifier",
    "Reference (FBrf)",
    "PMID",
    "Reference Title",
    "keyword_hit",
    "keyword_match_terms",
    "n_reagents_for_gene",
]


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Write an Excel workbook of published Drosophila temperature-related "
            "phenotypic reagents from local FlyBase reports."
        )
    )
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help=f"Output workbook (default: {DEFAULT_OUTPUT.relative_to(REPO_ROOT)})",
    )
    parser.add_argument(
        "--skip-pubmed-fetch",
        action="store_true",
        help="Use cached/FlyBase reference metadata only; do not query PubMed.",
    )
    return parser


def ordered_unique(values: Iterable[Any]) -> str:
    """Semicolon-join unique non-empty values in encounter order."""
    seen = set()
    result: List[str] = []
    for value in values:
        if value is None or (isinstance(value, float) and pd.isna(value)):
            continue
        text = str(value).strip()
        if not text or text in seen:
            continue
        seen.add(text)
        result.append(text)
    return "; ".join(result)


def temperature_keyword_terms(phenotype_name: Any, qualifier_names: Any) -> List[str]:
    """Return matched temperature terms without treating ``sheath`` as heat."""
    text = f"{phenotype_name or ''} {qualifier_names or ''}"
    matched = []
    for match in TEMPERATURE_WORD_RE.finditer(text):
        token = match.group(0).lower()
        canonical = "thermo" if token.startswith("thermo") else token
        if canonical not in matched:
            matched.append(canonical)
    return matched


def is_uas_or_rnai_reagent(*text_values: Any) -> bool:
    """Return True when construct/allele text marks a reagent as UAS or RNAi.

    This is a text-match proxy (whole-word "UAS"/"RNAi", plus the common VDRC
    knockdown allele-bracket families GD/KK/VSH/HMS/JF) used to keep the
    workbook restricted to classical alleles/insertions rather than
    overexpression or knockdown transgenes.
    """
    combined = " ".join(
        str(value or "").strip() for value in text_values if str(value or "").strip()
    )
    if not combined:
        return False
    if UAS_RNAI_WORD_RE.search(combined):
        return True
    return bool(VDRC_KNOCKDOWN_ALLELE_RE.search(combined))


def split_fbrfs(value: Any) -> List[str]:
    """Extract ordered unique FlyBase publication IDs."""
    return list(dict.fromkeys(match for match in FBRF_RE.findall(str(value or ""))))


def parse_miniref(value: Any) -> Dict[str, str]:
    """Best-effort author/year/journal fallback for FlyBase miniref text."""
    text = str(value or "").strip()
    if not text:
        return {"title": "", "authors": "", "year": "", "journal": ""}
    parts = [part.strip() for part in text.split(",") if part.strip()]
    authors = parts[0] if parts else ""
    year = ""
    year_index: Optional[int] = None
    for index, part in enumerate(parts[1:], start=1):
        match = re.search(r"\b(?:19|20)\d{2}\b", part)
        if match:
            year = match.group(0)
            year_index = index
            break
    journal = (
        ", ".join(parts[year_index + 1 :]).strip()
        if year_index is not None
        else ", ".join(parts[1:]).strip()
    )
    return {"title": text, "authors": authors, "year": year, "journal": journal}


def load_dmel_gene_catalog(synonym_path: Path) -> pd.DataFrame:
    """Load current Dmel FBgn symbols from FlyBase's commented-header report."""
    columns = [
        "primary_FBid",
        "organism_abbreviation",
        "current_symbol",
        "current_fullname",
        "fullname_synonyms",
        "symbol_synonyms",
    ]
    catalog = pd.read_csv(
        synonym_path,
        sep="\t",
        comment="#",
        names=columns,
        dtype=str,
        keep_default_na=False,
    )
    catalog = catalog[
        catalog["organism_abbreviation"].eq("Dmel")
        & catalog["primary_FBid"].str.startswith("FBgn")
    ][["primary_FBid", "current_symbol"]].drop_duplicates(subset=["primary_FBid"])
    return catalog.rename(
        columns={
            "primary_FBid": "flybase_gene_id",
            "current_symbol": "current_gene_symbol",
        }
    )


def build_included_assertions(
    phenotype_df: pd.DataFrame,
    fbal_to_gene_df: pd.DataFrame,
    dmel_gene_catalog: Optional[pd.DataFrame] = None,
) -> pd.DataFrame:
    """Return cited, temperature-keyword, allele/insertion phenotype assertions.

    Inclusion is a direct text match: the phenotype/qualifier text must
    contain one of the temperature keywords (word-boundary "temperature",
    "heat", "cold", or a "thermo" prefix), and the reagent must be an
    allele/insertion rather than a UAS/RNAi construct.
    """
    frame = phenotype_df.copy()
    for column in (
        "genotype_symbols",
        "genotype_FBids",
        "phenotype_name",
        "qualifier_names",
        "reference",
    ):
        if column not in frame.columns:
            frame[column] = ""
        frame[column] = frame[column].fillna("").astype(str).str.strip()

    frame["keyword_match_terms"] = frame.apply(
        lambda row: "; ".join(
            temperature_keyword_terms(row["phenotype_name"], row["qualifier_names"])
        ),
        axis=1,
    )
    frame["keyword_hit"] = frame["keyword_match_terms"].ne("")
    frame["FBrf_list"] = frame["reference"].apply(split_fbrfs)
    frame = frame[
        frame["FBrf_list"].str.len().gt(0) & frame["keyword_hit"]
    ].copy()

    frame["Allele ID"] = frame["genotype_FBids"].apply(
        lambda value: [
            component_id
            for component_id in _extract_flybase_ids(value)
            if component_id.startswith("FBal")
        ]
    )
    frame = frame.explode("Allele ID")
    frame = frame[frame["Allele ID"].fillna("").astype(str).ne("")].copy()
    frame["Reference (FBrf)"] = frame["FBrf_list"]
    frame = frame.explode("Reference (FBrf)")

    mapping = fbal_to_gene_df.copy()
    for column in ("AlleleID", "AlleleSymbol", "GeneID", "GeneSymbol"):
        if column not in mapping.columns:
            mapping[column] = ""
        mapping[column] = mapping[column].fillna("").astype(str).str.strip()
    mapping = mapping.drop_duplicates(subset=["AlleleID"])
    mapping = mapping.rename(
        columns={
            "AlleleID": "Allele ID",
            "AlleleSymbol": "Allele symbol",
            "GeneID": "flybase_gene_id",
            "GeneSymbol": "Gene",
        }
    )
    frame = frame.merge(mapping, on="Allele ID", how="left", validate="many_to_one")
    frame = frame[
        frame["flybase_gene_id"].fillna("").astype(str).str.startswith("FBgn")
    ].copy()
    frame["is_allele_or_insertion"] = ~frame["Allele symbol"].apply(is_uas_or_rnai_reagent)
    frame = frame[frame["is_allele_or_insertion"]].copy()
    if dmel_gene_catalog is not None:
        frame = frame.merge(
            dmel_gene_catalog,
            on="flybase_gene_id",
            how="inner",
            validate="many_to_one",
        )
        current_symbols = frame["current_gene_symbol"].fillna("").astype(str).str.strip()
        frame["Gene"] = current_symbols.where(current_symbols.ne(""), frame["Gene"])
        frame = frame.drop(columns=["current_gene_symbol"])
    return frame.reset_index(drop=True)


def stock_center_label(collection: Any, custom_stock: Any) -> str:
    """Return a display stock-center label: BDSC, Vienna, ..., or Custom."""
    if bool(custom_stock):
        return "Custom"
    text = str(collection or "").strip()
    lowered = text.lower()
    if "bloomington" in lowered:
        return "BDSC"
    if "vienna" in lowered:
        return "Vienna"
    return text or "Unknown"


def attach_reagents(
    assertions: pd.DataFrame,
    derived_stock_components: pd.DataFrame,
) -> pd.DataFrame:
    """Attach orderable stocks and synthesize custom rows for unstocked alleles."""
    derived = derived_stock_components.copy()
    for column in (
        "FBst",
        "stock_number",
        "collection",
        "FB_genotype",
        "derived_stock_component",
    ):
        if column not in derived.columns:
            derived[column] = ""
        derived[column] = derived[column].fillna("").astype(str).str.strip()

    derived = derived.drop_duplicates(
        subset=[
            "FBst",
            "stock_number",
            "collection",
            "FB_genotype",
            "derived_stock_component",
        ]
    )
    stock_backed_alleles = set(
        derived.loc[
            derived["derived_stock_component"].str.startswith("FBal"),
            "derived_stock_component",
        ]
    )
    included_alleles = set(assertions["Allele ID"].dropna().astype(str))
    stock_lookup = derived[
        derived["derived_stock_component"].isin(included_alleles)
    ][
        [
            "FBst",
            "stock_number",
            "collection",
            "FB_genotype",
            "derived_stock_component",
        ]
    ].rename(
        columns={
            "stock_number": "Stock #",
            "collection": "Collection",
            "FB_genotype": "Genotype",
            "derived_stock_component": "Allele ID",
        }
    )
    stock_rows = assertions.merge(stock_lookup, on="Allele ID", how="inner")
    stock_rows["custom_stock"] = False

    custom_rows = assertions[
        ~assertions["Allele ID"].isin(stock_backed_alleles)
    ].copy()
    if len(custom_rows):
        custom_rows["FBst"] = ""
        custom_rows["Stock #"] = custom_rows["Allele symbol"]
        custom_rows["Collection"] = "Custom phenotype reagent"
        custom_rows["Genotype"] = custom_rows["genotype_symbols"]
        custom_rows["custom_stock"] = True

    combined = pd.concat([stock_rows, custom_rows], ignore_index=True, sort=False)
    combined["Stock Center"] = combined.apply(
        lambda row: stock_center_label(row["Collection"], row["custom_stock"]),
        axis=1,
    )
    combined["reagent_key"] = combined.apply(
        lambda row: (
            clean_id(row.get("FBst", ""))
            or f"custom:{str(row.get('Allele ID', '')).strip()}"
        ),
        axis=1,
    )
    return combined


def exclude_genes(
    assertions: pd.DataFrame,
    excluded_gene_ids: Mapping[str, str] = EXCLUDED_GENES,
) -> pd.DataFrame:
    """Remove explicitly blacklisted genes (by FBgn id AND symbol) before any
    downstream reagent/reference expansion, so a blacklisted gene can never
    reappear via a stale/alternate symbol.
    """
    if not excluded_gene_ids or assertions.empty:
        return assertions.copy()
    excluded_ids = {
        str(gene_id).strip() for gene_id in excluded_gene_ids if str(gene_id).strip()
    }
    excluded_symbols = {
        str(symbol).strip().lower()
        for symbol in excluded_gene_ids.values()
        if str(symbol).strip()
    }
    gene_ids = assertions["flybase_gene_id"].fillna("").astype(str)
    gene_symbols = assertions.get("Gene", pd.Series("", index=assertions.index))
    gene_symbols = gene_symbols.fillna("").astype(str).str.strip().str.lower()
    return assertions[
        ~gene_ids.isin(excluded_ids) & ~gene_symbols.isin(excluded_symbols)
    ].copy()


def deduplicate_reagent_rows(rows: pd.DataFrame) -> pd.DataFrame:
    """Collapse to unique reagent -> phenotype -> reference rows."""
    if rows.empty:
        return pd.DataFrame(columns=REAGENT_COLUMNS + ["reagent_key"])
    key = ["reagent_key", "phenotype_name", "Reference (FBrf)"]
    aggregations: Dict[str, Any] = {
        "Gene": ordered_unique,
        "flybase_gene_id": ordered_unique,
        "Collection": ordered_unique,
        "Stock Center": ordered_unique,
        "Stock #": ordered_unique,
        "FBst": ordered_unique,
        "custom_stock": "max",
        "Genotype": ordered_unique,
        "Allele symbol": ordered_unique,
        "Allele ID": ordered_unique,
        "qualifier_names": ordered_unique,
        "keyword_hit": "max",
        "keyword_match_terms": ordered_unique,
    }
    result = rows.groupby(key, sort=False, as_index=False).agg(aggregations)
    result = result.rename(
        columns={
            "phenotype_name": "Phenotype",
            "qualifier_names": "Qualifier",
        }
    )
    return result


def load_reference_map(refs_path: Path) -> pd.DataFrame:
    """Load one normalized FBrf-to-identifier row per FlyBase reference."""
    refs = load_flybase_tsv(refs_path)
    for column in ("FBrf", "PMID", "PMCID", "DOI", "miniref"):
        if column not in refs.columns:
            refs[column] = ""
        refs[column] = refs[column].fillna("").astype(str).str.strip()
    refs["FBrf"] = refs["FBrf"].apply(clean_id)
    refs["PMID"] = refs["PMID"].apply(clean_id)
    return refs[["FBrf", "PMID", "PMCID", "DOI", "miniref"]].drop_duplicates(
        subset=["FBrf"], keep="first"
    )


def enrich_references(
    reagent_rows: pd.DataFrame,
    reference_map: pd.DataFrame,
    pubmed_cache: PubMedCache,
    pubmed_client: PubMedClient,
    fetch_pubmed: bool = True,
) -> Tuple[pd.DataFrame, Dict[str, int]]:
    """Attach PMID and publication metadata, fetching incomplete PMIDs when allowed."""
    rows = reagent_rows.merge(
        reference_map,
        left_on="Reference (FBrf)",
        right_on="FBrf",
        how="left",
        validate="many_to_one",
    )
    rows["PMID"] = rows["PMID"].fillna("").astype(str).apply(clean_id)
    cache_keys_before = set(pubmed_cache.load())
    pmids = list(dict.fromkeys(p for p in rows["PMID"].tolist() if p))
    if fetch_pubmed:
        metadata = pubmed_client.fetch_author_date_metadata(pmids)
    else:
        metadata = {}
        for pmid in pmids:
            cached = pubmed_cache.get(pmid) or {}
            authors = cached.get("authors", [])
            if isinstance(authors, str):
                authors = [item.strip() for item in authors.split(";") if item.strip()]
            metadata[pmid] = {
                "title": str(cached.get("title", "") or ""),
                "all_authors": "; ".join(authors),
                "date_published": str(cached.get("year", "") or ""),
                "journal": str(cached.get("journal", "") or ""),
            }

    titles: List[str] = []
    authors_values: List[str] = []
    dates: List[str] = []
    journals: List[str] = []
    for _, row in rows.iterrows():
        fallback = parse_miniref(row.get("miniref", ""))
        meta = metadata.get(str(row.get("PMID", "")), {})
        titles.append(str(meta.get("title", "") or fallback["title"]))
        authors_values.append(str(meta.get("all_authors", "") or fallback["authors"]))
        dates.append(str(meta.get("date_published", "") or fallback["year"]))
        journals.append(str(meta.get("journal", "") or fallback["journal"]))
    rows["Reference Title"] = titles
    rows["Reference Authors"] = authors_values
    rows["Date of publication"] = dates
    rows["Reference Journal"] = journals
    cache_after = pubmed_cache.load()
    stats = {
        "unique_pmids": len(pmids),
        "pmids_cached_before": sum(pmid in cache_keys_before for pmid in pmids),
        "pmids_with_metadata": sum(bool(metadata.get(pmid)) for pmid in pmids),
        "cache_entries_after": len(cache_after),
    }
    return rows, stats


def sort_reagent_rows(rows: pd.DataFrame) -> pd.DataFrame:
    """Group by gene count, then order alphabetically inside each gene."""
    rows = rows.copy()
    counts = rows.groupby("Gene", dropna=False)["reagent_key"].nunique()
    rows["n_reagents_for_gene"] = rows["Gene"].map(counts).fillna(0).astype(int)
    rows["_stock_sort"] = rows["Stock #"].astype(str).str.lower()
    rows = rows.sort_values(
        by=[
            "n_reagents_for_gene",
            "Gene",
            "Phenotype",
            "Collection",
            "_stock_sort",
        ],
        ascending=[False, True, True, True, True],
        kind="mergesort",
        na_position="last",
    ).drop(columns=["_stock_sort"])
    return rows.reset_index(drop=True)


def build_genes_sheet(reagents: pd.DataFrame) -> pd.DataFrame:
    """Create one summary row per displayed gene group."""
    records = []
    for gene, group in reagents.groupby("Gene", sort=False, dropna=False):
        phenotype_pmids = ordered_unique(
            group.apply(
                lambda row: (
                    f"{row['Phenotype']} (PMID: {row['PMID']})"
                    if str(row.get("PMID", "")).strip()
                    else (
                        f"{row['Phenotype']} "
                        f"(PMID: unavailable; FBrf: {row['Reference (FBrf)']})"
                    )
                ),
                axis=1,
            )
        )
        records.append(
            {
                "Gene": gene,
                "flybase_gene_id": ordered_unique(group["flybase_gene_id"]),
                "n_reagents": int(group["reagent_key"].nunique()),
                "n_phenotype_refs": int(
                    group[["Phenotype", "Reference (FBrf)"]].drop_duplicates().shape[0]
                ),
                "keyword_hit": bool(group["keyword_hit"].any()),
                "Phenotype (PMID)": phenotype_pmids,
            }
        )
    return pd.DataFrame(records)


def build_references_sheet(reagents: pd.DataFrame) -> pd.DataFrame:
    """Build unique reference -> reagent rows in first encounter order."""
    frame = reagents.reset_index(drop=True).copy()
    frame["_encounter_order"] = frame.index
    frame["_reference_key"] = frame.apply(
        lambda row: str(row.get("PMID", "")).strip()
        or str(row.get("Reference (FBrf)", "")).strip(),
        axis=1,
    )
    records = []
    for (_, _), group in frame.groupby(
        ["_reference_key", "reagent_key"], sort=False, dropna=False
    ):
        first = group.iloc[0]
        phenotypes = ordered_unique(group["Phenotype"])
        center = stock_center_label(first["Collection"], first["custom_stock"])
        stock_number = str(first["Stock #"]).strip()
        records.append(
            {
                "_encounter_order": int(group["_encounter_order"].min()),
                "PMID": first["PMID"],
                "Title": first["Reference Title"],
                "Authors": first["Reference Authors"],
                "Date of publication": first["Date of publication"],
                "Journal": first["Reference Journal"],
                "Associated gene": ordered_unique(group["Gene"]),
                "Associated reagent": f"{stock_number}, {phenotypes}, {center}",
                "FBrf": ordered_unique(group["Reference (FBrf)"]),
                "Stock #": stock_number,
                "Collection": first["Collection"],
                "Phenotype": phenotypes,
                "custom_stock": bool(first["custom_stock"]),
            }
        )
    references = pd.DataFrame(records).sort_values(
        "_encounter_order", kind="mergesort"
    )
    return references.drop(columns=["_encounter_order"])[REFERENCE_COLUMNS].reset_index(
        drop=True
    )


def build_summary_sheet(
    reagents: pd.DataFrame,
    references: pd.DataFrame,
    genes: pd.DataFrame,
    source_paths: Mapping[str, Path],
    pubmed_stats: Mapping[str, int],
) -> pd.DataFrame:
    """Create visible assumptions, provenance, and result counts."""
    rows = [
        ("Workbook purpose", "Published Drosophila temperature-related phenotypic reagents"),
        (
            "Inclusion rule",
            "Cited phenotype AND word-boundary keyword hit AND allele/insertion "
            "(not UAS/RNAi) reagent",
        ),
        ("Keyword targets", "Temperature; heat; cold; thermo"),
        ("Keyword matching", "temperature/heat/cold use word boundaries; thermo uses a prefix match"),
        (
            "UAS/RNAi exclusion",
            "Word-boundary 'UAS'/'RNAi' or VDRC knockdown allele brackets "
            "(GD/KK/VSH/HMS/JF) in the allele symbol",
        ),
        ("Reagent uniqueness", "(FBst or custom allele key, phenotype, FBrf)"),
        ("References uniqueness", "(PMID if available else FBrf, reagent key)"),
        (
            "Excluded genes",
            "; ".join(
                f"{symbol} ({gene_id})" for gene_id, symbol in EXCLUDED_GENES.items()
            ),
        ),
        ("Reagent rows", len(reagents)),
        ("Reference-reagent rows", len(references)),
        ("Unique genes/groups", len(genes)),
        ("Distinct reagents", int(reagents["reagent_key"].nunique())),
        ("Stock-center reagents", int(reagents.loc[~reagents["custom_stock"], "reagent_key"].nunique())),
        ("Custom reagents", int(reagents.loc[reagents["custom_stock"], "reagent_key"].nunique())),
        ("Keyword-hit rows", int(reagents["keyword_hit"].sum())),
        ("Unique PMIDs", int(pubmed_stats.get("unique_pmids", 0))),
        ("PMIDs cached before", int(pubmed_stats.get("pmids_cached_before", 0))),
        ("PMIDs with metadata", int(pubmed_stats.get("pmids_with_metadata", 0))),
    ]
    rows.extend((f"Source: {label}", path.name) for label, path in source_paths.items())
    return pd.DataFrame(rows, columns=["Item", "Value"])


def _style_sheet(ws, *, max_width: int = 60) -> None:
    """Apply a compact professional Excel style."""
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    if ws.max_row and ws.max_column:
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 32
    for column_index, cells in enumerate(ws.iter_cols(), start=1):
        values = [str(cell.value or "") for cell in list(cells)[:500]]
        width = min(max(max(map(len, values), default=0) + 2, 11), max_width)
        ws.column_dimensions[get_column_letter(column_index)].width = width


def write_temperature_workbook(
    output_path: Path,
    reagents: pd.DataFrame,
    references: pd.DataFrame,
    genes: pd.DataFrame,
    summary: pd.DataFrame,
) -> None:
    """Write the final workbook with References as the second sheet."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    display_reagents = reagents.copy()
    for column in REAGENT_COLUMNS:
        if column not in display_reagents.columns:
            display_reagents[column] = ""
    display_reagents = display_reagents[REAGENT_COLUMNS]

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        display_reagents.to_excel(writer, sheet_name="Reagents", index=False)
        references.to_excel(writer, sheet_name="References", index=False)
        genes.to_excel(writer, sheet_name="Genes", index=False)
        summary.to_excel(writer, sheet_name="Summary", index=False)

        workbook = writer.book
        for worksheet in workbook.worksheets:
            _style_sheet(worksheet)

        for worksheet_name in ("Reagents", "References"):
            ws = workbook[worksheet_name]
            header_to_column = {cell.value: cell.column for cell in ws[1]}
            pmid_column = header_to_column.get("PMID")
            if pmid_column:
                for row_index in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row_index, column=pmid_column)
                    pmid = clean_id(cell.value)
                    if pmid:
                        cell.hyperlink = f"https://pubmed.ncbi.nlm.nih.gov/{pmid}/"
                        cell.style = "Hyperlink"
                        cell.font = Font(name="Arial", size=10, color="0563C1", underline="single")


def run_export(
    output_path: Path,
    fetch_pubmed: bool = True,
    settings: Optional[Settings] = None,
) -> Dict[str, int]:
    """Execute the TSV -> keyword-match -> stock -> workbook pipeline."""
    settings = settings or Settings()

    alleles_dir = settings.flybase_alleles_stocks_path
    refs_dir = settings.flybase_references_path
    phenotype_path = find_latest_tsv(alleles_dir, "genotype_phenotype_data")
    fbal_path = find_latest_tsv(alleles_dir, "fbal_to_fbgn")
    synonym_path = find_latest_tsv(settings.flybase_data_path / "genes", "fb_synonym")
    refs_path = find_latest_tsv(refs_dir, "fbrf_pmid_pmcid_doi")
    derived_path = Path(settings.flybase_derived_stock_components_path)
    source_paths = {
        "phenotypes": phenotype_path,
        "allele-to-gene": fbal_path,
        "Dmel gene catalog": synonym_path,
        "stock components": derived_path,
        "reference identifiers": refs_path,
    }

    print(f"Loading FlyBase phenotypes: {phenotype_path.name}")
    phenotype_df = load_flybase_tsv(phenotype_path)
    fbal_df = load_flybase_tsv(fbal_path)
    dmel_gene_catalog = load_dmel_gene_catalog(synonym_path)

    assertions = build_included_assertions(
        phenotype_df,
        fbal_df,
        dmel_gene_catalog=dmel_gene_catalog,
    )
    assertions_before_exclusion = len(assertions)
    assertions = exclude_genes(assertions)
    excluded_assertions = assertions_before_exclusion - len(assertions)
    if excluded_assertions:
        excluded_labels = ", ".join(
            f"{symbol} ({gene_id})" for gene_id, symbol in EXCLUDED_GENES.items()
        )
        print(
            f"Excluded {excluded_assertions:,} phenotype assertions for "
            f"{excluded_labels}"
        )
    print(f"Included {len(assertions):,} allele/phenotype/reference assertions")
    derived_df = pd.read_csv(derived_path, dtype=str, keep_default_na=False)
    reagent_rows = attach_reagents(assertions, derived_df)
    reagent_rows = deduplicate_reagent_rows(reagent_rows)

    reference_map = load_reference_map(refs_path)
    pubmed_cache = PubMedCache(Path(settings.pubmed_cache_path))
    pubmed_client = PubMedClient(
        cache=pubmed_cache,
        email=DEFAULT_CONTACT_EMAIL,
        api_key=settings.ncbi_api_key,
        batch_size=settings.batch_size,
    )
    reagent_rows, pubmed_stats = enrich_references(
        reagent_rows,
        reference_map,
        pubmed_cache,
        pubmed_client,
        fetch_pubmed=fetch_pubmed,
    )
    reagent_rows = sort_reagent_rows(reagent_rows)
    # Defense-in-depth: re-apply the gene blacklist to the fully-assembled
    # reagent rows so a blacklisted gene's reagents/references can never
    # reach the workbook, even if it re-entered via a later join/fallback.
    reagent_rows = exclude_genes(reagent_rows)
    references = build_references_sheet(reagent_rows)
    genes = build_genes_sheet(reagent_rows)
    summary = build_summary_sheet(
        reagent_rows,
        references,
        genes,
        source_paths,
        pubmed_stats,
    )
    write_temperature_workbook(output_path, reagent_rows, references, genes, summary)
    print(
        f"Wrote {output_path} ({len(reagent_rows):,} reagent rows, "
        f"{len(references):,} reference-reagent rows, {len(genes):,} gene groups)"
    )
    return {
        "reagent_rows": len(reagent_rows),
        "reference_rows": len(references),
        "genes": len(genes),
    }


def main() -> int:
    args = build_parser().parse_args()
    run_export(
        args.output,
        fetch_pubmed=not args.skip_pubmed_fetch,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
