#!/usr/bin/env python3
"""
Summarize organized stocker workbook counts by gene set and combination.

The stocker writes one organized workbook per input gene set. Each workbook's
Contents sheet includes a "Prioritized sheet breakdown" table with one row per
JSON config combination and the workbook-native # Stocks / # Alleles / # Genes
counts. This script aggregates those tables across a directory tree.
"""

from __future__ import annotations

import argparse
import json
import math
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, Font, Side


SUMMARY_COLUMNS = [
    "relative_path",
    "gene_set",
    "combination",
    "num_genes",
    "num_stocks",
    "num_alleles",
    "sheet_name",
    "workbook_path",
]


@dataclass(frozen=True)
class CombinationCount:
    relative_path: str
    gene_set: str
    combination: str
    num_genes: int
    num_stocks: int
    num_alleles: int
    sheet_name: str
    workbook_path: Path


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Scan organized stocker workbooks and summarize # genes, # stocks, "
            "and # alleles by input gene set and JSON config combination."
        )
    )
    parser.add_argument(
        "input_dir",
        type=Path,
        help="Directory to scan recursively, e.g. a Per Gene Set Runs directory.",
    )
    parser.add_argument(
        "--config",
        type=Path,
        default=None,
        help=(
            "Optional stock-splitting JSON config. When provided, combinations "
            "are emitted in config order and missing workbook rows are filled "
            "with zero counts."
        ),
    )
    parser.add_argument(
        "--output-name",
        default="combination_counts_summary.xlsx",
        help="Summary workbook filename to write into input_dir.",
    )
    parser.add_argument(
        "--csv-output-name",
        default=None,
        help=(
            "Optional CSV filename. Defaults to the output workbook stem with "
            ".csv appended."
        ),
    )
    return parser


def _normalize_cell(value: object) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except TypeError:
        pass
    text = str(value).strip()
    return "" if text.lower() == "nan" else text


def _coerce_non_negative_int(value: object, context: str) -> int:
    numeric = pd.to_numeric(pd.Series([value]), errors="coerce").iloc[0]
    if pd.isna(numeric):
        raise ValueError(f"{context} is not numeric: {value!r}")
    numeric_float = float(numeric)
    numeric_int = int(round(numeric_float))
    if (
        not math.isfinite(numeric_float)
        or numeric_float < 0
        or not math.isclose(numeric_float, numeric_int, rel_tol=0.0, abs_tol=1e-9)
    ):
        raise ValueError(f"{context} must be a non-negative integer: {value!r}")
    return numeric_int


def _load_config_combinations(config_path: Optional[Path]) -> Optional[List[str]]:
    if config_path is None:
        return None
    with Path(config_path).open("r", encoding="utf-8") as f:
        config = json.load(f)
    combinations = config.get("combinations", [])
    labels: List[str] = []
    for combo in combinations:
        if not isinstance(combo, list):
            raise ValueError(f"Config combination is not a list: {combo!r}")
        labels.append(" >> ".join(str(part) for part in combo))
    return labels


def _iter_rows(contents_df: pd.DataFrame) -> Iterable[List[str]]:
    for _, row in contents_df.iterrows():
        yield [_normalize_cell(value) for value in row.tolist()]


def _find_breakdown_header(contents_df: pd.DataFrame) -> Optional[tuple[int, Dict[str, int]]]:
    required = {"Sheet criteria", "# Stocks", "# Alleles", "# Genes"}
    rows = list(_iter_rows(contents_df))
    start_idx = 0
    for row_idx, row in enumerate(rows):
        if row and row[0] == "Prioritized sheet breakdown":
            start_idx = row_idx + 1
            break

    for row_idx, row in enumerate(rows[start_idx:], start=start_idx):
        row_positions = {value: idx for idx, value in enumerate(row)}
        if required.issubset(row_positions):
            return row_idx, row_positions
    return None


def _derive_gene_set_name(workbook_path: Path) -> str:
    stem = workbook_path.stem
    suffix = "_aggregated"
    if stem.endswith(suffix):
        return stem[: -len(suffix)]
    return stem


def _derive_relative_path(workbook_path: Path) -> str:
    """Return the input file's folder path relative to the input folder.

    Runs are staged at ``<root>/Per Gene Set Runs/<rel...>/<gene_set_run>/
    Stocks/[Organized Stocks/]<file>.xlsx``, where ``<rel...>`` mirrors the input
    CSV's folders under the input directory. The parts between ``Per Gene Set
    Runs`` and the gene-set run directory are the input file's relative folder
    path (using "/" as the separator). For example, an input file at
    ``Genes/A/B/C/set.csv`` yields ``A/B/C``. Returns "." for a top-level file or
    when ``Per Gene Set Runs`` is absent from the path.

    The gene-set run directory is located via its ``Stocks`` subfolder rather
    than by counting fixed trailing components, so this is layout-agnostic: it
    works whether the workbook still lives under ``Stocks/Organized Stocks/`` or
    has been flattened up into ``Stocks/``.
    """
    parts = workbook_path.parts
    try:
        runs_idx = parts.index("Per Gene Set Runs")
    except ValueError:
        return "."
    rel = parts[runs_idx + 1 :]
    # rel == (<rel...>, <gene_set_run>, "Stocks"[, "Organized Stocks"], <file>).
    # Use the last "Stocks" occurrence so a stray "Stocks" in <rel...> can't
    # shadow the gene-set run's own Stocks folder; <gene_set_run> sits directly
    # before it, and everything earlier is the input file's relative folder path.
    stocks_positions = [i for i, part in enumerate(rel) if part == "Stocks"]
    if not stocks_positions:
        return "."
    gene_set_run_idx = stocks_positions[-1] - 1
    folder_parts = rel[:gene_set_run_idx]
    return "/".join(folder_parts) if folder_parts else "."


def parse_workbook_counts(
    workbook_path: Path,
    *,
    config_combinations: Optional[Sequence[str]] = None,
    relative_path: Optional[str] = None,
) -> List[CombinationCount]:
    try:
        contents_df = pd.read_excel(workbook_path, sheet_name="Contents", header=None)
    except ValueError as exc:
        raise ValueError("missing Contents sheet") from exc
    except Exception as exc:
        raise ValueError(f"could not read Contents sheet: {exc}") from exc

    header = _find_breakdown_header(contents_df)
    if header is None:
        raise ValueError("missing prioritized sheet breakdown header")
    header_idx, positions = header

    combo_idx = positions["Sheet criteria"]
    stocks_idx = positions["# Stocks"]
    alleles_idx = positions["# Alleles"]
    genes_idx = positions["# Genes"]
    sheet_idx = positions.get("Sheet Name")

    gene_set = _derive_gene_set_name(workbook_path)
    if relative_path is None:
        relative_path = _derive_relative_path(workbook_path)
    parsed_by_combo: Dict[str, CombinationCount] = {}
    parsed_order: List[str] = []

    for row_num, row in enumerate(_iter_rows(contents_df.iloc[header_idx + 1 :]), start=header_idx + 2):
        if combo_idx >= len(row):
            continue
        combination = row[combo_idx]
        if not combination:
            continue
        if combination.startswith("Total"):
            break

        count = CombinationCount(
            relative_path=relative_path,
            gene_set=gene_set,
            combination=combination,
            num_genes=_coerce_non_negative_int(
                row[genes_idx] if genes_idx < len(row) else "",
                f"# Genes for {combination!r} in {workbook_path} row {row_num}",
            ),
            num_stocks=_coerce_non_negative_int(
                row[stocks_idx] if stocks_idx < len(row) else "",
                f"# Stocks for {combination!r} in {workbook_path} row {row_num}",
            ),
            num_alleles=_coerce_non_negative_int(
                row[alleles_idx] if alleles_idx < len(row) else "",
                f"# Alleles for {combination!r} in {workbook_path} row {row_num}",
            ),
            sheet_name=row[sheet_idx] if sheet_idx is not None and sheet_idx < len(row) else "",
            workbook_path=workbook_path,
        )
        parsed_by_combo[combination] = count
        parsed_order.append(combination)

    if not parsed_by_combo:
        raise ValueError("prioritized sheet breakdown contained no combination rows")

    ordered_combinations = list(config_combinations) if config_combinations is not None else parsed_order
    counts: List[CombinationCount] = []
    for combination in ordered_combinations:
        if combination in parsed_by_combo:
            counts.append(parsed_by_combo[combination])
        else:
            counts.append(
                CombinationCount(
                    relative_path=relative_path,
                    gene_set=gene_set,
                    combination=combination,
                    num_genes=0,
                    num_stocks=0,
                    num_alleles=0,
                    sheet_name="-",
                    workbook_path=workbook_path,
                )
            )
    return counts


def discover_organized_workbooks(input_dir: Path, output_path: Path) -> List[Path]:
    workbooks: List[Path] = []
    for path in sorted(input_dir.rglob("*_aggregated.xlsx")):
        if path.name.startswith("~$") or path.name.startswith("."):
            continue
        if path.resolve() == output_path.resolve():
            continue
        if path.name.endswith("_aggregated_similarity_tiers.xlsx"):
            continue
        if "_aggregated_similarity" in str(path):
            continue
        workbooks.append(path)
    return workbooks


def counts_to_dataframe(counts: Sequence[CombinationCount]) -> pd.DataFrame:
    rows = [
        {
            "relative_path": count.relative_path,
            "gene_set": count.gene_set,
            "combination": count.combination,
            "num_genes": count.num_genes,
            "num_stocks": count.num_stocks,
            "num_alleles": count.num_alleles,
            "sheet_name": count.sheet_name,
            "workbook_path": str(count.workbook_path),
        }
        for count in counts
    ]
    return pd.DataFrame(rows, columns=SUMMARY_COLUMNS)


def build_totals_dataframe(summary_df: pd.DataFrame) -> pd.DataFrame:
    if summary_df.empty:
        return pd.DataFrame(
            columns=[
                "relative_path",
                "combination",
                "num_genes",
                "num_stocks",
                "num_alleles",
            ]
        )
    return (
        summary_df.groupby(
            ["relative_path", "combination"], sort=False
        )[["num_genes", "num_stocks", "num_alleles"]]
        .sum()
        .reset_index()
    )


def _natural_sort_key(text: object) -> str:
    """Return a lexicographically sortable key that orders numbers naturally.

    Digit runs are zero-padded to a fixed width so that, e.g.,
    ``frequency_4`` sorts before ``frequency_5`` and ``frequency_9`` before
    ``frequency_10`` (rather than lexical order, which would place ``10`` before
    ``2``).
    """
    return re.sub(r"\d+", lambda match: match.group().zfill(12), str(text))


def _sort_summary_rows(
    summary_df: pd.DataFrame,
    config_combinations: Optional[Sequence[str]],
) -> pd.DataFrame:
    """Order rows by folder, then gene set (natural sort), then combination.

    Folders and the files within each folder are ordered with a natural
    (numeric-aware) sort, so ``..._frequency_5...`` follows ``..._frequency_4...``
    and ``..._frequency_10...`` follows ``..._frequency_9...``. Combination order
    follows the JSON config when provided, otherwise the order combinations
    first appear in the data. The sort is stable so the per-workbook ordering is
    preserved within each group.
    """
    if summary_df.empty:
        return summary_df
    if config_combinations:
        order = {combo: idx for idx, combo in enumerate(config_combinations)}
    else:
        order = {
            combo: idx
            for idx, combo in enumerate(dict.fromkeys(summary_df["combination"]))
        }
    combo_rank = summary_df["combination"].map(lambda c: order.get(c, len(order)))
    return (
        summary_df.assign(
            _rel_key=summary_df["relative_path"].map(_natural_sort_key),
            _gene_key=summary_df["gene_set"].map(_natural_sort_key),
            _combo_rank=combo_rank,
        )
        .sort_values(
            by=["_rel_key", "_gene_key", "_combo_rank"],
            kind="stable",
        )
        .drop(columns=["_rel_key", "_gene_key", "_combo_rank"])
        .reset_index(drop=True)
    )


def _autosize_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        values = ["" if cell.value is None else str(cell.value) for cell in column_cells]
        width = max((len(value) for value in values), default=10) + 2
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(width, 12), 80)


def _draw_group_borders(worksheet, summary_df: pd.DataFrame) -> None:
    """Draw horizontal lines that group rows by relative_path and gene_set.

    A medium top border starts each new ``relative_path`` (input folder) group;
    a thin top border starts each new ``gene_set`` (the output block derived from
    one input file) within a group. Rows are assumed pre-sorted so each group is
    contiguous.
    """
    if summary_df.empty:
        return
    medium = Side(style="medium")
    thin = Side(style="thin")
    num_cols = len(summary_df.columns)
    rel_values = summary_df["relative_path"].tolist()
    gene_values = summary_df["gene_set"].tolist()

    prev_rel = None
    prev_gene = None
    for offset in range(len(summary_df)):
        # Data rows start at worksheet row 2 (row 1 is the header).
        excel_row = offset + 2
        rel = rel_values[offset]
        gene = gene_values[offset]
        new_group = rel != prev_rel
        new_file = new_group or gene != prev_gene
        if offset > 0 and new_file:
            style = medium if new_group else thin
            for col in range(1, num_cols + 1):
                cell = worksheet.cell(row=excel_row, column=col)
                existing = cell.border
                cell.border = Border(
                    left=existing.left,
                    right=existing.right,
                    bottom=existing.bottom,
                    top=style,
                )
        prev_rel = rel
        prev_gene = gene


def write_summary_outputs(
    *,
    output_path: Path,
    csv_output_path: Path,
    summary_df: pd.DataFrame,
    totals_df: pd.DataFrame,
    warnings: Sequence[str],
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    csv_output_path.parent.mkdir(parents=True, exist_ok=True)

    summary_df.to_csv(csv_output_path, index=False)

    workbook = Workbook()
    workbook.remove(workbook.active)
    header_font = Font(bold=True)

    for sheet_name, df in (
        ("Combination Counts", summary_df),
        ("Totals by Combination", totals_df),
        ("Warnings", pd.DataFrame({"warning": list(warnings)})),
    ):
        worksheet = workbook.create_sheet(sheet_name)
        worksheet.append(list(df.columns))
        for cell in worksheet[1]:
            cell.font = header_font
        for row in df.itertuples(index=False, name=None):
            worksheet.append(list(row))
        if sheet_name == "Combination Counts":
            _draw_group_borders(worksheet, df)
        _autosize_columns(worksheet)

    workbook.save(output_path)


def summarize_combination_counts(
    input_dir: Path,
    *,
    config_path: Optional[Path] = None,
    output_name: str = "combination_counts_summary.xlsx",
    csv_output_name: Optional[str] = None,
) -> int:
    """Scan organized workbooks under ``input_dir`` and write the summary.

    Importable entry point shared by the CLI script and the ``run`` pipeline so
    the summary can be generated automatically at the end of a run.
    """
    input_dir = Path(input_dir).resolve()
    if not input_dir.exists():
        raise SystemExit(f"Input directory does not exist: {input_dir}")
    if not input_dir.is_dir():
        raise SystemExit(f"Input path is not a directory: {input_dir}")

    output_path = input_dir / output_name
    if output_path.suffix.lower() != ".xlsx":
        output_path = output_path.with_suffix(".xlsx")
    resolved_csv_name = csv_output_name or f"{output_path.stem}.csv"
    csv_output_path = input_dir / resolved_csv_name

    config_combinations = _load_config_combinations(config_path)
    workbooks = discover_organized_workbooks(input_dir, output_path)

    warnings: List[str] = []
    all_counts: List[CombinationCount] = []
    for workbook_path in workbooks:
        try:
            all_counts.extend(
                parse_workbook_counts(
                    workbook_path,
                    config_combinations=config_combinations,
                    relative_path=_derive_relative_path(workbook_path),
                )
            )
        except ValueError as exc:
            warnings.append(f"Skipping {workbook_path}: {exc}")

    if not all_counts:
        print(f"No organized *_aggregated.xlsx workbooks found under {input_dir}")
        for warning in warnings:
            print(warning)
        return 0

    summary_df = _sort_summary_rows(
        counts_to_dataframe(all_counts), config_combinations
    )
    totals_df = build_totals_dataframe(summary_df)
    write_summary_outputs(
        output_path=output_path,
        csv_output_path=csv_output_path,
        summary_df=summary_df,
        totals_df=totals_df,
        warnings=warnings,
    )

    for warning in warnings:
        print(warning)
    print(f"Saved summary workbook: {output_path}")
    print(f"Saved summary CSV: {csv_output_path}")
    print(f"  Workbooks parsed: {summary_df['workbook_path'].nunique()}")
    print(f"  Rows written: {len(summary_df)}")
    return 0


def main(argv: Optional[Sequence[str]] = None) -> int:
    args = build_parser().parse_args(argv)
    return summarize_combination_counts(
        args.input_dir,
        config_path=args.config,
        output_name=args.output_name,
        csv_output_name=args.csv_output_name,
    )


if __name__ == "__main__":
    raise SystemExit(main())
