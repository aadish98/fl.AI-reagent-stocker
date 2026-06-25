#!/usr/bin/env python3
"""Improve readability of aggregated stock workbooks without changing cell values."""

from __future__ import annotations

import argparse
import re
import sys
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

DATA_SHEET_PREFIXES = ("Sheet",)
DATA_SHEET_NAMES = {
    "References",
    "All Phenotypic Stocks Sheet",
}

WIDE_TEXT_PATTERNS = (
    r"genotype",
    r"phenotype",
    r"title",
    r"author",
    r"journal",
    r"notes",
    r"definition",
    r"rationale",
    r"reference",
    r"provenance",
    r"construct",
    r"component",
    r"balancers",
    r"meaning",
    r"description",
)

NARROW_PATTERNS = (
    r"^FBst$",
    r"^PMID$",
    r"^DOI$",
    r"^Gene$",
    r"^UAS$",
    r"^GAL4$",
    r"^mutant$",
    r"^Other$",
    r"^Qualifier$",
    r"^stock_number$",
    r"^collection$",
    r"^stock_count$",
    r"^num_",
    r"^attP_count$",
    r"^FBti_count$",
    r"^ID #$",
    r"cosine similarity",
    r"max cosine",
    r"relevance score",
    r"^RNAi Type$",
    r"^Data Set$",
    r"^RNAi$",
    r"^RNAi Shorthand$",
    r"^ordered\?",
    r"^Finished",
    r"^Screening Group$",
    r"^# ",
    r"^Sheet Name$",
)


def _header_text(header: object) -> str:
    return str(header or "").strip()


def _is_data_sheet(name: str) -> bool:
    if name in DATA_SHEET_NAMES:
        return True
    return any(name.startswith(prefix) for prefix in DATA_SHEET_PREFIXES)


def _width_bounds(header: str) -> tuple[float, float]:
    lower = header.lower()
    for pattern in NARROW_PATTERNS:
        if re.search(pattern, lower, flags=re.IGNORECASE):
            return 10.0, 18.0
    if re.fullmatch(r"column \d+", lower):
        return 20.0, 40.0
    if "title" in lower or "author" in lower:
        return 24.0, 48.0
    if "genotype" in lower:
        return 24.0, 42.0
    for pattern in WIDE_TEXT_PATTERNS:
        if re.search(pattern, lower, flags=re.IGNORECASE):
            if "score" in lower and "cosine" not in lower and "relevance" in lower:
                return 14.0, 22.0
            return 18.0, 36.0
    if len(header) > 40:
        return 16.0, 28.0
    return 12.0, 24.0


def _longest_display_line(value: object, cap: int = 80) -> int:
    if value is None:
        return 0
    text = str(value).replace("\r\n", "\n").replace("\r", "\n")
    if not text:
        return 0
    return max(min(len(part), cap) for part in text.split("\n"))


def _compute_column_width(
    header: object,
    values: list[object],
    *,
    sample_size: int = 250,
) -> float:
    header_text = _header_text(header)
    min_w, max_w = _width_bounds(header_text)
    sample = values[:sample_size]
    longest = _longest_display_line(header_text, cap=int(max_w))
    for value in sample:
        longest = max(longest, _longest_display_line(value, cap=int(max_w)))
    if header_text and len(header_text) > 24:
        longest = max(longest, min(len(header_text) // 2 + 4, max_w))
    return min(max(longest + 2, min_w), max_w)


def _apply_alignment(cell, *, wrap: bool, vertical: str = "top") -> None:
    alignment = copy(cell.alignment)
    alignment.wrap_text = wrap
    alignment.vertical = vertical or alignment.vertical or "top"
    cell.alignment = alignment


def _estimate_row_height(text: object, col_width: float, *, base: float = 15.0) -> float:
    if text is None or str(text).strip() == "":
        return base
    chars_per_line = max(int(col_width * 1.15), 8)
    lines = 0
    for part in str(text).replace("\r\n", "\n").replace("\r", "\n").split("\n"):
        part_len = len(part)
        lines += max(1, (part_len + chars_per_line - 1) // chars_per_line)
    return min(base * lines, 120.0)


def _format_contents_sheet(ws: Worksheet) -> None:
    ws.column_dimensions["A"].width = 72
    ws.column_dimensions["B"].width = 84
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14

    for row_idx in range(1, ws.max_row + 1):
        row_height = 15.0
        has_content = False
        for col_idx in range(1, 6):
            cell = ws.cell(row_idx, col_idx)
            if cell.value is None:
                continue
            has_content = True
            wrap = col_idx <= 2
            _apply_alignment(cell, wrap=wrap)
            col_width = ws.column_dimensions[get_column_letter(col_idx)].width or 12
            row_height = max(row_height, _estimate_row_height(cell.value, col_width))
        if has_content:
            ws.row_dimensions[row_idx].height = row_height


def _format_data_sheet(ws: Worksheet) -> None:
    max_col = ws.max_column
    max_row = ws.max_row
    if max_col == 0 or max_row == 0:
        return

    headers = [ws.cell(1, col_idx).value for col_idx in range(1, max_col + 1)]
    col_widths: list[float] = []
    for col_idx, header in enumerate(headers, start=1):
        values = [ws.cell(row_idx, col_idx).value for row_idx in range(2, max_row + 1)]
        width = _compute_column_width(header, values)
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = width
        col_widths.append(width)

    header_height = 15.0
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(1, col_idx)
        _apply_alignment(cell, wrap=True, vertical="center")
        header_height = max(
            header_height,
            _estimate_row_height(header, col_widths[col_idx - 1], base=15.0),
        )
    ws.row_dimensions[1].height = max(header_height, 30.0)

    for row_idx in range(2, max_row + 1):
        row_height = 15.0
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row_idx, col_idx)
            _apply_alignment(cell, wrap=True, vertical="top")
            row_height = max(
                row_height,
                _estimate_row_height(cell.value, col_widths[col_idx - 1]),
            )
        ws.row_dimensions[row_idx].height = row_height

    ws.freeze_panes = "A2"
    if max_col >= 1 and max_row >= 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"


def _normalize_cell_value(value: object) -> object:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, int) and not isinstance(value, bool):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def _workbook_values_snapshot(path: Path) -> list[tuple[str, int, int, object]]:
    wb = load_workbook(path, data_only=False)
    snapshot: list[tuple[str, int, int, object]] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                snapshot.append(
                    (sheet_name, cell.row, cell.column, _normalize_cell_value(cell.value))
                )
    wb.close()
    return snapshot


def refactor_workbook_layout(path: Path) -> None:
    before = _workbook_values_snapshot(path)
    wb = load_workbook(path)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if sheet_name == "Contents":
            _format_contents_sheet(ws)
        elif _is_data_sheet(sheet_name):
            _format_data_sheet(ws)
    wb.save(path)
    wb.close()

    after = _workbook_values_snapshot(path)
    if before != after:
        raise RuntimeError(f"Cell values changed after layout refactor: {path}")


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Improve aggregated workbook readability without changing values.",
    )
    parser.add_argument("workbook", type=Path, help="Path to *_aggregated.xlsx")
    args = parser.parse_args()
    path = args.workbook.resolve()
    if not path.exists():
        print(f"ERROR: workbook not found: {path}", file=sys.stderr)
        return 1
    refactor_workbook_layout(path)
    print(f"Refactored layout: {path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
