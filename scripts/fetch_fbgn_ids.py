#!/usr/bin/env python3
"""
Map input gene symbols to FlyBase FBgn IDs.

Never overwrites the user's CSV. Writes sidecar review files instead:

    validated_<original>.csv
    validated_<original>.xlsx   (FBgn cells link to flybase.org/reports/<FBgn>)
    needs-review.csv            (unmatched rows from every input, plus source_file)

Usage:
    python scripts/fetch_fbgn_ids.py <input_directory> [input_gene_col]

Do not run stocker until a human has reviewed the xlsx. If needs-review.csv
has data rows, validate at https://flybase.org/convert/id first.
"""

from __future__ import annotations

import argparse
import gzip
import os
import sys
from glob import glob
from itertools import combinations
from pathlib import Path

import pandas as pd

symbol_to_name = {
    "α": "alpha",
    "β": "beta",
    "γ": "gamma",
    "δ": "delta",  # Lowercase delta
    "ε": "epsilon",
    "ζ": "zeta",
    "η": "eta",
    "θ": "theta",
    "ι": "iota",
    "κ": "kappa",
    "λ": "lambda",
    "μ": "mu",
    "ν": "nu",
    "ξ": "xi",
    "ο": "omicron",
    "π": "pi",
    "ρ": "rho",
    "σ": "sigma",
    "τ": "tau",
    "υ": "upsilon",
    "φ": "phi",
    "χ": "chi",
    "ψ": "psi",
    "ω": "omega",
    "Α": "Alpha",  # Uppercase Alpha
    "Β": "Beta",
    "Γ": "Gamma",
    "Δ": "Delta",  # Uppercase Delta
    "Ε": "Epsilon",
    "Ζ": "Zeta",
    "Η": "Eta",
    "Θ": "Theta",
    "Ι": "Iota",
    "Κ": "Kappa",
    "Λ": "Lambda",
    "Μ": "Mu",
    "Ν": "Nu",
    "Ξ": "Xi",
    "Ο": "Omicron",
    "Π": "Pi",
    "Ρ": "Rho",
    "Σ": "Sigma",
    "Τ": "Tau",
    "Υ": "Upsilon",
    "Φ": "Phi",
    "Χ": "Chi",
    "Ψ": "Psi",
    "Ω": "Omega",
}


_SCRIPT_DIR = Path(__file__).parent.resolve()
_REPO_ROOT = _SCRIPT_DIR.parent
DEFAULT_FLYBASE_DATA = _REPO_ROOT / "data" / "flybase"
VALIDATED_PREFIX = "validated_"
NEEDS_REVIEW_NAME = "needs-review.csv"
FLYBASE_REPORT_URL = "https://flybase.org/reports/{fbgn}"
FLYBASE_CONVERT_URL = "https://flybase.org/convert/id"


def replace_symbol(gene_series):
    for symbol, name in symbol_to_name.items():
        gene_series = gene_series.str.replace(symbol, name, regex=False)
    return gene_series


def create_expanded_mappings(mappings_df):
    """
    Create expanded mappings for each relevant column to ensure all entries are processed
    (e.g., pipe-separated values in 'fullname_synonym(s)' and 'symbol_synonym(s)').
    """

    def expand_synonyms(column_name):
        """
        Expand a column containing pipe-separated synonyms into individual mappings.
        """
        if column_name not in mappings_df.columns:
            raise ValueError(f"Column '{column_name}' not found in the DataFrame.")

        # Select relevant rows where the column is not null
        relevant_rows = mappings_df.dropna(subset=[column_name]).copy()

        # Split pipe-separated values into lists
        relevant_rows[column_name] = relevant_rows[column_name].str.split("|")

        # Explode to create individual rows for each synonym
        expanded = relevant_rows.explode(column_name).rename(columns={column_name: "synonym"})
        expanded["synonym"] = expanded["synonym"].str.strip()  # Remove whitespace

        return expanded[["synonym", "primary_FBid"]]

    # Create mappings for each relevant column
    columns_to_expand = [
        "current_symbol",
        "current_fullname",
        "fullname_synonym(s)",
        "symbol_synonym(s)",
    ]
    all_mappings = pd.DataFrame()

    for column in columns_to_expand:
        print(f"Processing column: {column}")
        expanded_mapping = expand_synonyms(column)
        all_mappings = pd.concat([all_mappings, expanded_mapping], ignore_index=True)

    # Drop duplicates to ensure unique mappings
    all_mappings = all_mappings.drop_duplicates()

    # Convert to a dictionary for efficient lookups
    synonym_to_fbgnid_map = dict(zip(all_mappings["synonym"], all_mappings["primary_FBid"]))

    return synonym_to_fbgnid_map


def map_gene_ids(df, gene_to_fbgnid_main, gene_to_fbgnid_synonym, gene_col, corrected_col):
    """
    Efficiently maps ext_gene in df to flybase_gene_id using main and synonym mappings,
    while applying prioritized step combinations in a vectorized manner.

    `corrected_col` records the exact (possibly spelling-/case-normalized) form that
    produced each successful match, leaving the caller's original column untouched.
    """
    import re

    def clean_gene_vectorized(series, steps):
        """
        Apply a sequence of cleaning steps to a pandas Series in a vectorized manner.
        """
        for step in steps:
            if step == "lowercase":
                series = series.str.lower()
            elif step == "remove_zeros":
                series = (
                    series.str.rstrip("0")
                    .str.replace("-0", "-", regex=False)
                    .str.replace("-00", "-", regex=False)
                )
            elif step == "capitalize":
                series = series.str.capitalize()
            elif step == "remove_hyphen":
                series = series.str.replace("-", "", regex=False)
            elif step == "number_to_end":
                series = series.str.replace(
                    r"^(\d+)-(\D+)$", r"\2\1", regex=True
                )  # Handles cases like "1-Sep" -> "Sep1"
            elif step == "number_to_start":
                series = series.str.replace(
                    r"^(\D+)-(\d+)$", r"\2\1", regex=True
                )  # Handles cases like "Sep-1" -> "1Sep"
            elif step == "add_cr":
                series = "CR" + series
            elif step == "uppercase":
                series = series.str.upper()
        return series

    # Step 1: Initial mapping. The corrected form starts as the symbol-normalized,
    # stripped input and is overwritten whenever a cleaning combination resolves a gene.
    df[gene_col] = df[gene_col].str.strip()
    df[corrected_col] = df[gene_col]
    df["flybase_gene_id"] = df[gene_col].map(gene_to_fbgnid_main)

    # Step 2: Synonym mapping for initially unmapped genes
    unmapped_genes_mask = df["flybase_gene_id"].isna()
    print(f"TOTAL GENES: {df.shape[0]}")
    print(f"Initially missed genes: {unmapped_genes_mask.sum()}")
    df.loc[unmapped_genes_mask, "flybase_gene_id"] = df.loc[unmapped_genes_mask, gene_col].map(
        gene_to_fbgnid_synonym
    )
    print(f'Unmatched synonyms: {df["flybase_gene_id"].isna().sum()}')
    print(f'{df[df["flybase_gene_id"].isna()]}')
    # Step 3: Prioritized step combinations
    steps = ["lowercase", "remove_zeros", "remove_hyphen", "number_to_end", "uppercase"]  # 'add_cr'
    all_step_combinations = [combo for r in range(1, len(steps) + 1) for combo in combinations(steps, r)]
    # for r in range(1, len(steps) + 1):
    #     all_step_combinations.extend(permutations(steps, r))

    prev_unmapped_count = df["flybase_gene_id"].isna().sum()

    # Step 4: Apply combinations iteratively
    for step_combo in all_step_combinations:
        unmapped_idx = df.index[df["flybase_gene_id"].isna()]
        if len(unmapped_idx) == 0:
            break  # Exit if all genes are mapped

        # Clean only the still-unmapped genes for this combination.
        cleaned_genes = clean_gene_vectorized(df.loc[unmapped_idx, gene_col], step_combo)

        # Try the main symbol mapping, recording the corrected spelling on hits.
        main_hits = cleaned_genes.map(gene_to_fbgnid_main).dropna()
        df.loc[main_hits.index, "flybase_gene_id"] = main_hits
        df.loc[main_hits.index, corrected_col] = cleaned_genes.loc[main_hits.index]

        # Fall back to the synonym mapping for whatever is still unmapped.
        remaining_idx = unmapped_idx.difference(main_hits.index)
        syn_hits = cleaned_genes.loc[remaining_idx].map(gene_to_fbgnid_synonym).dropna()
        df.loc[syn_hits.index, "flybase_gene_id"] = syn_hits
        df.loc[syn_hits.index, corrected_col] = cleaned_genes.loc[syn_hits.index]

        current_unmapped_count = df["flybase_gene_id"].isna().sum()
        if current_unmapped_count == prev_unmapped_count:
            continue  # Skip redundant combinations
        else:
            print(
                f"{step_combo} was succesful in resolving {prev_unmapped_count - current_unmapped_count} genes"
            )
        prev_unmapped_count = current_unmapped_count

    # Step 5: Log remaining unmapped genes
    remaining_unmapped_mask = df["flybase_gene_id"].isna()
    print(f"Remaining unmapped genes after all steps: {remaining_unmapped_mask.sum()}")
    if remaining_unmapped_mask.sum() > 0:
        print("Remaining unmapped genes:")
        print(df.loc[remaining_unmapped_mask, gene_col])

    return df


def find_latest_tsv(directory: Path, pattern: str) -> Path:
    """Find latest FlyBase TSV, preferring .tsv.gz over .tsv."""
    gz_files = sorted(glob(str(directory / f"{pattern}*.tsv.gz")), reverse=True)
    if gz_files:
        return Path(gz_files[0])

    tsv_files = sorted(glob(str(directory / f"{pattern}*.tsv")), reverse=True)
    if tsv_files:
        return Path(tsv_files[0])

    raise FileNotFoundError(f"No TSV file matching '{pattern}' found in {directory}")


def load_flybase_tsv(filepath, **kwargs) -> pd.DataFrame:
    """
    Load a FlyBase TSV file, handling metadata/comments and header variants.
    Supports both .tsv and .tsv.gz files.
    """
    filepath = Path(filepath)

    if filepath.suffix == ".gz" or str(filepath).endswith(".tsv.gz"):
        opener = lambda f: gzip.open(f, "rt", encoding="utf-8")
    else:
        opener = lambda f: open(f, "r", encoding="utf-8")

    skip_rows = 0
    header_line = None
    use_custom_header = False
    header_found = False

    with opener(filepath) as f:
        for i, line in enumerate(f):
            stripped = line.strip()

            if not stripped:
                skip_rows = i + 1
                continue

            if header_found:
                if stripped.startswith("#"):
                    skip_rows = i + 1
                    continue
                skip_rows = i
                break

            if stripped.startswith("#"):
                if "\t" not in stripped:
                    skip_rows = i + 1
                    continue

                # Some FlyBase files use "##col1\tcol2..." headers.
                header_content = stripped.lstrip("#")
                cols = header_content.split("\t")
                col_names = [c.strip() for c in cols if c.strip()]

                if len(col_names) >= 2:
                    header_line = col_names
                    use_custom_header = True
                    header_found = True
                    skip_rows = i + 1
                    continue
                else:
                    skip_rows = i + 1
                    continue
            else:
                cols = stripped.split("\t")
                first_field = cols[0].strip() if cols else ""
                if first_field.startswith("FB") and len(first_field) > 4:
                    break
                else:
                    break

    if use_custom_header and header_line:
        df = pd.read_csv(
            filepath,
            sep="\t",
            skiprows=skip_rows,
            names=header_line,
            low_memory=False,
            on_bad_lines="warn",
            **kwargs,
        )
    else:
        df = pd.read_csv(
            filepath,
            sep="\t",
            skiprows=skip_rows,
            low_memory=False,
            on_bad_lines="warn",
            **kwargs,
        )

    return df


def load_mappings(flybase_data_dir: Path):
    columns_as_strings = {
        "current_symbol": str,
        "current_fullname": str,
        "fullname_synonym(s)": str,
        "symbol_synonym(s)": str,
        "primary_FBid": str,
    }

    synonym_path = find_latest_tsv(flybase_data_dir / "Genes", "fb_synonym")
    print(f"Loading FlyBase synonym table: {synonym_path}")
    mappings_df = load_flybase_tsv(
        synonym_path,
        dtype=columns_as_strings,
        keep_default_na=False,
    )
    mappings_df = mappings_df[mappings_df.organism_abbreviation == "Dmel"]
    mappings_df = mappings_df[mappings_df["primary_FBid"].str.startswith("FBgn", na=False)]

    synonym_to_fbgnid_map = create_expanded_mappings(mappings_df)
    gene_to_fbgnid_main = dict(
        zip(
            mappings_df["current_symbol"].astype(str).str.strip(),
            mappings_df["primary_FBid"].astype(str).str.strip(),
        )
    )
    fbgnid_to_symbol_map = dict(
        zip(
            mappings_df["primary_FBid"].astype(str).str.strip(),
            mappings_df["current_symbol"].astype(str).str.strip(),
        )
    )
    return gene_to_fbgnid_main, synonym_to_fbgnid_map, fbgnid_to_symbol_map


def is_conversion_sidecar(path) -> bool:
    """Return True for files this script writes, so a re-run cannot consume them."""
    name = Path(path).name
    if name == NEEDS_REVIEW_NAME:
        return True
    return name.startswith(VALIDATED_PREFIX) and name.endswith(".csv")


def is_mapped_fbgn(value) -> bool:
    text = str(value).strip()
    return text.startswith("FBgn")


def select_source_csv_paths(input_dir: Path) -> list[Path]:
    """Return user CSVs only. Skip validated_*.csv and needs-review.csv."""
    return sorted(
        path
        for path in Path(input_dir).glob("*.csv")
        if not is_conversion_sidecar(path)
    )


def _mapped_mask(series: pd.Series) -> pd.Series:
    return series.astype(str).str.strip().str.startswith("FBgn")


def write_review_xlsx(xlsx_path: Path, gene_column: str, validated_df: pd.DataFrame, review_df: pd.DataFrame) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils.dataframe import dataframe_to_rows

    corrected_col = "corrected_" + gene_column
    review_cols = [gene_column, corrected_col, "primary_symbol", "flybase_gene_id"]
    for extra in review_cols:
        if extra not in validated_df.columns and not validated_df.empty:
            raise KeyError(f"Validated table is missing column {extra}")

    workbook = Workbook()
    validated_sheet = workbook.active
    validated_sheet.title = "Validated"
    export_validated = (
        validated_df.loc[:, [c for c in review_cols if c in validated_df.columns]]
        if not validated_df.empty
        else pd.DataFrame(columns=review_cols)
    )
    for row in dataframe_to_rows(export_validated, index=False, header=True):
        validated_sheet.append(row)
    id_col = review_cols.index("flybase_gene_id") + 1
    link_font = Font(color="0563C1", underline="single")
    for row_idx in range(2, validated_sheet.max_row + 1):
        cell = validated_sheet.cell(row=row_idx, column=id_col)
        fbgn = str(cell.value or "").strip()
        if is_mapped_fbgn(fbgn):
            cell.hyperlink = FLYBASE_REPORT_URL.format(fbgn=fbgn)
            cell.font = link_font

    review_sheet = workbook.create_sheet("Needs review")
    export_review = (
        review_df.loc[:, [c for c in review_cols if c in review_df.columns]]
        if not review_df.empty
        else pd.DataFrame(columns=review_cols)
    )
    for row in dataframe_to_rows(export_review, index=False, header=True):
        review_sheet.append(row)

    workbook.save(xlsx_path)


def process_csv_file(
    csv_file, gene_column, gene_to_fbgnid_main, synonym_to_fbgnid_map, fbgnid_to_symbol_map
):
    """Map genes and write sidecar files. Never overwrites ``csv_file``.

    Returns the unmatched rows (with ``source_file``) or None if the file was skipped.
    """
    csv_path = Path(csv_file)
    print(f"Processing file: {csv_path}")

    header_row = 0
    try:
        temp_df = pd.read_csv(csv_path, nrows=10, header=None, encoding="utf-8-sig")
        for i, row in temp_df.iterrows():
            if gene_column in row.astype(str).values:
                header_row = i
                print(f"Detected header at row {i}")
                break
    except Exception as e:
        print(f"Header detection skipped/failed (using default 0): {e}")

    custom_na_values = [
        "",
        "#N/A",
        "#N/A N/A",
        "#NA",
        "-1.#IND",
        "-1.#QNAN",
        "-NaN",
        "-nan",
        "1.#IND",
        "1.#QNAN",
        "<NA>",
        "N/A",
        "NA",
        "NULL",
        "NaN",
        "n/a",
        "nan",
        "null",
    ]
    df = pd.read_csv(
        csv_path,
        encoding="utf-8-sig",
        header=header_row,
        na_values=custom_na_values,
        keep_default_na=False,
        low_memory=False,
    )

    if gene_column not in df.columns:
        if "GO" in csv_path.name and "Genes" in df.columns:
            df["Genes"] = df["Genes"].astype(str).str.split(", ")
            df = df.explode("Genes").reset_index(drop=True)
            df[gene_column] = df["Genes"]
        else:
            print(f"'{gene_column}' column not found in {csv_path}. Skipping file.")
            return None

    working_col = gene_column + "_new"
    corrected_col = "corrected_" + gene_column

    for existing_col in ("flybase_gene_id", "primary_symbol", corrected_col):
        if existing_col in df.columns:
            df.drop(columns=[existing_col], inplace=True)

    # Keep the user's original `gene_column` untouched. All spelling/symbol/
    # capitalization normalization happens on a throwaway working column; the
    # form that actually matched is captured in `corrected_col`.
    df[gene_column] = df[gene_column].astype(str)
    df[working_col] = replace_symbol(df[gene_column])
    df = df.dropna(subset=[working_col])
    df = map_gene_ids(
        df, gene_to_fbgnid_main, synonym_to_fbgnid_map, working_col, corrected_col
    )
    df["primary_symbol"] = df["flybase_gene_id"].map(fbgnid_to_symbol_map)

    # Only surface `corrected_col` when it actually differs from the user's
    # original input; otherwise mark it as "-" to signal no spelling correction.
    no_correction_mask = df[corrected_col].astype(str) == df[gene_column].astype(str)
    df.loc[no_correction_mask, corrected_col] = "-"

    df.fillna("-", inplace=True)
    df.drop(columns=[working_col], inplace=True)

    # Order the derived columns up front, preserving any other original columns after.
    ordered = [gene_column, corrected_col, "primary_symbol", "flybase_gene_id"]
    remaining = [c for c in df.columns if c not in ordered]
    df = df[ordered + remaining]

    mapped_mask = _mapped_mask(df["flybase_gene_id"])
    validated_df = df.loc[mapped_mask].copy()
    review_df = df.loc[~mapped_mask].copy()
    review_df.insert(0, "source_file", csv_path.name)

    validated_csv = csv_path.with_name(f"{VALIDATED_PREFIX}{csv_path.name}")
    validated_xlsx = csv_path.with_name(f"{VALIDATED_PREFIX}{csv_path.stem}.xlsx")
    validated_df.to_csv(validated_csv, index=False, encoding="utf-8-sig")
    write_review_xlsx(validated_xlsx, gene_column, validated_df, review_df.drop(columns=["source_file"]))

    print(f"Left original untouched: {csv_path}")
    print(f"Wrote {validated_csv} ({len(validated_df)} mapped rows).")
    print(f"Wrote {validated_xlsx} (click each FBgn to open the FlyBase gene report).")
    print(f"Unmatched rows from this file: {len(review_df)}")
    return review_df


def _print_review_gate(validated_paths, unmatched_count: int) -> None:
    print("\n" + "=" * 72)
    print("STOP. Do not run stocker yet.")
    print("The original gene-list CSV was not edited.")
    print("Review the companion xlsx (FBgn cells are FlyBase hyperlinks) for precision:")
    for path in validated_paths:
        print(f"  {path}")
    print(f"ID validator: {FLYBASE_CONVERT_URL}")
    if unmatched_count:
        print(
            f"{unmatched_count} row(s) are in {NEEDS_REVIEW_NAME}. "
            "Paste those symbols at the validator before treating any list as ready."
        )
    else:
        print(
            "No unmatched rows. Still open the xlsx and confirm each FBgn "
            "before running stocker on validated_*.csv only."
        )
    print("Stocker must run only on validated_*.csv after that human review.")
    print("=" * 72)


def main(argv=None):
    parser = argparse.ArgumentParser(
        description=(
            "Map fly gene symbols to FBgn IDs. Never overwrites the input CSV; "
            "writes validated_*.csv, validated_*.xlsx, and needs-review.csv."
        )
    )
    parser.add_argument("input_directory", help="Directory containing input CSV files")
    parser.add_argument(
        "input_gene_col",
        nargs="?",
        default="ext_gene",
        help="Input gene symbol column (default: ext_gene)",
    )
    parser.add_argument(
        "--flybase-data-dir",
        default=os.environ.get("FLYBASE_DATA_DIR", str(DEFAULT_FLYBASE_DATA)),
        help="Path to FlyBase data directory (default: Data/FlyBase path)",
    )
    args = parser.parse_args(argv)

    input_dir = Path(args.input_directory)
    if not input_dir.is_dir():
        print(f"Error: '{input_dir}' is not a valid directory")
        return 1

    flybase_data_dir = Path(args.flybase_data_dir)
    if not flybase_data_dir.exists():
        print(f"Error: FlyBase data directory not found at: {flybase_data_dir}")
        return 1

    csv_files = select_source_csv_paths(input_dir)
    if not csv_files:
        print(f"No source CSV files found in {input_dir} (sidecars are ignored).")
        return 0

    try:
        gene_to_fbgnid_main, synonym_to_fbgnid_map, fbgnid_to_symbol_map = load_mappings(
            flybase_data_dir
        )
    except Exception as e:
        print(f"Error loading FlyBase synonym mappings: {e}")
        return 1

    review_frames = []
    validated_xlsx_paths = []
    for csv_file in csv_files:
        review_df = process_csv_file(
            csv_file,
            args.input_gene_col,
            gene_to_fbgnid_main,
            synonym_to_fbgnid_map,
            fbgnid_to_symbol_map,
        )
        if review_df is None:
            continue
        review_frames.append(review_df)
        validated_xlsx_paths.append(csv_file.with_name(f"{VALIDATED_PREFIX}{csv_file.stem}.xlsx"))

    needs_review_path = input_dir / NEEDS_REVIEW_NAME
    if review_frames:
        combined_review = pd.concat(review_frames, ignore_index=True)
    else:
        combined_review = pd.DataFrame(columns=["source_file", args.input_gene_col, "flybase_gene_id"])
    combined_review.to_csv(needs_review_path, index=False, encoding="utf-8-sig")
    print(f"Wrote {needs_review_path} ({len(combined_review)} unmatched rows).")

    _print_review_gate(validated_xlsx_paths, len(combined_review))
    print("All files processed.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
