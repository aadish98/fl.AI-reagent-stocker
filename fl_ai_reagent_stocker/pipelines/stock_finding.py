"""
Stage 1: find-stocks.

This module builds the stock/reference workbook by:
1. Loading input genes from CSV files
2. Mapping genes to FlyBase stocks via derived stock components
3. Getting references for gene-relevant alleles
4. Scoring references against config keywords
5. Saving the Stage 1 Excel workbook

Usage:
    from fl_ai_reagent_stocker.pipelines.stock_finding import StockFindingPipeline
    from fl_ai_reagent_stocker.config import Settings

    settings = Settings()
    pipeline = StockFindingPipeline(settings)
    pipeline.run(input_dir, keywords=["sleep", "circadian"], run_functional_validation=False)
"""

import gc
import subprocess
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Set

import pandas as pd

from ..config import Settings, ValidationStatus, GET_FBGN_IDS_SCRIPT
from ..utils import (
    clean_id,
    parse_semicolon_list,
    unique_join,
    find_latest_tsv,
    load_flybase_tsv,
    generate_keyword_column_name,
    get_gpt_derived_columns,
    apply_experimental_prefix,
    EXPERIMENTAL_PREFIX,
)
from ..integrations.pubmed import PubMedClient, PubMedCache
from ..integrations.fulltext import FullTextFetcher, FunctionalValidator
from ..validation_runner import run_functional_validation


def _apply_grey_fill_openpyxl(ws, df: pd.DataFrame) -> None:
    """
    Apply light grey background fill to GPT-derived column **headers** in an
    openpyxl worksheet.  Only the header row (row 1) is shaded.

    GPT-derived columns are detected by the '[EXPERIMENTAL] ' prefix that
    :func:`apply_experimental_prefix` adds to their header names.
    """
    from openpyxl.styles import PatternFill

    grey_fill = PatternFill(
        start_color='D9D9D9', end_color='D9D9D9', fill_type='solid'
    )
    # openpyxl columns are 1-indexed; row 1 is the header
    for col_idx, col_name in enumerate(df.columns, 1):
        if str(col_name).startswith(EXPERIMENTAL_PREFIX):
            ws.cell(row=1, column=col_idx).fill = grey_fill


class StockFindingPipeline:
    """
    Pipeline for mapping genes to stocks and references with functional validation.
    
    This is Pipeline 1 in the fly-stocker workflow:
    Input: CSV files with gene IDs
    Output: Excel file with stocks, references, and validation results
    """
    
    def __init__(self, settings: Optional[Settings] = None):
        """
        Initialize the pipeline.
        
        Args:
            settings: Configuration settings (uses defaults if None)
        """
        self.settings = settings or Settings()
        
        # Validate settings
        issues = self.settings.validate()
        for issue in issues:
            print(f"    Warning: {issue}")
        
        # Initialize components
        self._pubmed_cache = PubMedCache(self.settings.pubmed_cache_path)
        self._pubmed_client = PubMedClient(
            cache=self._pubmed_cache,
            api_key=self.settings.ncbi_api_key
        )
        self._fulltext_fetcher = FullTextFetcher(
            unpaywall_token=self.settings.unpaywall_token,
            ncbi_api_key=self.settings.ncbi_api_key,
            method_cache_path=self.settings.fulltext_method_cache_path
        )
        self._validator = FunctionalValidator(
            openai_api_key=self.settings.openai_api_key,
            model=self.settings.openai_model,
            log_dir=self.settings.gpt_log_dir / datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            if self.settings.enable_gpt_logging else None
        )
        
        # Persist details of filtered-in references that still have no PMID.
        self._no_pmid_fbrf_records: List[Dict[str, str]] = []
        
        # Data tables (loaded lazily)
        self._tables: Optional[dict] = None
    
    def _load_tables(self) -> dict:
        """Load FlyBase TSV files and FlyBase-derived helper tables."""
        if self._tables is not None:
            return self._tables
        
        print("\nLoading data tables...")
        tables = {}
        
        # Load FlyBase-derived stock components
        derived_components_path = self.settings.flybase_derived_components_path
        print(f"  Loading: {derived_components_path.name}")
        tables["derived_stock_components"] = pd.read_csv(
            derived_components_path,
            dtype=str,
        ).fillna("")
        if "collection_short_name" in tables["derived_stock_components"].columns:
            tables["derived_stock_components"] = tables["derived_stock_components"].rename(
                columns={"collection_short_name": "collection"}
            )
        for col in (
            "FBst",
            "stock_number",
            "collection",
            "FB_genotype",
            "derived_stock_component",
            "embedded_type",
            "object_symbol",
            "AlleleSymbol",
            "FBgnID",
            "GeneSymbol",
        ):
            if col not in tables["derived_stock_components"].columns:
                tables["derived_stock_components"][col] = ""
        tables["derived_stock_components"]["FBst"] = tables["derived_stock_components"]["FBst"].apply(clean_id)
        tables["derived_stock_components"]["stock_number"] = (
            tables["derived_stock_components"]["stock_number"].apply(clean_id)
        )
        print(
            "    Loaded "
            f"{len(tables['derived_stock_components'])} FlyBase derived stock-component rows"
        )
        
        # Load entity_publication
        entity_pub_path = find_latest_tsv(self.settings.flybase_references_path, "entity_publication")
        print(f"  Loading: {entity_pub_path.name}")
        tables["entity_publication"] = load_flybase_tsv(entity_pub_path)
        print(f"    Loaded {len(tables['entity_publication'])} entity-publication entries")
        
        # Load reference metadata
        fbrf_path = find_latest_tsv(self.settings.flybase_references_path, "fbrf_pmid_pmcid_doi")
        print(f"  Loading: {fbrf_path.name}")
        tables["ref_metadata"] = load_flybase_tsv(fbrf_path)
        
        # Load fbal_to_fbgn for allele-to-gene mapping
        fbal_to_fbgn_path = find_latest_tsv(self.settings.flybase_alleles_stocks_path, "fbal_to_fbgn")
        print(f"  Loading: {fbal_to_fbgn_path.name}")
        tables["fbal_to_fbgn"] = load_flybase_tsv(fbal_to_fbgn_path)
        print(f"    Loaded {len(tables['fbal_to_fbgn'])} allele-to-gene mappings")

        # Load transgenic construct descriptions (FBal -> FBtp)
        try:
            construct_path = find_latest_tsv(
                self.settings.flybase_transgenic_constructs_path,
                "transgenic_construct_descriptions",
            )
            print(f"  Loading: {construct_path.name}")
            tables["transgenic_construct_descriptions"] = load_flybase_tsv(construct_path)
            print(
                "    Loaded "
                f"{len(tables['transgenic_construct_descriptions'])} transgenic construct description rows"
            )
        except FileNotFoundError:
            print(
                "  Warning: transgenic_construct_descriptions TSV not found; "
                "FBtp joins will be skipped"
            )
            tables["transgenic_construct_descriptions"] = pd.DataFrame()

        # Load optional chado-derived FBtp -> FBti mapping (construct -> insertion)
        fbtp_to_fbti_path = self.settings.flybase_transgenic_insertions_path / "fbtp_to_fbti.csv"
        if fbtp_to_fbti_path.exists():
            print(f"  Loading: {fbtp_to_fbti_path.name}")
            tables["fbtp_to_fbti"] = pd.read_csv(fbtp_to_fbti_path, dtype=str).fillna("")
            print(
                "    Loaded "
                f"{len(tables['fbtp_to_fbti'])} FBtp-to-FBti mapping rows"
            )
        else:
            print(
                "  Warning: fbtp_to_fbti.csv not found; "
                "FBtp -> FBti expansion will be skipped"
            )
            tables["fbtp_to_fbti"] = pd.DataFrame(columns=["FBtp", "FBti"])

        # Load classical/insertion allele descriptions (FBal -> FBti)
        try:
            insertion_path = find_latest_tsv(
                self.settings.flybase_alleles_stocks_path,
                "dmel_classical_and_insertion_allele_descriptions",
            )
            print(f"  Loading: {insertion_path.name}")
            tables["insertion_allele_descriptions"] = load_flybase_tsv(insertion_path)
            print(
                "    Loaded "
                f"{len(tables['insertion_allele_descriptions'])} classical/insertion allele description rows"
            )
        except FileNotFoundError:
            print(
                "  Warning: dmel_classical_and_insertion_allele_descriptions TSV not found; "
                "FBti joins will be skipped"
            )
            tables["insertion_allele_descriptions"] = pd.DataFrame()

        # Load FlyBase stocks TSV for collection/genotype metadata
        try:
            stocks_tsv_path = find_latest_tsv(self.settings.flybase_alleles_stocks_path, "stocks")
            print(f"  Loading: {stocks_tsv_path.name}")
            fb_stocks = load_flybase_tsv(stocks_tsv_path)
            if "collection_short_name" in fb_stocks.columns:
                fb_stocks = fb_stocks.rename(columns={"collection_short_name": "collection"})
            if 'stock_number' in fb_stocks.columns:
                fb_stocks['stock_number'] = fb_stocks['stock_number'].apply(clean_id)
            if 'FBst' in fb_stocks.columns:
                fb_stocks['FBst'] = fb_stocks['FBst'].apply(clean_id)
            tables["fb_stocks"] = fb_stocks
            print(f"    Loaded {len(fb_stocks)} stock entries from FlyBase stocks TSV")
        except FileNotFoundError:
            print("  Warning: FlyBase stocks TSV not found, stock metadata refresh will be skipped")
            tables["fb_stocks"] = None
        
        # Load genotype_phenotype_data for custom phenotype reagent detection
        try:
            phenotype_path = find_latest_tsv(
                self.settings.flybase_alleles_stocks_path, "genotype_phenotype_data"
            )
            print(f"  Loading: {phenotype_path.name}")
            tables["genotype_phenotype_data"] = load_flybase_tsv(phenotype_path)
            print(
                f"    Loaded {len(tables['genotype_phenotype_data'])} genotype-phenotype entries"
            )
        except FileNotFoundError:
            print(
                "  Warning: genotype_phenotype_data TSV not found, "
                "custom phenotype rows will be skipped"
            )
            tables["genotype_phenotype_data"] = pd.DataFrame()

        self._tables = tables
        return tables
    
    def _run_fbgnid_conversion(self, input_dir: Path, gene_col: str = "ext_gene") -> bool:
        """Run the FBgn ID helper script to convert gene names to FlyBase IDs."""
        print(f"\n{'='*60}")
        print("Step 0: Running FBgnID conversion on input files...")
        print(f"{'='*60}")
        
        if not GET_FBGN_IDS_SCRIPT.exists():
            print(f"ERROR: FBgn ID helper script not found at {GET_FBGN_IDS_SCRIPT}")
            return False
        
        cmd = [sys.executable, str(GET_FBGN_IDS_SCRIPT), str(input_dir), gene_col,
               "--flybase-data-dir", str(self.settings.flybase_data_path)]
        print(f"  Running: python {GET_FBGN_IDS_SCRIPT.name} {input_dir} {gene_col}")
        
        try:
            result = subprocess.run(cmd, capture_output=True, text=True, check=False)
            print(result.stdout)
            if result.stderr:
                print(result.stderr)
        except Exception as e:
            print(f"ERROR: Failed to run {GET_FBGN_IDS_SCRIPT.name}: {e}")
            return False
        
        return True
    
    def _validate_fbgnids(self, csv_path: Path, gene_col: str = "flybase_gene_id") -> List[str]:
        """Validate that all rows have valid FlyBase gene IDs."""
        df = pd.read_csv(csv_path)
        
        if gene_col not in df.columns:
            return []
        
        invalid_mask = (
            df[gene_col].isna() |
            (df[gene_col].astype(str).str.strip() == "") |
            (df[gene_col].astype(str).str.strip() == "-") |
            ~df[gene_col].astype(str).str.startswith("FBgn")
        )
        
        if invalid_mask.any():
            gene_symbol_col = None
            for col in ["ext_gene", "gene_symbol", "gene", "Gene"]:
                if col in df.columns:
                    gene_symbol_col = col
                    break
            
            if gene_symbol_col:
                return df.loc[invalid_mask, gene_symbol_col].dropna().unique().tolist()
            else:
                return df.loc[invalid_mask, gene_col].astype(str).unique().tolist()
        
        return []
    
    def _aggregate_input_genes(self, input_dir: Path, gene_col: str) -> Tuple[List[str], List[Path]]:
        """Read all CSV files and aggregate unique genes."""
        csv_files = sorted(input_dir.glob("*.csv"))
        
        if not csv_files:
            return [], []
        
        all_genes = set()
        valid_files = []
        
        print(f"\nAggregating genes from {len(csv_files)} CSV file(s)...")
        
        for csv_path in csv_files:
            try:
                gene_series_chunks = []
                for chunk in pd.read_csv(csv_path, usecols=[gene_col], chunksize=10000):
                    gene_series_chunks.append(chunk[gene_col].dropna())
                
                if gene_series_chunks:
                    genes = pd.concat(gene_series_chunks, ignore_index=True).unique()
                    genes = [g for g in genes if isinstance(g, str) and g.startswith("FBgn")]
                    all_genes.update(genes)
                    valid_files.append(csv_path)
                    print(f"  {csv_path.name}: {len(genes)} unique genes")
            except (KeyError, ValueError) as e:
                print(f"  {csv_path.name}: Skipped (no '{gene_col}' column or read error)")
                continue
        
        unique_genes = sorted(all_genes)
        print(f"\nTotal: {len(unique_genes)} unique genes from {len(valid_files)} file(s)")
        
        return unique_genes, valid_files
    
    def _build_gene_component_tables(
        self,
        input_genes: List[str],
        fbal_to_fbgn_df: pd.DataFrame,
        transgenic_constructs_df: Optional[pd.DataFrame] = None,
        insertion_alleles_df: Optional[pd.DataFrame] = None,
        fbtp_to_fbti_df: Optional[pd.DataFrame] = None,
    ) -> Dict[str, Any]:
        """Build per-allele construct and insertion lookup tables for input genes."""
        input_gene_set = set(input_genes)

        fbal_df = fbal_to_fbgn_df.copy()
        fbal_df["GeneID"] = fbal_df["GeneID"].astype(str).str.strip()
        fbal_df["AlleleID"] = fbal_df["AlleleID"].astype(str).str.strip()
        fbal_df["GeneSymbol"] = fbal_df.get("GeneSymbol", "").astype(str).str.strip()
        fbal_df["AlleleSymbol"] = fbal_df.get("AlleleSymbol", "").astype(str).str.strip()
        gene_alleles = fbal_df[fbal_df["GeneID"].isin(input_gene_set)].copy()

        empty_alleles = pd.DataFrame(
            columns=[
                "FBgnID",
                "GeneSymbol",
                "AlleleID",
                "AlleleSymbol",
                "AlleleClassTerm",
                "AlleleSupportingFBrf",
            ]
        )
        empty_constructs = pd.DataFrame(
            columns=[
                "FBgnID",
                "GeneSymbol",
                "AlleleID",
                "AlleleSymbol",
                "AlleleClassTerm",
                "AlleleSupportingFBrf",
                "FBtpID",
                "FBtpSymbol",
                "TransgenicProductClassTerm",
                "ConstructSupportingFBrf",
            ]
        )
        empty_insertions = pd.DataFrame(
            columns=[
                "FBgnID",
                "GeneSymbol",
                "AlleleID",
                "AlleleSymbol",
                "AlleleClassTerm",
                "AlleleSupportingFBrf",
                "FBtiID",
                "FBtiSymbol",
                "InsertionSupportingFBrf",
            ]
        )

        if len(gene_alleles) == 0:
            return {
                "gene_alleles": gene_alleles,
                "alleles": empty_alleles,
                "constructs": empty_constructs,
                "insertions": empty_insertions,
                "constructs_by_allele": {},
                "insertions_by_allele": {},
            }

        def _split_pipe_values(value: Any) -> List[str]:
            if pd.isna(value):
                return []
            return [item.strip() for item in str(value).split("|") if item.strip()]

        def _pipe_unique_join(series: pd.Series) -> str:
            values: List[str] = []
            for raw in series:
                values.extend(_split_pipe_values(raw))
            return unique_join(values)

        allele_rows = gene_alleles[["GeneID", "GeneSymbol", "AlleleID", "AlleleSymbol"]].copy()
        allele_rows = allele_rows.rename(columns={"GeneID": "FBgnID"})
        allele_rows = allele_rows.drop_duplicates(subset=["FBgnID", "GeneSymbol", "AlleleID", "AlleleSymbol"])

        allele_ids = set(allele_rows["AlleleID"].astype(str).str.strip())

        insertion_source = pd.DataFrame()
        if insertion_alleles_df is not None and len(insertion_alleles_df) > 0:
            insertion_source = insertion_alleles_df.copy()
            for col in (
                "Allele (id)",
                "Insertion (id)",
                "Insertion (symbol)",
                "Allele Class (term)",
                "Description (supporting reference)",
            ):
                if col not in insertion_source.columns:
                    insertion_source[col] = ""
            insertion_source["Allele (id)"] = insertion_source["Allele (id)"].apply(clean_id)
            insertion_source = insertion_source[
                insertion_source["Allele (id)"].isin(allele_ids)
            ].copy()

        if len(insertion_source) > 0:
            allele_meta = (
                insertion_source.groupby("Allele (id)", sort=False)
                .agg(
                    AlleleClassTerm=("Allele Class (term)", _pipe_unique_join),
                    AlleleSupportingFBrf=("Description (supporting reference)", _pipe_unique_join),
                )
                .reset_index()
                .rename(columns={"Allele (id)": "AlleleID"})
            )
            allele_rows = allele_rows.merge(allele_meta, on="AlleleID", how="left")
        else:
            allele_rows["AlleleClassTerm"] = ""
            allele_rows["AlleleSupportingFBrf"] = ""

        for col in ("AlleleClassTerm", "AlleleSupportingFBrf"):
            if col not in allele_rows.columns:
                allele_rows[col] = ""
            allele_rows[col] = allele_rows[col].fillna("").astype(str).str.strip()

        allele_meta_by_id = (
            allele_rows.drop_duplicates(subset=["AlleleID"])
            .set_index("AlleleID")
            .to_dict("index")
        )

        construct_rows: List[Dict[str, str]] = []
        if transgenic_constructs_df is not None and len(transgenic_constructs_df) > 0:
            construct_source = transgenic_constructs_df.copy()
            for col in (
                "Component Allele (id)",
                "Transgenic Construct (id)",
                "Transgenic Construct (symbol)",
                "Transgenic Product class (term)",
                "Description (supporting reference)",
            ):
                if col not in construct_source.columns:
                    construct_source[col] = ""
            construct_source["Component Allele (id)"] = construct_source["Component Allele (id)"].apply(clean_id)
            construct_source = construct_source[
                construct_source["Component Allele (id)"].isin(allele_ids)
            ].copy()

            for _, row in construct_source.iterrows():
                allele_id = clean_id(row.get("Component Allele (id)", ""))
                allele_meta = allele_meta_by_id.get(allele_id)
                if not allele_meta:
                    continue
                construct_ids = [
                    clean_id(value)
                    for value in _split_pipe_values(row.get("Transgenic Construct (id)", ""))
                    if clean_id(value)
                ]
                if not construct_ids:
                    continue
                construct_symbols = _split_pipe_values(row.get("Transgenic Construct (symbol)", ""))
                class_term = unique_join(_split_pipe_values(row.get("Transgenic Product class (term)", "")))
                supporting_fbrf = unique_join(
                    _split_pipe_values(row.get("Description (supporting reference)", ""))
                )
                for idx, construct_id in enumerate(construct_ids):
                    construct_symbol = ""
                    if idx < len(construct_symbols):
                        construct_symbol = construct_symbols[idx]
                    elif len(construct_symbols) == 1:
                        construct_symbol = construct_symbols[0]
                    construct_rows.append({
                        "FBgnID": str(allele_meta.get("FBgnID", "")),
                        "GeneSymbol": str(allele_meta.get("GeneSymbol", "")),
                        "AlleleID": allele_id,
                        "AlleleSymbol": str(allele_meta.get("AlleleSymbol", "")),
                        "AlleleClassTerm": str(allele_meta.get("AlleleClassTerm", "")),
                        "AlleleSupportingFBrf": str(allele_meta.get("AlleleSupportingFBrf", "")),
                        "FBtpID": construct_id,
                        "FBtpSymbol": construct_symbol,
                        "TransgenicProductClassTerm": class_term,
                        "ConstructSupportingFBrf": supporting_fbrf,
                    })

        construct_df = pd.DataFrame(construct_rows, columns=empty_constructs.columns)

        insertion_rows: List[Dict[str, str]] = []
        if len(insertion_source) > 0:
            for _, row in insertion_source.iterrows():
                allele_id = clean_id(row.get("Allele (id)", ""))
                allele_meta = allele_meta_by_id.get(allele_id)
                if not allele_meta:
                    continue
                insertion_ids = [
                    clean_id(value)
                    for value in _split_pipe_values(row.get("Insertion (id)", ""))
                    if clean_id(value)
                ]
                if not insertion_ids:
                    continue
                insertion_symbols = _split_pipe_values(row.get("Insertion (symbol)", ""))
                supporting_fbrf = unique_join(
                    _split_pipe_values(row.get("Description (supporting reference)", ""))
                )
                for idx, insertion_id in enumerate(insertion_ids):
                    insertion_symbol = ""
                    if idx < len(insertion_symbols):
                        insertion_symbol = insertion_symbols[idx]
                    elif len(insertion_symbols) == 1:
                        insertion_symbol = insertion_symbols[0]
                    insertion_rows.append({
                        "FBgnID": str(allele_meta.get("FBgnID", "")),
                        "GeneSymbol": str(allele_meta.get("GeneSymbol", "")),
                        "AlleleID": allele_id,
                        "AlleleSymbol": str(allele_meta.get("AlleleSymbol", "")),
                        "AlleleClassTerm": str(allele_meta.get("AlleleClassTerm", "")),
                        "AlleleSupportingFBrf": str(allele_meta.get("AlleleSupportingFBrf", "")),
                        "FBtiID": insertion_id,
                        "FBtiSymbol": insertion_symbol,
                        "InsertionSupportingFBrf": supporting_fbrf,
                    })

        if (
            fbtp_to_fbti_df is not None
            and len(fbtp_to_fbti_df) > 0
            and len(construct_df) > 0
        ):
            mapping_df = fbtp_to_fbti_df.copy()
            for col in ("FBtp", "FBti"):
                if col not in mapping_df.columns:
                    mapping_df[col] = ""
                mapping_df[col] = mapping_df[col].apply(clean_id)
            mapping_df = mapping_df[
                mapping_df["FBtp"].astype(bool) & mapping_df["FBti"].astype(bool)
            ].drop_duplicates(subset=["FBtp", "FBti"])

            if len(mapping_df) > 0:
                mapped_insertions = 0
                construct_df_for_mapping = construct_df.drop_duplicates(
                    subset=["AlleleID", "FBtpID"]
                ).copy()
                fbtp_to_fbti_ids: Dict[str, List[str]] = (
                    mapping_df.groupby("FBtp", sort=False)["FBti"].agg(list).to_dict()
                )

                for _, construct_row in construct_df_for_mapping.iterrows():
                    fbti_ids = fbtp_to_fbti_ids.get(clean_id(construct_row.get("FBtpID", "")), [])
                    if not fbti_ids:
                        continue
                    for fbti_id in fbti_ids:
                        insertion_rows.append(
                            {
                                "FBgnID": str(construct_row.get("FBgnID", "")),
                                "GeneSymbol": str(construct_row.get("GeneSymbol", "")),
                                "AlleleID": str(construct_row.get("AlleleID", "")),
                                "AlleleSymbol": str(construct_row.get("AlleleSymbol", "")),
                                "AlleleClassTerm": str(construct_row.get("AlleleClassTerm", "")),
                                "AlleleSupportingFBrf": str(
                                    construct_row.get("AlleleSupportingFBrf", "")
                                ),
                                "FBtiID": fbti_id,
                                "FBtiSymbol": "",
                                "InsertionSupportingFBrf": str(
                                    construct_row.get("ConstructSupportingFBrf", "")
                                ),
                            }
                        )
                        mapped_insertions += 1
                if mapped_insertions > 0:
                    print(
                        "    Expanded "
                        f"{mapped_insertions} FBti insertion links from FBtp constructs"
                    )

        insertion_df = pd.DataFrame(insertion_rows, columns=empty_insertions.columns)

        def _group_component_rows(df: pd.DataFrame, id_col: str, symbol_col: str, class_col: str, ref_col: str) -> Dict[str, Dict[str, str]]:
            grouped: Dict[str, Dict[str, str]] = {}
            if df is None or len(df) == 0:
                return grouped
            for allele_id, group in df.groupby("AlleleID", sort=False):
                grouped[str(allele_id).strip()] = {
                    "IDs": unique_join(group[id_col].tolist()),
                    "Symbols": unique_join(group[symbol_col].tolist()),
                    "ClassTerm": unique_join(group[class_col].tolist()) if class_col in group.columns else "",
                    "SupportingFBrf": unique_join(group[ref_col].tolist()) if ref_col in group.columns else "",
                }
            return grouped

        return {
            "gene_alleles": gene_alleles,
            "alleles": allele_rows,
            "constructs": construct_df.drop_duplicates(),
            "insertions": insertion_df.drop_duplicates(),
            "constructs_by_allele": _group_component_rows(
                construct_df,
                "FBtpID",
                "FBtpSymbol",
                "TransgenicProductClassTerm",
                "ConstructSupportingFBrf",
            ),
            "insertions_by_allele": _group_component_rows(
                insertion_df,
                "FBtiID",
                "FBtiSymbol",
                "AlleleClassTerm",
                "InsertionSupportingFBrf",
            ),
        }

    def _build_stock_mapping(
        self,
        input_genes: List[str],
        derived_components_df: pd.DataFrame,
        fbal_to_fbgn_df: pd.DataFrame,
        transgenic_constructs_df: Optional[pd.DataFrame] = None,
        insertion_alleles_df: Optional[pd.DataFrame] = None,
        fbtp_to_fbti_df: Optional[pd.DataFrame] = None,
        fb_stocks_df: Optional[pd.DataFrame] = None,
    ) -> pd.DataFrame:
        """
        Build mapping from input genes to FlyBase stocks using exact ID joins.

        Stocks can be recovered through three direct match paths in the derived
        component CSV: FBal allele IDs, FBtp construct IDs, and FBti insertion IDs.
        Reference aggregation later uses the union of allele, construct, and
        insertion IDs associated with the source allele(s) that matched the stock.
        """
        if not input_genes:
            return pd.DataFrame()

        component_tables = self._build_gene_component_tables(
            input_genes,
            fbal_to_fbgn_df,
            transgenic_constructs_df,
            insertion_alleles_df,
            fbtp_to_fbti_df,
        )
        gene_alleles = component_tables["gene_alleles"]
        allele_rows = component_tables["alleles"]
        construct_rows = component_tables["constructs"]
        insertion_rows = component_tables["insertions"]
        constructs_by_allele = component_tables["constructs_by_allele"]
        insertions_by_allele = component_tables["insertions_by_allele"]

        if len(gene_alleles) == 0 or len(allele_rows) == 0:
            print(f"    Warning: No alleles found for {len(set(input_genes))} input genes in fbal_to_fbgn")
            return pd.DataFrame()

        derived_df = derived_components_df.copy()
        derived_df["derived_stock_component"] = derived_df["derived_stock_component"].astype(str).str.strip()

        allele_rows_unique = allele_rows.drop_duplicates(subset=["AlleleID"]).copy()
        allele_meta_by_id = allele_rows_unique.set_index("AlleleID").to_dict("index")

        allele_ids = set(allele_rows_unique["AlleleID"].astype(str).str.strip())
        construct_rows = construct_rows.drop_duplicates(subset=["AlleleID", "FBtpID"]).copy()
        insertion_rows = insertion_rows.drop_duplicates(subset=["AlleleID", "FBtiID"]).copy()
        construct_ids = set(construct_rows["FBtpID"].astype(str).str.strip()) if len(construct_rows) > 0 else set()
        insertion_ids = set(insertion_rows["FBtiID"].astype(str).str.strip()) if len(insertion_rows) > 0 else set()

        genes_with_alleles = gene_alleles["GeneID"].nunique()
        print(f"    Found {len(allele_ids)} alleles for {genes_with_alleles}/{len(set(input_genes))} input genes")
        print(f"    Linked component IDs: {len(construct_ids)} FBtp constructs, {len(insertion_ids)} FBti insertions")

        match_frames: List[pd.DataFrame] = []

        if allele_ids:
            allele_matches = derived_df[derived_df["derived_stock_component"].isin(allele_ids)].copy()
            if len(allele_matches) > 0:
                allele_lookup = allele_rows_unique[[
                    "AlleleID",
                    "FBgnID",
                    "GeneSymbol",
                    "AlleleSymbol",
                    "AlleleClassTerm",
                    "AlleleSupportingFBrf",
                ]].rename(
                    columns={
                        "AlleleID": "SourceAlleleID",
                        "FBgnID": "RelevantFBgnID",
                        "GeneSymbol": "RelevantGeneSymbol",
                        "AlleleSymbol": "SourceAlleleSymbol",
                        "AlleleClassTerm": "SourceAlleleClassTerm",
                        "AlleleSupportingFBrf": "SourceAlleleSupportingFBrf",
                    }
                )
                allele_matches = allele_matches.merge(
                    allele_lookup,
                    left_on="derived_stock_component",
                    right_on="SourceAlleleID",
                    how="left",
                )
                allele_matches["FBgnID"] = allele_matches["RelevantFBgnID"]
                allele_matches["GeneSymbol"] = allele_matches["RelevantGeneSymbol"]
                allele_matches["MatchComponentType"] = "FBal"
                match_frames.append(allele_matches)

        if construct_ids:
            construct_matches = derived_df[derived_df["derived_stock_component"].isin(construct_ids)].copy()
            if len(construct_matches) > 0:
                construct_lookup = construct_rows.rename(
                    columns={
                        "AlleleID": "SourceAlleleID",
                        "AlleleSymbol": "SourceAlleleSymbol",
                        "AlleleClassTerm": "SourceAlleleClassTerm",
                        "AlleleSupportingFBrf": "SourceAlleleSupportingFBrf",
                        "FBgnID": "RelevantFBgnID",
                        "GeneSymbol": "RelevantGeneSymbol",
                    }
                )
                construct_matches = construct_matches.merge(
                    construct_lookup,
                    left_on="derived_stock_component",
                    right_on="FBtpID",
                    how="left",
                )
                construct_matches["FBgnID"] = construct_matches["RelevantFBgnID"]
                construct_matches["GeneSymbol"] = construct_matches["RelevantGeneSymbol"]
                construct_matches["MatchComponentType"] = "FBtp"
                match_frames.append(construct_matches)

        if insertion_ids:
            insertion_matches = derived_df[derived_df["derived_stock_component"].isin(insertion_ids)].copy()
            if len(insertion_matches) > 0:
                insertion_lookup = insertion_rows.rename(
                    columns={
                        "AlleleID": "SourceAlleleID",
                        "AlleleSymbol": "SourceAlleleSymbol",
                        "AlleleClassTerm": "SourceAlleleClassTerm",
                        "AlleleSupportingFBrf": "SourceAlleleSupportingFBrf",
                        "FBgnID": "RelevantFBgnID",
                        "GeneSymbol": "RelevantGeneSymbol",
                    }
                )
                insertion_matches = insertion_matches.merge(
                    insertion_lookup,
                    left_on="derived_stock_component",
                    right_on="FBtiID",
                    how="left",
                )
                insertion_matches["FBgnID"] = insertion_matches["RelevantFBgnID"]
                insertion_matches["GeneSymbol"] = insertion_matches["RelevantGeneSymbol"]
                insertion_matches["MatchComponentType"] = "FBti"
                match_frames.append(insertion_matches)

        if not match_frames:
            print(
                f"    Warning: No FlyBase stocks found for {len(allele_ids)} alleles, "
                f"{len(construct_ids)} constructs, or {len(insertion_ids)} insertions"
            )
            return pd.DataFrame()

        matches = pd.concat(match_frames, ignore_index=True, sort=False).fillna("")
        matches = matches[matches["FBst"].astype(str).str.strip() != ""].copy()
        matches = matches.drop_duplicates(
            subset=["FBst", "derived_stock_component", "SourceAlleleID", "MatchComponentType"]
        )

        matched_genes = matches["FBgnID"].astype(str).str.strip().replace("", pd.NA).dropna().nunique()
        matching_fbsts = matches["FBst"].astype(str).str.strip().unique()
        print(f"    Found {len(matching_fbsts)} stocks for {matched_genes}/{len(set(input_genes))} input genes")

        all_rows = derived_df[derived_df["FBst"].isin(matching_fbsts)].copy()
        match_groups = {fbst: group.copy() for fbst, group in matches.groupby("FBst", sort=False)}

        def _gather_from_alleles(source_allele_ids: List[str], mapping: Dict[str, Dict[str, str]], key: str) -> str:
            values: List[str] = []
            for allele_id in source_allele_ids:
                raw_value = mapping.get(allele_id, {}).get(key, "")
                if not raw_value:
                    continue
                values.extend(parse_semicolon_list(raw_value))
            return unique_join(values)

        def _gather_from_allele_meta(source_allele_ids: List[str], field: str) -> str:
            values: List[str] = []
            for allele_id in source_allele_ids:
                raw_value = str(allele_meta_by_id.get(allele_id, {}).get(field, "") or "").strip()
                if not raw_value:
                    continue
                values.extend(parse_semicolon_list(raw_value))
            return unique_join(values)

        stock_rows: List[Dict[str, object]] = []
        for fbst, group in all_rows.groupby("FBst", sort=False):
            match_group = match_groups.get(fbst)
            if match_group is None or len(match_group) == 0:
                continue

            source_allele_ids = sorted({
                clean_id(value)
                for value in match_group["SourceAlleleID"].tolist()
                if clean_id(value)
            })
            relevant_fbal_ids = unique_join(source_allele_ids)
            relevant_fbal_symbols = _gather_from_allele_meta(source_allele_ids, "AlleleSymbol")
            relevant_fbtp_ids = _gather_from_alleles(source_allele_ids, constructs_by_allele, "IDs")
            relevant_fbtp_symbols = _gather_from_alleles(source_allele_ids, constructs_by_allele, "Symbols")
            relevant_fbti_ids = _gather_from_alleles(source_allele_ids, insertions_by_allele, "IDs")
            relevant_fbti_symbols = _gather_from_alleles(source_allele_ids, insertions_by_allele, "Symbols")
            relevant_component_ids = unique_join(
                parse_semicolon_list(relevant_fbal_ids)
                + parse_semicolon_list(relevant_fbtp_ids)
                + parse_semicolon_list(relevant_fbti_ids)
            )
            relevant_component_symbols = unique_join(
                parse_semicolon_list(relevant_fbal_symbols)
                + parse_semicolon_list(relevant_fbtp_symbols)
                + parse_semicolon_list(relevant_fbti_symbols)
            )
            relevant_gene_symbols = _gather_from_allele_meta(source_allele_ids, "GeneSymbol")
            relevant_flybase_gene_ids = _gather_from_allele_meta(source_allele_ids, "FBgnID")
            transgenic_product_class_terms = _gather_from_alleles(
                source_allele_ids, constructs_by_allele, "ClassTerm"
            )
            allele_class_terms = _gather_from_allele_meta(source_allele_ids, "AlleleClassTerm")
            construct_supporting_fbrfs = _gather_from_alleles(
                source_allele_ids, constructs_by_allele, "SupportingFBrf"
            )
            insertion_supporting_fbrfs = _gather_from_alleles(
                source_allele_ids, insertions_by_allele, "SupportingFBrf"
            )

            all_component_symbols = unique_join(group["object_symbol"].tolist())
            all_component_ids = unique_join(group["derived_stock_component"].tolist())
            all_stock_genes = unique_join(group["GeneSymbol"].tolist())
            all_stock_constructs = unique_join(
                group.loc[group["embedded_type"].isin(["FBti", "FBtp"]), "object_symbol"].tolist()
            )
            num_balancers = int((group["embedded_type"] == "FBba").sum())
            balancer_symbols = unique_join(
                group.loc[group["embedded_type"] == "FBba", "object_symbol"].tolist()
            )
            genotype = unique_join(group["FB_genotype"].tolist())
            collection = unique_join(group["collection"].tolist())
            stock_number = unique_join(group["stock_number"].tolist())

            stock_rows.append(
                {
                    "FBst": fbst,
                    "stock_number": stock_number,
                    "collection": collection,
                    "genotype": genotype,
                    "all_component_symbols": all_component_symbols,
                    "all_component_ids": all_component_ids,
                    "all_stock_genes": all_stock_genes,
                    "all_stock_constructs": all_stock_constructs,
                    "num_Balancers": num_balancers,
                    "Balancers": balancer_symbols or "-",
                    "FBti_count": int((group["embedded_type"] == "FBti").sum()),
                    "attP_count": str(genotype).lower().count("attp"),
                    "matched_component_ids": unique_join(match_group["derived_stock_component"].tolist()),
                    "matched_component_symbols": unique_join(match_group["object_symbol"].tolist()),
                    "matched_component_types": unique_join(match_group["MatchComponentType"].tolist()),
                    "relevant_component_symbols": relevant_component_symbols,
                    "relevant_component_ids": relevant_component_ids,
                    "relevant_gene_symbols": relevant_gene_symbols,
                    "relevant_flybase_gene_ids": relevant_flybase_gene_ids,
                    "relevant_fbal_symbols": relevant_fbal_symbols,
                    "relevant_fbal_ids": relevant_fbal_ids,
                    "relevant_fbtp_symbols": relevant_fbtp_symbols,
                    "relevant_fbtp_ids": relevant_fbtp_ids,
                    "relevant_fbti_symbols": relevant_fbti_symbols,
                    "relevant_fbti_ids": relevant_fbti_ids,
                    "transgenic_product_class_terms": transgenic_product_class_terms,
                    "allele_class_terms": allele_class_terms,
                    "construct_supporting_fbrfs": construct_supporting_fbrfs,
                    "insertion_supporting_fbrfs": insertion_supporting_fbrfs,
                }
            )

        result = pd.DataFrame(stock_rows)

        if fb_stocks_df is not None and len(fb_stocks_df) > 0:
            fb_meta = fb_stocks_df.copy()
            if "collection_short_name" in fb_meta.columns and "collection" not in fb_meta.columns:
                fb_meta = fb_meta.rename(columns={"collection_short_name": "collection"})
            for col in ("FBst", "stock_number", "collection", "FB_genotype"):
                if col not in fb_meta.columns:
                    fb_meta[col] = ""
            fb_meta["FBst"] = fb_meta["FBst"].apply(clean_id)
            fb_meta["stock_number"] = fb_meta["stock_number"].apply(clean_id)
            fb_meta = fb_meta[["FBst", "stock_number", "collection", "FB_genotype"]].drop_duplicates(
                subset=["FBst"]
            )
            fb_meta = fb_meta.rename(
                columns={
                    "stock_number": "stock_number_fb",
                    "collection": "collection_fb",
                    "FB_genotype": "genotype_fb",
                }
            )
            result = pd.merge(result, fb_meta, on="FBst", how="left")
            for current_col, refreshed_col in (
                ("stock_number", "stock_number_fb"),
                ("collection", "collection_fb"),
                ("genotype", "genotype_fb"),
            ):
                refreshed = result[refreshed_col].fillna("").astype(str).str.strip()
                current = result[current_col].fillna("").astype(str).str.strip()
                result[current_col] = refreshed.where(refreshed.astype(bool), current)
            result = result.drop(
                columns=["stock_number_fb", "collection_fb", "genotype_fb"],
                errors="ignore",
            )

        result["collection"] = result["collection"].fillna("").astype(str).str.strip()
        result["genotype"] = result["genotype"].fillna("").astype(str).str.strip()
        result["attP_count"] = result["genotype"].str.lower().str.count("attp")
        result["flybase_gene_id"] = result["relevant_flybase_gene_ids"]

        print(f"    Built stock mapping: {len(result)} stocks")
        return result

    def _build_custom_phenotype_rows(
        self,
        input_genes: List[str],
        fbal_to_fbgn_df: pd.DataFrame,
        stock_mapping: pd.DataFrame,
        phenotype_df: pd.DataFrame,
        transgenic_constructs_df: Optional[pd.DataFrame] = None,
        insertion_alleles_df: Optional[pd.DataFrame] = None,
        fbtp_to_fbti_df: Optional[pd.DataFrame] = None,
    ) -> pd.DataFrame:
        """Build custom_stock rows for phenotype-backed reagents without FBst stock."""
        import re

        if phenotype_df is None or len(phenotype_df) == 0:
            return pd.DataFrame()

        component_tables = self._build_gene_component_tables(
            input_genes,
            fbal_to_fbgn_df,
            transgenic_constructs_df,
            insertion_alleles_df,
            fbtp_to_fbti_df,
        )
        allele_rows = component_tables["alleles"]
        constructs_by_allele = component_tables["constructs_by_allele"]
        insertions_by_allele = component_tables["insertions_by_allele"]

        if len(allele_rows) == 0:
            return pd.DataFrame()

        stock_backed_allele_ids: Set[str] = set()
        if len(stock_mapping) > 0 and "relevant_fbal_ids" in stock_mapping.columns:
            for ids_str in stock_mapping["relevant_fbal_ids"].fillna("").astype(str):
                stock_backed_allele_ids.update(
                    cid for cid in parse_semicolon_list(ids_str) if cid
                )

        allele_meta_by_id = (
            allele_rows.drop_duplicates(subset=["AlleleID"])
            .set_index("AlleleID")
            .to_dict("index")
        )
        non_stock_allele_ids = set(allele_meta_by_id.keys()) - stock_backed_allele_ids

        if not non_stock_allele_ids:
            return pd.DataFrame()

        id_re = re.compile(r"FB[a-zA-Z]{2,4}[0-9]+")

        custom_data: Dict[str, Set[str]] = {}
        for _, pheno_row in phenotype_df.iterrows():
            raw_fbids = str(pheno_row.get("genotype_FBids", "") or "")
            ids_in_row = {clean_id(m) for m in id_re.findall(raw_fbids) if clean_id(m)}
            matched = ids_in_row & non_stock_allele_ids
            if not matched:
                continue
            genotype_symbols = str(pheno_row.get("genotype_symbols", "") or "").strip()
            for allele_id in matched:
                if allele_id not in custom_data:
                    custom_data[allele_id] = set()
                if genotype_symbols:
                    custom_data[allele_id].add(genotype_symbols)

        if not custom_data:
            return pd.DataFrame()

        custom_rows: List[Dict[str, object]] = []
        for allele_id, genotypes in custom_data.items():
            meta = allele_meta_by_id.get(allele_id)
            if not meta:
                continue
            allele_symbol = str(meta.get("AlleleSymbol", "") or "").strip()
            gene_id = str(meta.get("FBgnID", "") or "").strip()
            gene_symbol = str(meta.get("GeneSymbol", "") or "").strip()
            allele_class = str(meta.get("AlleleClassTerm", "") or "").strip()
            if not allele_symbol:
                continue

            genotype = unique_join(genotypes)

            construct_info = constructs_by_allele.get(allele_id, {})
            insertion_info = insertions_by_allele.get(allele_id, {})
            rel_fbtp_ids = construct_info.get("IDs", "")
            rel_fbtp_symbols = construct_info.get("Symbols", "")
            rel_fbti_ids = insertion_info.get("IDs", "")
            rel_fbti_symbols = insertion_info.get("Symbols", "")

            relevant_component_ids = unique_join(
                [allele_id]
                + parse_semicolon_list(rel_fbtp_ids)
                + parse_semicolon_list(rel_fbti_ids)
            )
            relevant_component_symbols = unique_join(
                [allele_symbol]
                + parse_semicolon_list(rel_fbtp_symbols)
                + parse_semicolon_list(rel_fbti_symbols)
            )

            custom_rows.append(
                {
                    "FBst": "",
                    "stock_number": allele_symbol,
                    "collection": "Custom phenotype reagent",
                    "genotype": genotype,
                    "all_component_symbols": allele_symbol,
                    "all_component_ids": allele_id,
                    "all_stock_genes": gene_symbol,
                    "all_stock_constructs": "",
                    "num_Balancers": 0,
                    "Balancers": "-",
                    "FBti_count": 0,
                    "attP_count": str(genotype).lower().count("attp"),
                    "matched_component_ids": allele_id,
                    "matched_component_symbols": allele_symbol,
                    "matched_component_types": allele_id[:4],
                    "relevant_component_symbols": relevant_component_symbols,
                    "relevant_component_ids": relevant_component_ids,
                    "relevant_gene_symbols": gene_symbol,
                    "relevant_flybase_gene_ids": gene_id,
                    "relevant_fbal_symbols": allele_symbol,
                    "relevant_fbal_ids": allele_id,
                    "relevant_fbtp_symbols": rel_fbtp_symbols,
                    "relevant_fbtp_ids": rel_fbtp_ids,
                    "relevant_fbti_symbols": rel_fbti_symbols,
                    "relevant_fbti_ids": rel_fbti_ids,
                    "transgenic_product_class_terms": construct_info.get(
                        "ClassTerm", ""
                    ),
                    "allele_class_terms": allele_class,
                    "construct_supporting_fbrfs": construct_info.get(
                        "SupportingFBrf", ""
                    ),
                    "insertion_supporting_fbrfs": insertion_info.get(
                        "SupportingFBrf", ""
                    ),
                    "custom_stock": True,
                }
            )

        result = pd.DataFrame(custom_rows)
        if len(result) > 0:
            result["flybase_gene_id"] = result["relevant_flybase_gene_ids"]
        print(f"    Built custom phenotype rows: {len(result)} non-stock reagents")
        return result

    def _refresh_stock_metadata_from_flybase(
        self,
        stocks_df: pd.DataFrame,
        fb_stocks_df: Optional[pd.DataFrame] = None,
    ) -> pd.DataFrame:
        """
        Refresh collection, stock number, and genotype from the latest FlyBase
        stocks TSV when those values are available.
        """
        if fb_stocks_df is None or len(fb_stocks_df) == 0:
            return stocks_df

        refreshed = stocks_df.copy()
        fb_meta = fb_stocks_df.copy()
        if "collection_short_name" in fb_meta.columns and "collection" not in fb_meta.columns:
            fb_meta = fb_meta.rename(columns={"collection_short_name": "collection"})
        for col in ("FBst", "stock_number", "collection", "FB_genotype"):
            if col not in fb_meta.columns:
                fb_meta[col] = ""
        fb_meta["FBst"] = fb_meta["FBst"].apply(clean_id)
        fb_meta["stock_number"] = fb_meta["stock_number"].apply(clean_id)
        fb_meta = fb_meta[["FBst", "stock_number", "collection", "FB_genotype"]].drop_duplicates(
            subset=["FBst"]
        )
        fb_meta = fb_meta.rename(
            columns={
                "stock_number": "stock_number_fb",
                "collection": "collection_fb",
                "FB_genotype": "genotype_fb",
            }
        )
        refreshed = pd.merge(refreshed, fb_meta, on="FBst", how="left")
        for current_col, refreshed_col in (
            ("stock_number", "stock_number_fb"),
            ("collection", "collection_fb"),
            ("genotype", "genotype_fb"),
        ):
            refreshed_values = refreshed[refreshed_col].fillna("").astype(str).str.strip()
            current_values = refreshed[current_col].fillna("").astype(str).str.strip()
            refreshed[current_col] = refreshed_values.where(refreshed_values.astype(bool), current_values)
        return refreshed.drop(
            columns=["stock_number_fb", "collection_fb", "genotype_fb"],
            errors="ignore",
        )
    
    def _get_references_for_stocks(
        self,
        stocks_df: pd.DataFrame,
        entity_pub: pd.DataFrame,
        ref_metadata: pd.DataFrame,
        keywords: List[str] = None,
        input_genes: List[str] = None
    ) -> Tuple[pd.DataFrame, pd.DataFrame, Dict[str, Set[str]], Dict[str, dict]]:
        """
        Get references for stocks from entity_publication.

        References are tracked separately for FBal, FBtp, and FBti component
        buckets, while Ref scoring continues to use the combined union.
        """
        stocks_df = stocks_df.copy()

        type_column_map = {
            "FBal": "relevant_fbal_ids",
            "FBtp": "relevant_fbtp_ids",
            "FBti": "relevant_fbti_ids",
        }
        type_pmid_cols = {
            "FBal": "FBal PMID",
            "FBtp": "FBtp PMID",
            "FBti": "FBti PMID",
        }
        type_no_pmid_cols = {
            "FBal": "FBal FBrf without PMID",
            "FBtp": "FBtp FBrf without PMID",
            "FBti": "FBti FBrf without PMID",
        }

        for component_type, id_col in type_column_map.items():
            if id_col not in stocks_df.columns:
                stocks_df[id_col] = ""
            stocks_df[id_col] = stocks_df[id_col].fillna("").astype(str)
            stocks_df[type_pmid_cols[component_type]] = ""
            stocks_df[type_no_pmid_cols[component_type]] = ""

        relevant_ids = set()
        for id_col in type_column_map.values():
            for ids_str in stocks_df[id_col].dropna():
                relevant_ids.update(parse_semicolon_list(ids_str))
        if not relevant_ids and "relevant_component_ids" in stocks_df.columns:
            for ids_str in stocks_df["relevant_component_ids"].dropna():
                relevant_ids.update(parse_semicolon_list(ids_str))

        effective_ref_metadata = ref_metadata.copy()
        for col in ['FBrf', 'PMID', 'PMCID', 'DOI', 'pub_type', 'miniref']:
            if col not in effective_ref_metadata.columns:
                effective_ref_metadata[col] = ''
            effective_ref_metadata[col] = effective_ref_metadata[col].fillna('').astype(str).str.strip()
        effective_ref_metadata['FBrf'] = effective_ref_metadata['FBrf'].apply(clean_id)
        effective_ref_metadata['PMID'] = effective_ref_metadata['PMID'].apply(clean_id)
        effective_ref_metadata['PMCID'] = effective_ref_metadata['PMCID'].apply(clean_id)
        effective_ref_metadata['DOI'] = effective_ref_metadata['DOI'].apply(clean_id)
        effective_ref_metadata['pub_type'] = effective_ref_metadata['pub_type'].str.lower()

        max_genes_per_pub = 50
        if 'pub_type' in effective_ref_metadata.columns:
            paper_fbrfs = set(
                effective_ref_metadata.loc[
                    effective_ref_metadata['pub_type'] == 'paper',
                    'FBrf'
                ].dropna()
            )
        else:
            paper_fbrfs = set(entity_pub['FlyBase_publication_id'].dropna().unique())

        gene_entity_pub = entity_pub[entity_pub['entity_id'].str.startswith('FBgn', na=False)]
        gene_counts = gene_entity_pub.groupby('FlyBase_publication_id')['entity_name'].nunique()
        low_gene_pubs = set(gene_counts[gene_counts <= max_genes_per_pub].index)
        pubs_with_genes = set(gene_counts.index)
        pubs_without_genes = set(entity_pub['FlyBase_publication_id'].dropna().unique()) - pubs_with_genes
        low_gene_pubs = low_gene_pubs | pubs_without_genes
        excluded_pubs = len(pubs_with_genes) - len(gene_counts[gene_counts <= max_genes_per_pub])

        valid_pubs = paper_fbrfs & low_gene_pubs
        print(f"    Pre-filtered references: {len(valid_pubs)} papers with ≤{max_genes_per_pub} associated genes (excluded {excluded_pubs} large-scale publications)")

        entity_pub_filtered = entity_pub[
            entity_pub['FlyBase_publication_id'].isin(valid_pubs) &
            entity_pub['entity_id'].isin(relevant_ids)
        ].copy()

        if len(entity_pub_filtered) == 0:
            stocks_df['PMID'] = ""
            stocks_df['PMCID'] = ""
            stocks_df['FBrf without PMID'] = ""
            stocks_df['references_without_pmid_count'] = 0
            stocks_df['ref_count'] = 0
            stocks_df['keyword_ref_count'] = 0
            stocks_df['keyword_ref_pmids'] = ''
            stocks_df['total_refs'] = 0
            self._no_pmid_fbrf_records = []
            if keywords:
                kw_col_name = generate_keyword_column_name(keywords)
                stocks_df[kw_col_name] = False
            return stocks_df, pd.DataFrame(), {}, {}, {}

        fbrfs = [clean_id(x) for x in entity_pub_filtered['FlyBase_publication_id'].dropna().unique()]
        refs_df = effective_ref_metadata[effective_ref_metadata['FBrf'].isin(fbrfs)].copy()
        if 'pub_type' in refs_df.columns:
            refs_df = refs_df[refs_df['pub_type'] == 'paper'].copy()

        fbrf_to_pmid = dict(zip(refs_df['FBrf'], refs_df['PMID'].apply(clean_id)))
        fbrf_to_pmcid = dict(zip(
            refs_df['FBrf'],
            refs_df['PMCID'].apply(clean_id) if 'PMCID' in refs_df.columns else [''] * len(refs_df)
        ))

        component_id_to_symbol: Dict[str, str] = {}
        for _, eprow in entity_pub_filtered.iterrows():
            eid = str(eprow.get('entity_id', '')).strip()
            ename = str(eprow.get('entity_name', '')).strip()
            if eid and ename:
                component_id_to_symbol[eid] = ename

        comp_fbrfs = entity_pub_filtered.groupby('entity_id')['FlyBase_publication_id'].apply(set).to_dict()

        component_to_pmids: Dict[str, Set[str]] = {}
        component_to_fbrfs_without_pmid: Dict[str, Set[str]] = {}
        for comp, fbrf_set in comp_fbrfs.items():
            pmids = set()
            no_pmid_fbrfs = set()
            for fbrf_raw in fbrf_set:
                fbrf = clean_id(fbrf_raw)
                pmid = fbrf_to_pmid.get(fbrf)
                if pmid:
                    pmids.add(pmid)
                else:
                    no_pmid_fbrfs.add(fbrf)
            if pmids:
                component_to_pmids[comp] = pmids
            if no_pmid_fbrfs:
                component_to_fbrfs_without_pmid[comp] = no_pmid_fbrfs

        def _collect_component_pmids(component_ids: List[str]) -> str:
            stock_pmids: Set[str] = set()
            for cid in component_ids:
                stock_pmids.update(component_to_pmids.get(cid, set()))
            return unique_join(list(stock_pmids))

        def _collect_component_no_pmid_fbrfs(component_ids: List[str]) -> str:
            fbrfs_without_pmid: Set[str] = set()
            for cid in component_ids:
                fbrfs_without_pmid.update(component_to_fbrfs_without_pmid.get(cid, set()))
            return unique_join(list(fbrfs_without_pmid))

        pmid_to_stocks: Dict[str, Set[str]] = {}
        pmid_to_genes: Dict[str, Set[str]] = {}
        for _, row in stocks_df.iterrows():
            stock_num = clean_id(row['stock_number'])
            comp_ids = parse_semicolon_list(row.get('relevant_component_ids', ''))
            genes = parse_semicolon_list(row.get('relevant_gene_symbols', ''))
            for cid in comp_ids:
                if cid in component_to_pmids:
                    for pmid in component_to_pmids[cid]:
                        pmid_to_stocks.setdefault(pmid, set()).add(stock_num)
                        pmid_to_genes.setdefault(pmid, set()).update(genes)

        for component_type, id_col in type_column_map.items():
            pmid_col = type_pmid_cols[component_type]
            no_pmid_col = type_no_pmid_cols[component_type]
            stocks_df[pmid_col] = stocks_df[id_col].apply(
                lambda value: _collect_component_pmids(parse_semicolon_list(value))
            )
            stocks_df[no_pmid_col] = stocks_df[id_col].apply(
                lambda value: _collect_component_no_pmid_fbrfs(parse_semicolon_list(value))
            )

        def _union_join_from_columns(row: pd.Series, columns: List[str]) -> str:
            values: List[str] = []
            for col in columns:
                values.extend(parse_semicolon_list(row.get(col, '')))
            return unique_join(values)

        stocks_df['PMID'] = stocks_df.apply(
            lambda row: _union_join_from_columns(row, list(type_pmid_cols.values())),
            axis=1,
        )
        stocks_df['FBrf without PMID'] = stocks_df.apply(
            lambda row: _union_join_from_columns(row, list(type_no_pmid_cols.values())),
            axis=1,
        )
        def _collect_component_pmcids(component_ids: List[str]) -> str:
            pmcids: Set[str] = set()
            for cid in component_ids:
                for fbrf in comp_fbrfs.get(cid, set()):
                    pmcid = fbrf_to_pmcid.get(clean_id(fbrf), '')
                    if pmcid:
                        pmcids.add(pmcid)
            return unique_join(list(pmcids))

        stocks_df['PMCID'] = stocks_df['relevant_component_ids'].apply(
            lambda value: _collect_component_pmcids(parse_semicolon_list(value))
        )
        stocks_df['references_without_pmid_count'] = stocks_df['FBrf without PMID'].apply(
            lambda x: len([f for f in str(x).split(';') if f.strip()]) if pd.notna(x) and x else 0
        )
        stocks_df['total_refs'] = stocks_df.apply(
            lambda row: (
                len([p for p in str(row.get('PMID', '')).split(';') if p.strip()]) +
                len([f for f in str(row.get('FBrf without PMID', '')).split(';') if f.strip()])
            ),
            axis=1
        )
        stocks_df['ref_count'] = stocks_df['total_refs']

        no_pmid_refs_df = refs_df[refs_df['PMID'].astype(str).str.strip() == ''].copy()
        no_pmid_records: List[Dict[str, str]] = []
        if len(no_pmid_refs_df) > 0:
            fbrf_to_components: Dict[str, Set[str]] = {}
            for cid, fset in comp_fbrfs.items():
                symbol = component_id_to_symbol.get(cid, cid)
                for f in fset:
                    f_clean = clean_id(f)
                    fbrf_to_components.setdefault(f_clean, set()).add(symbol)
            fbrf_to_stocks: Dict[str, Set[str]] = {}
            for _, row in stocks_df.iterrows():
                stock_num = clean_id(row.get('stock_number', ''))
                for f in parse_semicolon_list(row.get('FBrf without PMID', '')):
                    f_clean = clean_id(f)
                    if not f_clean:
                        continue
                    fbrf_to_stocks.setdefault(f_clean, set()).add(stock_num)

            for _, row in no_pmid_refs_df.iterrows():
                fbrf = clean_id(row.get('FBrf', ''))
                if not fbrf:
                    continue
                no_pmid_records.append({
                    'FBrf': fbrf,
                    'pub_type': str(row.get('pub_type', '')).strip(),
                    'miniref': str(row.get('miniref', '')).strip(),
                    'associated_stocks': ', '.join(sorted(fbrf_to_stocks.get(fbrf, set()))),
                    'associated_components': ', '.join(sorted(fbrf_to_components.get(fbrf, set()))),
                })
        self._no_pmid_fbrf_records = no_pmid_records

        unique_pmids = set()
        for pmids_set in component_to_pmids.values():
            unique_pmids.update(pmids_set)
        unique_pmids_list = sorted(unique_pmids)
        pubmed_metadata = self._pubmed_client.fetch_metadata(unique_pmids_list)

        kw_col_name = None
        keyword_pmids: Set[str] = set()
        if keywords:
            kw_col_name = generate_keyword_column_name(keywords)
            for pmid in unique_pmids:
                if self._pubmed_client.check_keywords(pmid, keywords, pubmed_metadata):
                    keyword_pmids.add(pmid)

            def check_keyword_refs(pmids_str):
                if pd.isna(pmids_str) or not str(pmids_str).strip():
                    return False
                pmids = [p.strip() for p in str(pmids_str).split(';') if p.strip()]
                return any(pmid in keyword_pmids for pmid in pmids)

            def count_keyword_refs(pmids_str):
                if pd.isna(pmids_str) or not str(pmids_str).strip():
                    return 0
                pmids = [p.strip() for p in str(pmids_str).split(';') if p.strip()]
                return sum(1 for pmid in pmids if pmid in keyword_pmids)

            def collect_keyword_pmids(pmids_str):
                if pd.isna(pmids_str) or not str(pmids_str).strip():
                    return ''
                pmids = [p.strip() for p in str(pmids_str).split(';') if p.strip()]
                matching = [pmid for pmid in pmids if pmid in keyword_pmids]
                return '; '.join(sorted(matching))

            stocks_df[kw_col_name] = stocks_df['PMID'].apply(check_keyword_refs)
            stocks_df['keyword_ref_count'] = stocks_df['PMID'].apply(count_keyword_refs)
            stocks_df['keyword_ref_pmids'] = stocks_df['PMID'].apply(collect_keyword_pmids)
            refs_df[kw_col_name] = refs_df['PMID'].apply(
                lambda x: clean_id(x) in keyword_pmids if pd.notna(x) else False
            )
        else:
            stocks_df['keyword_ref_count'] = 0
            stocks_df['keyword_ref_pmids'] = ''

        if len(refs_df) > 0:
            refs_df['associated_stocks'] = refs_df['PMID'].apply(
                lambda x: ', '.join(sorted(pmid_to_stocks.get(clean_id(x), set()))) if pd.notna(x) else ''
            )

            input_gene_symbols = set()
            if input_genes:
                for _, row in stocks_df.iterrows():
                    input_gene_symbols.update(parse_semicolon_list(row.get('relevant_gene_symbols', '')))

            def get_associated_genes(pmid):
                if pd.isna(pmid):
                    return ''
                pmid_clean = clean_id(pmid)
                genes = pmid_to_genes.get(pmid_clean, set())
                if input_gene_symbols:
                    genes = genes.intersection(input_gene_symbols)
                return ', '.join(sorted(genes))

            refs_df['associated_genes'] = refs_df['PMID'].apply(get_associated_genes)
            refs_df['stock_count'] = refs_df['PMID'].apply(
                lambda x: len(pmid_to_stocks.get(clean_id(x), set())) if pd.notna(x) else 0
            )

            def _get_pubmed_meta(pmid):
                if pd.isna(pmid):
                    return {}
                pmid_clean = clean_id(pmid)
                meta = pubmed_metadata.get(pmid_clean, {})
                return meta if isinstance(meta, dict) else {}

            def get_pub_date(pmid):
                meta = _get_pubmed_meta(pmid)
                return str(meta.get('year', '') or '')

            def get_authors(pmid):
                meta = _get_pubmed_meta(pmid)
                authors = meta.get('authors', [])
                if isinstance(authors, str):
                    return '; '.join([a.strip() for a in authors.split(';') if a.strip()])
                if isinstance(authors, list):
                    return '; '.join([str(a).strip() for a in authors if str(a).strip()])
                return ''

            refs_df['publication_date'] = refs_df['PMID'].apply(get_pub_date)
            refs_df['author(s)'] = refs_df['PMID'].apply(get_authors)
            refs_df['journal'] = refs_df['PMID'].apply(
                lambda x: str(_get_pubmed_meta(x).get('journal', '') or '')
            )
            refs_df['title'] = refs_df['PMID'].apply(
                lambda x: pubmed_metadata.get(clean_id(x), {}).get('title', '') if pd.notna(x) else ''
            )

        return stocks_df, refs_df, component_to_pmids, component_id_to_symbol, pubmed_metadata

    def _run_functional_validation(
        self,
        stocks_df: pd.DataFrame,
        references_df: pd.DataFrame,
        component_to_pmids: Dict[str, Set[str]],
        component_id_to_symbol: Dict[str, str],
        keywords: List[str] = None,
        soft_run: bool = False
    ) -> pd.DataFrame:
        """Run functional validation on keyword-hit references."""
        return run_functional_validation(
            stocks_df=stocks_df,
            references_df=references_df,
            keywords=keywords,
            soft_run=soft_run,
            validator=self._validator,
            fulltext_fetcher=self._fulltext_fetcher,
            pubmed_client=self._pubmed_client,
            pubmed_cache=self._pubmed_cache,
            max_gpt_calls_per_stock=self.settings.max_gpt_calls_per_stock,
            component_to_pmids=component_to_pmids,
            component_id_to_symbol=component_id_to_symbol,
        )
    
    def run(
        self,
        input_dir: Path,
        keywords: Optional[List[str]] = None,
        gene_col: str = "flybase_gene_id",
        input_gene_col: str = "ext_gene",
        skip_fbgnid_conversion: bool = False,
        run_functional_validation: bool = False,
    ) -> Optional[Path]:
        """
        Run the full pipeline.
        
        Args:
            input_dir: Directory containing CSV files with gene IDs
            keywords: List of keywords for filtering references
            gene_col: Column name for FlyBase gene IDs (after conversion)
            input_gene_col: Column name for gene symbols (before conversion)
            skip_fbgnid_conversion: Skip FBgnID conversion step
            run_functional_validation: Run GPT-based functional validation now
        
        Returns:
            Path to output Excel file, or None if failed
        """
        input_dir = Path(input_dir)
        if not input_dir.is_dir():
            raise ValueError(f"Input path is not a directory: {input_dir}")
        
        if keywords:
            print(f"Keywords: {', '.join(keywords)}")
        
        # Step 0: FBgnID conversion
        if not skip_fbgnid_conversion and not self.settings.skip_fbgnid_conversion:
            self._run_fbgnid_conversion(input_dir, input_gene_col)
            
            # Validate
            print("\nValidating FBgnID conversion...")
            csv_files = sorted(input_dir.glob("*.csv"))
            all_invalid = []
            for csv_path in csv_files:
                invalid = self._validate_fbgnids(csv_path, gene_col)
                if invalid:
                    all_invalid.extend(invalid)
                    print(f"  {csv_path.name}: {len(invalid)} genes without valid FBgnID")
            
            if all_invalid:
                print(f"\nERROR: {len(set(all_invalid))} genes without FBgnIDs")
                return None
            else:
                print("  All genes successfully mapped")
        
        # Create output directory
        output_dir = input_dir / "Stocks"
        output_dir.mkdir(exist_ok=True)
        
        # Aggregate genes
        unique_genes, valid_files = self._aggregate_input_genes(input_dir, gene_col)
        
        if not unique_genes:
            print("No genes found in input files")
            return None
        
        # Load tables
        tables = self._load_tables()
        
        # Build stock mapping
        print(f"\n{'='*60}")
        print("Building stock mapping...")
        stock_mapping = self._build_stock_mapping(
            unique_genes,
            tables["derived_stock_components"],
            tables["fbal_to_fbgn"],
            tables.get("transgenic_construct_descriptions"),
            tables.get("insertion_allele_descriptions"),
            tables.get("fbtp_to_fbti"),
            tables.get("fb_stocks"),
        )

        if len(stock_mapping) > 0:
            stock_mapping["custom_stock"] = False

        phenotype_df = tables.get("genotype_phenotype_data")
        if phenotype_df is not None and len(phenotype_df) > 0:
            custom_phenotype_rows = self._build_custom_phenotype_rows(
                unique_genes,
                tables["fbal_to_fbgn"],
                stock_mapping,
                phenotype_df,
                tables.get("transgenic_construct_descriptions"),
                tables.get("insertion_allele_descriptions"),
                tables.get("fbtp_to_fbti"),
            )
            if len(custom_phenotype_rows) > 0:
                stock_mapping = pd.concat(
                    [stock_mapping, custom_phenotype_rows],
                    ignore_index=True,
                    sort=False,
                )
                stock_mapping["custom_stock"] = stock_mapping["custom_stock"].fillna(False)

        if len(stock_mapping) == 0:
            print("No stocks or custom phenotype reagents found for input genes")
            return None
        
        # Get references
        print(f"\nGetting references...")
        stock_mapping, references_df, component_to_pmids, component_id_to_symbol, pubmed_metadata = self._get_references_for_stocks(
            stock_mapping,
            tables["entity_publication"],
            tables["ref_metadata"],
            keywords,
            input_genes=unique_genes
        )
        
        # Add allele symbol for downstream pipelines (before boolean checks so it's available)
        stock_mapping['AlleleSymbol'] = stock_mapping.get('relevant_fbal_symbols', stock_mapping['relevant_component_symbols'])

        genotype_series = stock_mapping['genotype'].fillna('').astype(str)
        collection_series = stock_mapping['collection'].fillna('').astype(str)

        stock_mapping['UAS'] = genotype_series.str.contains('uas', case=False, na=False)
        uas_count = stock_mapping['UAS'].sum()
        print(f"    UAS flag: {uas_count} stocks marked as UAS")

        stock_mapping['sgRNA'] = genotype_series.str.contains('sgrna', case=False, na=False)
        sgrna_count = stock_mapping['sgRNA'].sum()
        print(f"    sgRNA flag: {sgrna_count} stocks marked as sgRNA")

        is_vienna = collection_series.str.contains('vienna', case=False, na=False)
        transgenic_product_class_series = stock_mapping['transgenic_product_class_terms'].fillna('').astype(str)
        has_vdrc_knockdown_family = (
            genotype_series.str.contains(r'P\{GD\d+', regex=True, na=False)
            | genotype_series.str.contains(r'P\{KK\d+', regex=True, na=False)
            | genotype_series.str.contains(r'P\{VSH\d+', regex=True, na=False)
        )
        has_rnai_reagent_class = transgenic_product_class_series.str.contains(
            'rnai_reagent',
            case=False,
            na=False,
        )
        stock_mapping['RNAi'] = (
            genotype_series.str.contains('uas', case=False, na=False)
            | genotype_series.str.contains('rnai', case=False, na=False)
            | (is_vienna & has_vdrc_knockdown_family)
            | has_rnai_reagent_class
        )
        rnai_count = int(stock_mapping['RNAi'].sum())
        print(f"    RNAi proxy: {rnai_count} stocks marked as RNAi")
        
        # Rename keyword_ref_count and keyword_ref_pmids to include the actual keywords
        if keywords:
            kw_ref_count_col = f"{' OR '.join(keywords)} references count"
            kw_ref_pmids_col = f"Stock (FBal / FBtp / FBti) {' OR '.join(keywords)} references"
            rename_map = {}
            if 'keyword_ref_count' in stock_mapping.columns:
                rename_map['keyword_ref_count'] = kw_ref_count_col
            if 'keyword_ref_pmids' in stock_mapping.columns:
                rename_map['keyword_ref_pmids'] = kw_ref_pmids_col
            if rename_map:
                stock_mapping = stock_mapping.rename(columns=rename_map)
        
        # Run functional validation (now optional; typically executed in split-stocks)
        if run_functional_validation:
            print(f"\n{'='*60}")
            print("Running functional validation...")
            stock_mapping = self._run_functional_validation(
                stock_mapping,
                references_df,
                component_to_pmids,
                component_id_to_symbol,
                keywords,
                soft_run=self.settings.soft_run
            )
        else:
            print(f"\n{'='*60}")
            print("Skipping functional validation in find-stocks.")
            print("  Validation will be performed in validate-stocks for Ref++ output-sheet stocks.")
        
        # Save fulltext method cache and clean up
        self._fulltext_fetcher.save_cache()
        self._fulltext_fetcher.clear_cache()
        gc.collect()

        # Prepare Stocks sheet formatting
        if keywords is not None:
            old_kw_col = generate_keyword_column_name(keywords)
            if keywords:
                new_kw_col = f"Stock (FBal / FBtp / FBti) associated with at least 1 {' OR '.join(keywords)} reference?"
            else:
                new_kw_col = "Stock (allele) associated with at least 1 reference?"
            if old_kw_col in stock_mapping.columns:
                stock_mapping = stock_mapping.rename(columns={old_kw_col: new_kw_col})
        
        # Remove unwanted columns from Stocks sheet
        cols_to_drop = [
            'all_flybase_gene_ids',
            'all_gene_symbols',
            'flybase_gene_id',
            'PMCID',
            'pmcid',
            'ref_count',
            'is_custom_stock'
        ]
        stock_mapping = stock_mapping.drop(columns=[c for c in cols_to_drop if c in stock_mapping.columns])
        
        # Reorder columns for Stocks sheet readability
        preferred_order = [
            'FBst',
            'stock_number',
            'collection',
            'genotype',
            'num_Balancers',
            'Balancers',
            'FBti_count',
            'attP_count',
            'matched_component_symbols',
            'matched_component_ids',
            'matched_component_types',
            'relevant_fbal_symbols',
            'relevant_fbal_ids',
            'relevant_fbtp_symbols',
            'relevant_fbtp_ids',
            'relevant_fbti_symbols',
            'relevant_fbti_ids',
            'relevant_component_symbols',
            'relevant_component_ids',
            'relevant_gene_symbols',
            'relevant_flybase_gene_ids',
            'transgenic_product_class_terms',
            'allele_class_terms',
            'FBal PMID',
            'FBtp PMID',
            'FBti PMID',
            'FBal FBrf without PMID',
            'FBtp FBrf without PMID',
            'FBti FBrf without PMID',
            'all_component_symbols',
            'all_component_ids',
            'all_stock_genes',
            'all_stock_constructs',
            'custom_stock',
            'UAS',
            'sgRNA',
            'RNAi',
        ]
        ordered_cols = [c for c in preferred_order if c in stock_mapping.columns]
        for col in stock_mapping.columns:
            if col not in ordered_cols:
                ordered_cols.append(col)
        stock_mapping = stock_mapping[ordered_cols]
        
        # Process References sheet
        if references_df is not None and len(references_df) > 0:
            # Remove unwanted columns: FBrf, PMCID, miniref, pmid_added
            cols_to_remove = ['FBrf', 'PMCID', 'miniref', 'pmid_added', 'source', 'resolved_at', 'has_pmid', 'has_pubtype']
            for col in cols_to_remove:
                if col in references_df.columns:
                    references_df = references_df.drop(columns=[col])
            
            # Get keyword column name
            kw_col_name = generate_keyword_column_name(keywords) if keywords else None
            
            # Sort references: first by keyword hits (True first), then by stock_count (descending)
            if kw_col_name and kw_col_name in references_df.columns:
                references_df = references_df.sort_values(
                    by=[kw_col_name, 'stock_count'],
                    ascending=[False, False]
                ).reset_index(drop=True)
            elif 'stock_count' in references_df.columns:
                references_df = references_df.sort_values(
                    by=['stock_count'],
                    ascending=[False]
                ).reset_index(drop=True)
            
            # Reorder columns for better readability
            preferred_order = [
                'PMID', 'title', 'associated_stocks', 'associated_genes', 
                'stock_count', 'publication_date', 'author(s)', 'journal'
            ]
            if kw_col_name:
                preferred_order.append(kw_col_name)
            
            # Build final column order
            final_cols = []
            for col in preferred_order:
                if col in references_df.columns:
                    final_cols.append(col)
            # Add any remaining columns
            for col in references_df.columns:
                if col not in final_cols:
                    final_cols.append(col)
            references_df = references_df[final_cols]
        
        # Save output
        output_path = output_dir / "aggregated_stock_refs.xlsx"
        print(f"\nSaving output to: {output_path}")
        
        if self.settings.soft_run:
            # Soft run: drop GPT-derived columns entirely
            gpt_cols_to_drop = get_gpt_derived_columns(stock_mapping)
            stock_mapping_out = stock_mapping.drop(
                columns=[c for c in gpt_cols_to_drop if c in stock_mapping.columns]
            )
        else:
            # Apply [EXPERIMENTAL] prefix to GPT-derived column headers
            stock_mapping_out = apply_experimental_prefix(stock_mapping)
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            stock_mapping_out.to_excel(writer, sheet_name='Stocks', index=False)
            if references_df is not None and len(references_df) > 0:
                references_df.to_excel(writer, sheet_name='References', index=False)
            
            if not self.settings.soft_run:
                # Apply light grey fill to GPT-derived column headers only
                _apply_grey_fill_openpyxl(
                    writer.sheets['Stocks'], stock_mapping_out
                )
        
        # Write surviving filtered FBrfs without PMID to a sidecar report.
        no_pmid_report = output_dir / "references_without_pmid_fbrf.txt"
        no_pmid_rows = sorted(
            self._no_pmid_fbrf_records,
            key=lambda r: (r.get('associated_stocks', ''), r.get('FBrf', ''))
        )
        with open(no_pmid_report, 'w', encoding='utf-8') as f:
            f.write("Filtered-in FlyBase references without PMID\n")
            f.write(f"Generated: {datetime.now().isoformat(timespec='seconds')}\n")
            f.write(f"Count: {len(no_pmid_rows)}\n")
            f.write("\n")
            for row in no_pmid_rows:
                f.write(f"FBrf: {row.get('FBrf', '')}\n")
                f.write(f"  Stocks: {row.get('associated_stocks', '')}\n")
                f.write(f"  Components: {row.get('associated_components', '')}\n")
                f.write(f"  pub_type: {row.get('pub_type', '')}\n")
                if row.get('miniref', ''):
                    f.write(f"  miniref: {row.get('miniref', '')}\n")
                f.write("\n")
        print(f"  Saved no-PMID FBrf report: {no_pmid_report}")
        
        print(f"\n{'='*60}")
        print("COMPLETE")
        print(f"  Input files: {len(valid_files)}")
        print(f"  Total genes: {len(unique_genes)}")
        print(f"  Output: {output_path}")
        
        return output_path
