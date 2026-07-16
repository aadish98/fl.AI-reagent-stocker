import pandas as pd
from openpyxl import load_workbook

from scripts.export_gal4_driver_workbook import (
    CUSTOM_LAB_REAGENT_COLLECTION,
    Gal4StockMatch,
    INPUT_SYMBOLS_COLUMN,
    STOCK_ID_COLUMN,
    _build_gal4_construct_alias_fbti_index,
    _build_gal4_lookup_context,
    _build_coverage_sheet,
    _build_main_sheet,
    _build_references_sheet,
    _build_collection_sheets,
    _gal4_driver_symbols_match,
    _gal4_promoter_keys,
    _gal4_search_keys,
    _is_split_gal4_component,
    _resolve_gal4_symbol_to_stocks,
    write_gal4_workbook,
)


def test_gal4_search_keys_preserve_gawb_bracket_driver_alias():
    keys = _gal4_search_keys("P{GawB}elav[C155]")

    assert "elavc155" in keys


def test_gal4_search_keys_preserve_gmr_r_line_alias():
    keys = _gal4_search_keys("P{GMR23E10-GAL4}attP2")

    assert "r23e10gal4" in keys


def test_split_gal4_dbd_component_is_detected_for_resolver_guard():
    assert _gal4_driver_symbols_match(
        "R23E10-GAL4",
        "P{R23E10-GAL4.DBD}attP2",
    )
    assert _is_split_gal4_component("P{R23E10-GAL4.DBD}attP2")


def test_mef2_pr_resolves_through_construct_alias_to_bloomington_600193():
    derived_df = pd.DataFrame(
        [
            {
                "FBst": "FBst0600193",
                "stock_number": "600193",
                "collection": "Bloomington",
                "derived_stock_component": "FBti0230300",
                "object_symbol": "P{GAL4-Mef2.R}39-1",
                "GeneSymbol": "",
                "FB_genotype": "w[*]; P{w[+mC]=GAL4-Mef2.R}39-1",
            }
        ]
    )
    construct_df = pd.DataFrame(
        [
            {
                "Component Allele (symbol)": "Scer\\GAL4[Mef2.PR]",
                "Transgenic Construct (symbol)": "P{GAL4-Mef2.R}",
                "Transgenic Construct (id)": "FBtp0006434",
            }
        ]
    )
    fbtp_to_fbti_df = pd.DataFrame(
        [{"FBtp": "FBtp0006434", "FBti": "FBti0230300"}]
    )
    construct_alias_index = _build_gal4_construct_alias_fbti_index(
        construct_df,
        fbtp_to_fbti_df,
    )
    (
        component_id_to_symbol,
        component_id_to_gene_symbol,
        component_id_to_stock_candidate_details,
        component_rows_by_fbst,
        gal4_only_fbst,
        stock_candidate_label_to_driver_genotype,
        driver_stock_index,
        fbst_to_fb_genotype,
        resolution_index,
    ) = _build_gal4_lookup_context(derived_df)

    matches = _resolve_gal4_symbol_to_stocks(
        "Mef2.PR",
        gal4_only_fbst,
        driver_stock_index,
        component_id_to_symbol,
        component_id_to_gene_symbol,
        component_id_to_stock_candidate_details,
        component_rows_by_fbst,
        stock_candidate_label_to_driver_genotype,
        fbst_to_fb_genotype,
        resolution_index,
        construct_alias_index,
    )

    assert construct_alias_index["mef2pr"] == {"FBti0230300"}
    assert [(match.stock_number, match.collection) for match in matches] == [
        ("600193", "Bloomington")
    ]
    assert matches[0].driver_genotype == "P{GAL4-Mef2.R}39-1"


def test_generic_pu_drivers_resolve_to_simple_bdsc_promoter_family_stocks():
    rows = [
        ("FBst0051635", "51635", "P{nSyb-GAL4.S}3", "y[1] w[*]; P{nSyb-GAL4.S}3"),
        ("FBst0006479", "6479", "P{GawB}sca[109-68]", "y[1] w[*]; P{GawB}sca[109-68]/CyO"),
        ("FBst0008765", "8765", "P{GAL4-elav.L}2", "P{GAL4-elav.L}2/CyO"),
        (
            "FBst0042714",
            "42714",
            "P{nSyb-GAL4.S}3",
            "P{UAS-GFP-LAMP}2; P{nSyb-GAL4.S}3",
        ),
    ]
    derived_rows = []
    for index, (fbst, stock_number, symbol, genotype) in enumerate(rows):
        derived_rows.append(
            {
                "FBst": fbst,
                "stock_number": stock_number,
                "collection": "Bloomington",
                "derived_stock_component": f"FBti{index:07d}",
                "object_symbol": symbol,
                "GeneSymbol": "GAL4",
                "FB_genotype": genotype,
            }
        )
        if stock_number == "42714":
            derived_rows.append(
                {
                    "FBst": fbst,
                    "stock_number": stock_number,
                    "collection": "Bloomington",
                    "derived_stock_component": "FBti9999999",
                    "object_symbol": "P{UAS-GFP-LAMP}2",
                    "GeneSymbol": "Hsap\\LAMP1",
                    "FB_genotype": genotype,
                }
            )
    (
        component_id_to_symbol,
        component_id_to_gene_symbol,
        component_id_to_stock_candidate_details,
        component_rows_by_fbst,
        gal4_only_fbst,
        stock_candidate_label_to_driver_genotype,
        driver_stock_index,
        fbst_to_fb_genotype,
        resolution_index,
    ) = _build_gal4_lookup_context(pd.DataFrame(derived_rows))

    expected = {
        "nSyb.PU": "51635",
        "sca.PU": "6479",
        "elav.PU": "8765",
    }
    resolved_matches = []
    for input_symbol, stock_number in expected.items():
        matches = _resolve_gal4_symbol_to_stocks(
            input_symbol,
            gal4_only_fbst,
            driver_stock_index,
            component_id_to_symbol,
            component_id_to_gene_symbol,
            component_id_to_stock_candidate_details,
            component_rows_by_fbst,
            stock_candidate_label_to_driver_genotype,
            fbst_to_fb_genotype,
            resolution_index,
            {},
        )
        assert [match.stock_number for match in matches] == [stock_number]
        assert matches[0].resolution_method == "BDSC promoter-family substitution"
        resolved_matches.extend(matches)

    assert _gal4_promoter_keys("P{GawB}sca[109-68]") == {"sca"}
    coverage = _build_coverage_sheet(
        [
            *resolved_matches,
            Gal4StockMatch(
                input_symbol="missing.PU",
                fbst="",
                stock_number="",
                collection="",
                source_stock="",
                resolution_method="Unresolved",
                match_confidence="No orderable stock candidate found",
            ),
        ]
    )
    statuses = dict(zip(coverage["GAL4 symbol"], coverage["Coverage status"]))
    assert statuses["nSyb.PU"] == "Resolved candidate substitution"
    assert statuses["sca.PU"] == "Resolved candidate substitution"
    assert statuses["elav.PU"] == "Resolved candidate substitution"
    assert statuses["missing.PU"] == "Unresolved"


def test_main_sheet_labels_unresolved_reagents_as_custom_lab_reagents():
    main_df = _build_main_sheet(
        [
            Gal4StockMatch(
                input_symbol="alrm.PD",
                fbst="",
                stock_number="",
                collection="",
                source_stock="",
            )
        ],
        phenotype_df=pd.DataFrame(),
    )

    assert main_df.loc[0, "Collection (output)"] == CUSTOM_LAB_REAGENT_COLLECTION


def test_collection_sheets_drop_placeholder_only_custom_lab_rows():
    main_df = _build_main_sheet(
        [
            Gal4StockMatch(
                input_symbol="alrm.PD",
                fbst="",
                stock_number="",
                collection="",
                source_stock="",
            )
        ],
        phenotype_df=pd.DataFrame(),
    )

    sheets = _build_collection_sheets(main_df)

    assert list(sheets) == [CUSTOM_LAB_REAGENT_COLLECTION]
    assert len(sheets[CUSTOM_LAB_REAGENT_COLLECTION]) == 0


def test_main_sheet_collapses_multiple_input_symbols_to_unique_stock_rows():
    main_df = _build_main_sheet(
        [
            Gal4StockMatch(
                input_symbol="elav-GAL4",
                fbst="FBst0008765",
                stock_number="8765",
                collection="BDSC",
                source_stock="BDSC (8765)",
                driver_genotype="P{GawB}elav[C155]",
            ),
            Gal4StockMatch(
                input_symbol="P{GawB}elav[C155]",
                fbst="FBst0008765",
                stock_number="8765",
                collection="BDSC",
                source_stock="BDSC (8765)",
                driver_genotype="P{GawB}elav[C155]",
            ),
        ],
        phenotype_df=pd.DataFrame(),
    )

    assert len(main_df) == 1
    assert main_df.loc[0, STOCK_ID_COLUMN] == "FBst0008765"
    assert main_df.loc[0, INPUT_SYMBOLS_COLUMN] == "elav-GAL4; P{GawB}elav[C155]"


def test_main_sheet_pairs_phenotypes_with_reference_ids():
    main_df = _build_main_sheet(
        [
            Gal4StockMatch(
                input_symbol="elav-GAL4",
                fbst="FBst0008765",
                stock_number="8765",
                collection="BDSC",
                source_stock="BDSC (8765)",
                driver_genotype="P{GawB}elav[C155]",
            ),
        ],
        phenotype_df=pd.DataFrame(
            [
                {
                    "Source/ Stock #": "BDSC (8765)",
                    "Phenotype": "sleep decreased",
                    "PMID": "12345",
                    "PMCID": "",
                },
                {
                    "Source/ Stock #": "BDSC (8765)",
                    "Phenotype": "activity increased",
                    "PMID": "",
                    "PMCID": "PMC67890",
                },
            ]
        ),
    )

    assert (
        main_df.loc[0, "Phenotypes"]
        == "{PMID12345, sleep decreased},\n{PMCID67890, activity increased}"
    )


def test_collection_sheets_are_grouped_by_collection_with_references_last():
    main_df = pd.DataFrame(
        [
            {STOCK_ID_COLUMN: "FBst1", "Collection (output)": "BDSC"},
            {STOCK_ID_COLUMN: "custom-driver", "Collection (output)": CUSTOM_LAB_REAGENT_COLLECTION},
            {STOCK_ID_COLUMN: "FBst2", "Collection (output)": "VDRC"},
        ]
    )

    sheets = _build_collection_sheets(main_df)

    assert list(sheets) == ["BDSC", "VDRC", CUSTOM_LAB_REAGENT_COLLECTION]


def test_references_sheet_collapses_rows_by_reference_and_lists_associated_gal4_reagents():
    references_df = _build_references_sheet(
        [
            Gal4StockMatch(
                input_symbol="elav-GAL4",
                fbst="FBst0008765",
                stock_number="8765",
                collection="BDSC",
                source_stock="BDSC (8765)",
                driver_genotype="P{GawB}elav[C155]",
            ),
            Gal4StockMatch(
                input_symbol="repo-GAL4",
                fbst="FBst0007415",
                stock_number="7415",
                collection="BDSC",
                source_stock="BDSC (7415)",
                driver_genotype="repo-GAL4",
            ),
        ],
        phenotype_df=pd.DataFrame(
            [
                {
                    "Source/ Stock #": "BDSC (8765)",
                    "Phenotype": "sleep decreased",
                    "Qualifier": "dominant",
                    "Genotype": "elav genotype",
                    "PMID": "12345",
                    "PMCID": "",
                    "Reference": "A GAL4 paper",
                    "Authors": "Doe et al.",
                    "Journal": "Fly",
                    "Year of Publication": "2025",
                },
                {
                    "Source/ Stock #": "BDSC (7415)",
                    "Phenotype": "activity increased",
                    "Qualifier": "",
                    "Genotype": "repo genotype",
                    "PMID": "12345",
                    "PMCID": "",
                    "Reference": "A GAL4 paper",
                    "Authors": "Doe et al.",
                    "Journal": "Fly",
                    "Year of Publication": "2025",
                },
            ]
        ),
    )

    assert len(references_df) == 1
    assert references_df.loc[0, "Reference ID"] == "PMID12345"
    assert references_df.loc[0, "Associated GAL4 symbols (input)"] == "elav-GAL4; repo-GAL4"
    assert references_df.loc[0, "Associated stock IDs"] == "FBst0008765; FBst0007415"
    assert (
        references_df.loc[0, "Phenotypes"]
        == "{PMID12345, sleep decreased},\n{PMID12345, activity increased}"
    )


def test_written_workbook_wraps_and_sizes_phenotype_evidence(tmp_path):
    output_path = tmp_path / "gal4_stocks.xlsx"
    main_df = pd.DataFrame(
        [
            {
                STOCK_ID_COLUMN: "FBst0008765",
                INPUT_SYMBOLS_COLUMN: "elav-GAL4",
                "Resolved driver genotype (output)": "P{GawB}elav[C155]",
                "Stock genotype (output)": "P{GawB}elav[C155]/CyO",
                "Collection (output)": "BDSC",
                "Stock # (output)": "8765",
                "Source/ Stock # (output)": "BDSC (8765)",
                "FBst (output)": "FBst0008765",
                "Phenotypes": (
                    "{PMID12345, sleep decreased under constant darkness},\n"
                    "{PMCID67890, activity increased during subjective night}"
                ),
                "Qualifiers": "dominant",
                "Reference IDs": "PMID12345; PMCID67890",
                "Reference titles": "A GAL4 paper with a deliberately long title",
                "Authors": "Doe et al.",
                "Journals": "Fly",
                "Years": "2025",
                "Phenotype record genotypes": "elav genotype with GAL4 driver",
            }
        ]
    )
    references_df = pd.DataFrame(
        [
            {
                "Reference ID": "PMID12345",
                "PMID": "PMID12345",
                "PMCID": "-",
                "Title": "A GAL4 paper with a deliberately long title",
                "Authors": "Doe et al.",
                "Journal": "Fly",
                "Year": "2025",
                "Associated GAL4 symbols (input)": "elav-GAL4",
                "Associated driver genotypes": "P{GawB}elav[C155]",
                "Associated source stocks": "BDSC (8765)",
                "Associated stock IDs": "FBst0008765",
                "Associated collections": "BDSC",
                "Phenotypes": "sleep decreased",
                "Qualifiers": "dominant",
                "Phenotype record genotypes": "elav genotype with GAL4 driver",
            }
        ]
    )

    write_gal4_workbook(output_path, main_df, references_df)

    wb = load_workbook(output_path)
    ws = wb["BDSC"]
    headers = {cell.value: cell.column for cell in ws[1]}
    phenotype_col = headers["Phenotypes"]

    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref == f"A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}"
    assert ws.cell(row=2, column=phenotype_col).alignment.wrap_text is True
    assert ws.column_dimensions[ws.cell(row=1, column=phenotype_col).column_letter].width >= 50
    assert ws.row_dimensions[2].height > 36
