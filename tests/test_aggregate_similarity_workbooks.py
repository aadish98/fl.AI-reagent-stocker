from __future__ import annotations

import contextlib
import importlib.util
import io
import sys
import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook


REPO_ROOT = Path(__file__).resolve().parents[1]
SCRIPT_PATH = REPO_ROOT / "scripts" / "aggregate_similarity_workbook_counts.py"


def _load_script_module():
    spec = importlib.util.spec_from_file_location(
        "aggregate_similarity_workbook_counts",
        SCRIPT_PATH,
    )
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


AGGREGATOR = _load_script_module()


def _write_workbook(path: Path, sheets: dict[str, list[list[object]]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name, rows in sheets.items():
        worksheet = workbook.create_sheet(sheet_name)
        for row in rows:
            worksheet.append(list(row))
    workbook.save(path)


def _build_cosine_contents(tier_counts: dict[str, object]) -> list[list[object]]:
    rows: list[list[object]] = [
        ["Tier Workbook Contents"],
        [],
        ["Bucket assignment", "Rows are assigned by Max Cosine Similarity."],
        ["Thresholds", "Fixed 0.05 cosine bins are evaluated from 0.95-1.0 downward to <0.05; empty buckets are skipped."],
        ["Ordering within tiers", "Rows are grouped by gene."],
        [],
        ["Sheet", "Meaning", "Rows"],
        ["Gene Set", "Copied from the input gene-list CSV data when available.", ""],
        ["Stock Phenotype Sheet", "Full phenotype table used to build all similarity tiers.", ""],
    ]
    for label, count in tier_counts.items():
        rows.append([label, f"Tier for similarity range {label}.", count])
    return rows


def _append_simple_section(
    rows: list[list[object]],
    section_label: str,
    bucket_headers: list[str],
    collection_rows: dict[str, list[int]],
) -> None:
    rows.append([section_label])
    rows.append(["Collection", *bucket_headers, "Total"])
    column_totals = [0] * len(bucket_headers)
    grand_total = 0
    for collection, counts in collection_rows.items():
        row_total = sum(counts)
        rows.append([collection, *counts, row_total])
        column_totals = [left + right for left, right in zip(column_totals, counts)]
        grand_total += row_total
    rows.append(["Total", *column_totals, grand_total])
    rows.append([])


def _build_simple_contents(
    bucket_headers: list[str],
    stock_rows: dict[str, list[int]],
    allele_rows: dict[str, list[int]],
    gene_rows: dict[str, list[int]],
) -> list[list[object]]:
    rows: list[list[object]] = [
        ["Simple Bucket Workbook Contents"],
        [],
        ["Bucket assignment", "Each reagent is assigned to exactly one collection / UAS / sleep-circ / balancer combination."],
        ["Counting invariant", "Genes, alleles, and reagents are never double-counted within or across combinations."],
        [],
    ]
    _append_simple_section(rows, "Stocks per bucket", bucket_headers, stock_rows)
    _append_simple_section(rows, "Alleles per bucket", bucket_headers, allele_rows)
    _append_simple_section(rows, "Genes per bucket", bucket_headers, gene_rows)
    rows.extend(
        [
            ["Sheet legend"],
            ["Sheet name", "Combination"],
            ["sheet_a", "demo combination A"],
            ["sheet_b", "demo combination B"],
        ]
    )
    return rows


def _parse_cosine_summary(path: Path) -> dict[str, int]:
    df = pd.read_excel(path, sheet_name=AGGREGATOR.COSINE_SUMMARY_SHEET, header=None)
    header_idx = next(
        idx
        for idx, row in df.iterrows()
        if row.fillna("").astype(str).tolist()[:2] == ["Similarity Range", "Total Rows"]
    )
    counts: dict[str, int] = {}
    for _, row in df.iloc[header_idx + 1 :].iterrows():
        label = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
        if not label or label == "Total":
            continue
        counts[label] = int(float(row.iloc[1]))
    return counts


def _parse_simple_summary(path: Path) -> dict[str, dict[str, dict[str, int]]]:
    df = pd.read_excel(path, sheet_name=AGGREGATOR.SIMPLE_SUMMARY_SHEET, header=None)
    parsed: dict[str, dict[str, dict[str, int]]] = {}
    for section_label in ("Stocks per bucket", "Alleles per bucket", "Genes per bucket"):
        section_idx = next(
            idx
            for idx, row in df.iterrows()
            if str(row.iloc[0]).strip() == section_label
        )
        header_row = df.iloc[section_idx + 1].fillna("").astype(str).tolist()
        total_idx = header_row.index("Total")
        bucket_headers = [value for value in header_row[1:total_idx] if value]
        section_rows: dict[str, dict[str, int]] = {}
        for _, row in df.iloc[section_idx + 2 :].iterrows():
            label = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
            if not label:
                continue
            if label == "Total":
                break
            section_rows[label] = {
                bucket: int(float(row.iloc[bucket_idx + 1]))
                for bucket_idx, bucket in enumerate(bucket_headers)
            }
        parsed[section_label] = section_rows
    return parsed


class TestAggregateSimilarityWorkbooks(unittest.TestCase):
    def test_main_aggregates_cosine_workbooks_detected_by_contents_not_name(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            root = Path(tmp_dir)
            _write_workbook(
                root / "alpha" / "completely_arbitrary.xlsx",
                {"Contents": _build_cosine_contents({"0.95-1.0": 3, "0.9-0.95": 2})},
            )
            _write_workbook(
                root / "beta" / "nested" / "another_random_name.xlsx",
                {"Contents": _build_cosine_contents({"0.95-1.0": 1, "<0.05": 4})},
            )
            _write_workbook(
                root / "Organized Stocks" / "ordinary_output.xlsx",
                {"Contents": [["Contents"], ["Sheet1"], ["References"]]},
            )

            exit_code = AGGREGATOR.main([str(root)])

            self.assertEqual(exit_code, 0)
            output_path = root / "similarity_workbook_summary.xlsx"
            self.assertTrue(output_path.exists(), output_path)

            workbook = load_workbook(output_path, data_only=True)
            try:
                self.assertEqual(workbook.sheetnames, [AGGREGATOR.COSINE_SUMMARY_SHEET])
                worksheet = workbook[AGGREGATOR.COSINE_SUMMARY_SHEET]
                self.assertEqual(worksheet["B2"].value, 2)
            finally:
                workbook.close()

            self.assertEqual(
                _parse_cosine_summary(output_path),
                {"0.95-1.0": 4, "0.9-0.95": 2, "<0.05": 4},
            )

    def test_main_aggregates_simple_bucket_workbooks_into_one_sheet(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            root = Path(tmp_dir)
            bucket_headers = ["UAS", "non-UAS | bal"]
            _write_workbook(
                root / "simple_a.xlsx",
                {
                    "Contents": _build_simple_contents(
                        bucket_headers=bucket_headers,
                        stock_rows={"BDSC": [1, 2], "VDRC": [0, 4]},
                        allele_rows={"BDSC": [1, 1], "VDRC": [0, 2]},
                        gene_rows={"BDSC": [1, 1], "VDRC": [0, 1]},
                    )
                },
            )
            _write_workbook(
                root / "nested" / "simple_b.xlsx",
                {
                    "Contents": _build_simple_contents(
                        bucket_headers=bucket_headers,
                        stock_rows={"BDSC": [2, 0], "NIG": [3, 1]},
                        allele_rows={"BDSC": [1, 0], "NIG": [2, 1]},
                        gene_rows={"BDSC": [1, 0], "NIG": [1, 1]},
                    )
                },
            )

            exit_code = AGGREGATOR.main([str(root)])

            self.assertEqual(exit_code, 0)
            output_path = root / "similarity_workbook_summary.xlsx"
            self.assertTrue(output_path.exists(), output_path)

            workbook = load_workbook(output_path, data_only=True)
            try:
                self.assertEqual(workbook.sheetnames, [AGGREGATOR.SIMPLE_SUMMARY_SHEET])
                worksheet = workbook[AGGREGATOR.SIMPLE_SUMMARY_SHEET]
                self.assertEqual(worksheet["B2"].value, 2)
            finally:
                workbook.close()

            parsed = _parse_simple_summary(output_path)
            self.assertEqual(
                parsed["Stocks per bucket"],
                {
                    "BDSC": {"UAS": 3, "non-UAS | bal": 2},
                    "VDRC": {"UAS": 0, "non-UAS | bal": 4},
                    "NIG": {"UAS": 3, "non-UAS | bal": 1},
                },
            )
            self.assertEqual(
                parsed["Alleles per bucket"]["NIG"],
                {"UAS": 2, "non-UAS | bal": 1},
            )
            self.assertEqual(
                parsed["Genes per bucket"]["BDSC"],
                {"UAS": 2, "non-UAS | bal": 1},
            )

    def test_main_writes_both_summary_sheets_when_both_formats_exist(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            root = Path(tmp_dir)
            _write_workbook(
                root / "cosine.xlsx",
                {"Contents": _build_cosine_contents({"0.95-1.0": 2})},
            )
            _write_workbook(
                root / "simple.xlsx",
                {
                    "Contents": _build_simple_contents(
                        bucket_headers=["UAS"],
                        stock_rows={"BDSC": [1]},
                        allele_rows={"BDSC": [1]},
                        gene_rows={"BDSC": [1]},
                    )
                },
            )

            exit_code = AGGREGATOR.main([str(root)])

            self.assertEqual(exit_code, 0)
            workbook = load_workbook(root / "similarity_workbook_summary.xlsx", data_only=True)
            try:
                self.assertEqual(
                    workbook.sheetnames,
                    [AGGREGATOR.COSINE_SUMMARY_SHEET, AGGREGATOR.SIMPLE_SUMMARY_SHEET],
                )
            finally:
                workbook.close()

    def test_main_skips_missing_or_malformed_workbooks_and_avoids_empty_output(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            root = Path(tmp_dir)
            _write_workbook(root / "missing_contents.xlsx", {"Other": [["not relevant"]]})
            _write_workbook(
                root / "bad_cosine.xlsx",
                {"Contents": _build_cosine_contents({"0.95-1.0": "oops"})},
            )

            stdout = io.StringIO()
            with contextlib.redirect_stdout(stdout):
                exit_code = AGGREGATOR.main([str(root)])

            self.assertEqual(exit_code, 0)
            self.assertFalse((root / "similarity_workbook_summary.xlsx").exists())
            text = stdout.getvalue()
            self.assertIn("Skipping", text)
            self.assertIn("No supported similarity workbooks found", text)


if __name__ == "__main__":
    unittest.main()
