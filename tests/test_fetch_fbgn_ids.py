from __future__ import annotations

import importlib.util
import sys
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[1]
SCRIPT_PATH = REPO_ROOT / "scripts" / "fetch_fbgn_ids.py"


def _load_script():
    spec = importlib.util.spec_from_file_location("fetch_fbgn_ids", SCRIPT_PATH)
    if spec is None or spec.loader is None:
        raise ImportError(f"Could not load {SCRIPT_PATH}")
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


fbgn = _load_script()


class TestConversionSidecars(unittest.TestCase):
    def test_is_conversion_sidecar(self):
        self.assertTrue(fbgn.is_conversion_sidecar(Path("needs-review.csv")))
        self.assertTrue(fbgn.is_conversion_sidecar(Path("validated_genes.csv")))
        self.assertFalse(fbgn.is_conversion_sidecar(Path("genes.csv")))
        self.assertFalse(fbgn.is_conversion_sidecar(Path("validated_genes.xlsx")))

    def test_select_source_csv_paths_skips_sidecars(self):
        import tempfile

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            (root / "genes.csv").write_text("ext_gene\nsky\n", encoding="utf-8")
            (root / "validated_genes.csv").write_text("ext_gene\nsky\n", encoding="utf-8")
            (root / "needs-review.csv").write_text("ext_gene\nEGFP\n", encoding="utf-8")
            found = [p.name for p in fbgn.select_source_csv_paths(root)]
            self.assertEqual(found, ["genes.csv"])

    def test_process_csv_leaves_original_and_writes_sidecars(self):
        import tempfile

        original_text = "ext_gene\nsky\nEGFP\n"
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "genes.csv"
            source.write_text(original_text, encoding="utf-8")
            original_bytes = source.read_bytes()

            review = fbgn.process_csv_file(
                source,
                "ext_gene",
                {"sky": "FBgn0032901"},
                {},
                {"FBgn0032901": "sky"},
            )
            self.assertIsNotNone(review)
            self.assertEqual(source.read_bytes(), original_bytes)
            self.assertEqual(source.read_text(encoding="utf-8"), original_text)

            validated = pd.read_csv(root / "validated_genes.csv")
            self.assertEqual(list(validated["ext_gene"]), ["sky"])
            self.assertEqual(list(validated["flybase_gene_id"]), ["FBgn0032901"])

            self.assertEqual(list(review["ext_gene"]), ["EGFP"])
            self.assertEqual(list(review["source_file"]), ["genes.csv"])

            xlsx_path = root / "validated_genes.xlsx"
            self.assertTrue(xlsx_path.exists())
            workbook = load_workbook(xlsx_path)
            self.assertEqual(workbook.sheetnames, ["Validated", "Needs review"])
            validated_sheet = workbook["Validated"]
            self.assertEqual(validated_sheet["D2"].value, "FBgn0032901")
            self.assertEqual(
                str(validated_sheet["D2"].hyperlink.target),
                "https://flybase.org/reports/FBgn0032901",
            )
            needs_sheet = workbook["Needs review"]
            self.assertEqual(
                [cell.value for cell in needs_sheet[1]],
                ["ext_gene", "corrected_ext_gene", "primary_symbol", "flybase_gene_id"],
            )
            self.assertEqual(needs_sheet["A2"].value, "EGFP")
            self.assertIsNone(needs_sheet["D2"].hyperlink)

    def test_main_writes_combined_needs_review_and_ignores_sidecars_on_rerun(self):
        import tempfile

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            (root / "genes.csv").write_text("ext_gene\nsky\nEGFP\n", encoding="utf-8")
            fake_flybase = root / "flybase"
            fake_flybase.mkdir()

            def _fake_load(_flybase_data_dir):
                return {"sky": "FBgn0032901"}, {}, {"FBgn0032901": "sky"}

            original_load = fbgn.load_mappings
            fbgn.load_mappings = _fake_load
            try:
                first = fbgn.main([str(root), "ext_gene", "--flybase-data-dir", str(fake_flybase)])
                self.assertEqual(first, 0)
                needs = pd.read_csv(root / "needs-review.csv")
                self.assertEqual(list(needs["ext_gene"]), ["EGFP"])
                self.assertEqual(list(needs["source_file"]), ["genes.csv"])

                source_bytes = (root / "genes.csv").read_bytes()
                second = fbgn.main([str(root), "ext_gene", "--flybase-data-dir", str(fake_flybase)])
                self.assertEqual(second, 0)
                self.assertEqual((root / "genes.csv").read_bytes(), source_bytes)
                self.assertFalse((root / "validated_validated_genes.csv").exists())
                self.assertTrue((root / "validated_genes.csv").exists())
            finally:
                fbgn.load_mappings = original_load


if __name__ == "__main__":
    unittest.main()
