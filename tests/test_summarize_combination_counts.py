from __future__ import annotations

import importlib.util
import json
import sys
import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook


REPO_ROOT = Path(__file__).resolve().parents[1]
SCRIPT_PATH = REPO_ROOT / "scripts" / "summarize_combination_counts.py"


def _load_script_module():
    spec = importlib.util.spec_from_file_location(
        "summarize_combination_counts",
        SCRIPT_PATH,
    )
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


SUMMARY = _load_script_module()


def _write_organized_workbook(path: Path, rows: list[list[object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Contents"
    for row in rows:
        worksheet.append(row)
    workbook.save(path)


class TestSummarizeCombinationCounts(unittest.TestCase):
    def test_parses_contents_breakdown_and_derives_gene_set_name(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            workbook_path = (
                root
                / "GeneSetA"
                / "Stocks"
                / "Organized Stocks"
                / "GeneSetA_aggregated.xlsx"
            )
            _write_organized_workbook(
                workbook_path,
                [
                    ["What this workbook is"],
                    [],
                    ["Prioritized sheet breakdown"],
                    ["Sheet criteria", "# Stocks", "# Alleles", "# Genes", "Sheet Name"],
                    ["Bloomington >> UAS", 3, 2, 1, "Sheet1"],
                    ["Vienna >> UAS", 0, 0, 0, "-"],
                    ["Total", 3, 2, 1, ""],
                ],
            )

            counts = SUMMARY.parse_workbook_counts(workbook_path)

            self.assertEqual([count.combination for count in counts], ["Bloomington >> UAS", "Vienna >> UAS"])
            self.assertEqual(counts[0].gene_set, "GeneSetA")
            self.assertEqual(counts[0].num_stocks, 3)
            self.assertEqual(counts[0].num_alleles, 2)
            self.assertEqual(counts[0].num_genes, 1)
            # No "Per Gene Set Runs" ancestor -> relative_path defaults to ".".
            self.assertEqual(counts[0].relative_path, ".")

    def test_main_writes_summary_in_config_order_with_zero_fill(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            workbook_path = (
                root
                / "GeneSetA"
                / "Stocks"
                / "Organized Stocks"
                / "GeneSetA_aggregated.xlsx"
            )
            _write_organized_workbook(
                workbook_path,
                [
                    ["Prioritized sheet breakdown"],
                    ["Sheet criteria", "# Stocks", "# Alleles", "# Genes", "Sheet Name"],
                    ["Vienna >> UAS", 4, 3, 2, "Sheet1"],
                    ["Total", 4, 3, 2, ""],
                ],
            )
            config_path = root / "config.json"
            config_path.write_text(
                json.dumps(
                    {
                        "combinations": [
                            ["Bloomington", "UAS"],
                            ["Vienna", "UAS"],
                        ]
                    }
                ),
                encoding="utf-8",
            )

            result = SUMMARY.main([str(root), "--config", str(config_path)])

            self.assertEqual(result, 0)
            csv_path = root / "combination_counts_summary.csv"
            xlsx_path = root / "combination_counts_summary.xlsx"
            self.assertTrue(csv_path.exists())
            self.assertTrue(xlsx_path.exists())

            df = pd.read_csv(csv_path)
            self.assertEqual(df["combination"].tolist(), ["Bloomington >> UAS", "Vienna >> UAS"])
            self.assertEqual(df["num_stocks"].tolist(), [0, 4])
            self.assertEqual(df["num_alleles"].tolist(), [0, 3])
            self.assertEqual(df["num_genes"].tolist(), [0, 2])

            workbook = load_workbook(xlsx_path, read_only=True)
            self.assertIn("Combination Counts", workbook.sheetnames)
            self.assertIn("Totals by Combination", workbook.sheetnames)

    def test_derives_relative_path_from_relative_run_key(self):
        runs_layout = (
            "Per Gene Set Runs",
            "CSW FC0.5",
            "FC0.5_Sleep_freq2",
            "Stocks",
            "Organized Stocks",
            "FC0.5_Sleep_freq2_aggregated.xlsx",
        )
        path = Path("/root").joinpath(*runs_layout)
        self.assertEqual(SUMMARY._derive_relative_path(path), "CSW FC0.5")

        # Nested folders are joined with "/" (e.g. input Genes/A/B/C -> "A/B/C").
        nested = Path("/Genes").joinpath(
            "Per Gene Set Runs",
            "A",
            "B",
            "C",
            "gs",
            "Stocks",
            "Organized Stocks",
            "gs_aggregated.xlsx",
        )
        self.assertEqual(SUMMARY._derive_relative_path(nested), "A/B/C")

        # Top-level gene set (no subfolder) -> ".".
        flat = Path("/root").joinpath(
            "Per Gene Set Runs",
            "gs",
            "Stocks",
            "Organized Stocks",
            "gs_aggregated.xlsx",
        )
        self.assertEqual(SUMMARY._derive_relative_path(flat), ".")

    def test_main_combined_summary_groups_by_relative_path_with_borders(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            runs_root = root / "Per Gene Set Runs"
            # Two input files share folder "CSW FC0.5"; one is in "CSW FC1".
            for source, gene_set, stocks in (
                ("CSW FC0.5", "SetA", 3),
                ("CSW FC0.5", "SetB", 4),
                ("CSW FC1", "SetC", 5),
            ):
                _write_organized_workbook(
                    runs_root
                    / source
                    / gene_set
                    / "Stocks"
                    / "Organized Stocks"
                    / f"{gene_set}_aggregated.xlsx",
                    [
                        ["Prioritized sheet breakdown"],
                        ["Sheet criteria", "# Stocks", "# Alleles", "# Genes", "Sheet Name"],
                        ["Bloomington >> UAS", stocks, 2, 1, "Sheet1"],
                        ["Vienna >> UAS", stocks, 1, 1, "Sheet2"],
                        ["Total", stocks, 3, 2, ""],
                    ],
                )

            result = SUMMARY.main([str(root)])
            self.assertEqual(result, 0)

            df = pd.read_csv(root / "combination_counts_summary.csv")
            self.assertEqual(list(df.columns)[0], "relative_path")
            self.assertEqual(
                df["relative_path"].tolist(),
                ["CSW FC0.5", "CSW FC0.5", "CSW FC0.5", "CSW FC0.5", "CSW FC1", "CSW FC1"],
            )
            self.assertEqual(
                df["gene_set"].tolist(),
                ["SetA", "SetA", "SetB", "SetB", "SetC", "SetC"],
            )

            workbook = load_workbook(root / "combination_counts_summary.xlsx")
            sheet = workbook["Combination Counts"]
            # Rows (1=header). Data rows 2-7. Borders mark group starts:
            #   row 4 -> new gene_set (SetB, same folder)  => thin top border
            #   row 6 -> new relative_path (CSW FC1)        => medium top border
            self.assertEqual(sheet.cell(row=4, column=1).border.top.style, "thin")
            self.assertEqual(sheet.cell(row=6, column=1).border.top.style, "medium")
            # Rows that continue a block have no top border.
            self.assertIsNone(sheet.cell(row=3, column=1).border.top.style)
            self.assertIsNone(sheet.cell(row=5, column=1).border.top.style)

            totals = workbook["Totals by Combination"]
            rows = [list(r) for r in totals.iter_rows(values_only=True)]
            self.assertEqual(rows[0][0], "relative_path")
            self.assertEqual({r[0] for r in rows[1:]}, {"CSW FC0.5", "CSW FC1"})

    def test_derives_relative_path_for_clean_stocks_layout(self):
        # Cleaned layout: workbook directly under Stocks (no Organized Stocks).
        clean = Path("/Genes").joinpath(
            "Per Gene Set Runs",
            "CSW FC0.5",
            "gs",
            "Stocks",
            "gs_aggregated.xlsx",
        )
        self.assertEqual(SUMMARY._derive_relative_path(clean), "CSW FC0.5")

        nested_clean = Path("/Genes").joinpath(
            "Per Gene Set Runs",
            "A",
            "B",
            "C",
            "gs",
            "Stocks",
            "gs_aggregated.xlsx",
        )
        self.assertEqual(SUMMARY._derive_relative_path(nested_clean), "A/B/C")

    def test_summarize_function_on_clean_layout_tree(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            runs_root = root / "Per Gene Set Runs"
            # Clean layout: aggregated workbook directly under Stocks.
            _write_organized_workbook(
                runs_root / "CSW FC0.5" / "SetA" / "Stocks" / "SetA_aggregated.xlsx",
                [
                    ["Prioritized sheet breakdown"],
                    ["Sheet criteria", "# Stocks", "# Alleles", "# Genes", "Sheet Name"],
                    ["Bloomington >> UAS", 3, 2, 1, "Sheet1"],
                    ["Total", 3, 2, 1, ""],
                ],
            )

            self.assertEqual(SUMMARY.summarize_combination_counts(root), 0)
            df = pd.read_csv(root / "combination_counts_summary.csv")
            self.assertEqual(df["relative_path"].tolist(), ["CSW FC0.5"])
            self.assertEqual(df["gene_set"].tolist(), ["SetA"])
            self.assertEqual(df["num_stocks"].tolist(), [3])

    def test_files_within_folder_sorted_naturally(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            runs_root = root / "Per Gene Set Runs"
            # Created out of order, including a multi-digit frequency.
            for gene_set in (
                "FC0.5_Wake_frequency_10_n=3genes",
                "FC0.5_Wake_frequency_2_n=9genes",
                "FC0.5_Wake_frequency_4_n=7genes",
                "FC0.5_Wake_frequency_5_n=6genes",
            ):
                _write_organized_workbook(
                    runs_root
                    / "CSW FC0.5"
                    / gene_set
                    / "Stocks"
                    / "Organized Stocks"
                    / f"{gene_set}_aggregated.xlsx",
                    [
                        ["Prioritized sheet breakdown"],
                        ["Sheet criteria", "# Stocks", "# Alleles", "# Genes", "Sheet Name"],
                        ["Bloomington >> UAS", 1, 1, 1, "Sheet1"],
                        ["Total", 1, 1, 1, ""],
                    ],
                )

            self.assertEqual(SUMMARY.main([str(root)]), 0)
            df = pd.read_csv(root / "combination_counts_summary.csv")
            # frequency_2 < 4 < 5 < 10 (natural, not lexical) order.
            self.assertEqual(
                df["gene_set"].tolist(),
                [
                    "FC0.5_Wake_frequency_2_n=9genes",
                    "FC0.5_Wake_frequency_4_n=7genes",
                    "FC0.5_Wake_frequency_5_n=6genes",
                    "FC0.5_Wake_frequency_10_n=3genes",
                ],
            )


if __name__ == "__main__":
    unittest.main()
