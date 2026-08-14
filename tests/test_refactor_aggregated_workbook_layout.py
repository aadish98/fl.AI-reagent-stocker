from __future__ import annotations

import importlib.util
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook


REPO_ROOT = Path(__file__).resolve().parents[1]
SCRIPT_PATH = REPO_ROOT / "scripts" / "refactor_aggregated_workbook_layout.py"


def _load_script_module():
    spec = importlib.util.spec_from_file_location(
        "refactor_aggregated_workbook_layout",
        SCRIPT_PATH,
    )
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


REFACTOR = _load_script_module()


class TestRefactorAggregatedWorkbookLayout(unittest.TestCase):
    def test_configured_combination_sheet_is_formatted_as_data_sheet(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "named.xlsx"
            workbook = Workbook()
            contents = workbook.active
            contents.title = "Contents"
            contents.append(["Prioritized sheet breakdown"])
            contents.append(
                ["Sheet criteria", "# Stocks", "# Alleles", "# Genes", "Sheet Name"]
            )
            contents.append(["Bloomington >> RNAi", 1, 1, 1, "1_BDSC·RNAi·Pheno"])
            contents.append(["Total", 1, 1, 1, ""])
            data_sheet = workbook.create_sheet("1_BDSC·RNAi·Pheno")
            data_sheet.append(["stock_number", "genotype"])
            data_sheet.append(["100", "w[*]; P{UAS-test}"])
            workbook.save(path)

            REFACTOR.refactor_workbook_layout(path)

            updated = load_workbook(path)
            self.assertEqual(updated["1_BDSC·RNAi·Pheno"].freeze_panes, "A2")
            self.assertIsNotNone(updated["1_BDSC·RNAi·Pheno"].auto_filter.ref)
            updated.close()


if __name__ == "__main__":
    unittest.main()
