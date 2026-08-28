from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from scripts.build_temperature_phenotype_gene_list import (
    REAGENT_COLUMNS,
    attach_reagents,
    build_genes_sheet,
    build_included_assertions,
    build_references_sheet,
    deduplicate_reagent_rows,
    exclude_genes,
    is_uas_or_rnai_reagent,
    sort_reagent_rows,
    temperature_keyword_terms,
    write_temperature_workbook,
)


def test_temperature_keywords_use_word_boundaries_and_qualifiers():
    assert temperature_keyword_terms("gonadal sheath", "") == []
    assert temperature_keyword_terms("lethal", "heat sensitive|adult stage") == ["heat"]
    assert temperature_keyword_terms("abnormal thermotaxis", "") == ["thermo"]
    assert temperature_keyword_terms("abnormal cold stress response", "") == ["cold"]


def test_is_uas_or_rnai_reagent_detects_transgenic_constructs():
    assert is_uas_or_rnai_reagent("UAS-Trp[A1]") is True
    assert is_uas_or_rnai_reagent("Dcr-2[RNAi]") is True
    assert is_uas_or_rnai_reagent("P{KK108671}VIE-260B") is True
    assert is_uas_or_rnai_reagent("nAcRalpha-30D[1]") is False
    assert is_uas_or_rnai_reagent("") is False


def test_inclusion_requires_keyword_hit_and_allele_or_insertion():
    phenotypes = pd.DataFrame(
        [
            {
                "genotype_symbols": "a[1]",
                "genotype_FBids": "FBal0000001",
                "phenotype_name": "lethal",
                "qualifier_names": "heat sensitive",
                "reference": "FBrf0000001",
            },
            {
                "genotype_symbols": "b[1]",
                "genotype_FBids": "FBal0000002",
                "phenotype_name": "abnormal sensory behavior",
                "qualifier_names": "",
                "reference": "FBrf0000002",
            },
            {
                "genotype_symbols": "c[1]",
                "genotype_FBids": "FBal0000003",
                "phenotype_name": "gonadal sheath",
                "qualifier_names": "",
                "reference": "FBrf0000003",
            },
            {
                "genotype_symbols": "UAS-foreign[RNAi]",
                "genotype_FBids": "FBal0000004",
                "phenotype_name": "abnormal heat stress response",
                "qualifier_names": "",
                "reference": "FBrf0000004",
            },
        ]
    )
    mapping = pd.DataFrame(
        [
            {
                "AlleleID": "FBal0000001",
                "AlleleSymbol": "gene1[1]",
                "GeneID": "FBgn0000001",
                "GeneSymbol": "gene1",
            },
            {
                "AlleleID": "FBal0000002",
                "AlleleSymbol": "gene2[1]",
                "GeneID": "FBgn0000002",
                "GeneSymbol": "gene2",
            },
            {
                "AlleleID": "FBal0000003",
                "AlleleSymbol": "gene3[1]",
                "GeneID": "FBgn0000003",
                "GeneSymbol": "gene3",
            },
            {
                "AlleleID": "FBal0000004",
                "AlleleSymbol": "UAS-gene4[RNAi]",
                "GeneID": "FBgn0000004",
                "GeneSymbol": "gene4",
            },
        ]
    )
    dmel_catalog = pd.DataFrame(
        [
            {
                "flybase_gene_id": f"FBgn000000{index}",
                "current_gene_symbol": f"gene{index}",
            }
            for index in range(1, 5)
        ]
    )

    included = build_included_assertions(
        phenotypes,
        mapping,
        dmel_gene_catalog=dmel_catalog,
    )

    # "abnormal sensory behavior" and "gonadal sheath" have no keyword hit;
    # the UAS/RNAi row is excluded even though it has a keyword hit.
    assert included["Allele ID"].tolist() == ["FBal0000001"]
    assert included["keyword_hit"].tolist() == [True]
    assert included["Gene"].tolist() == ["gene1"]


def test_stock_join_custom_fallback_and_reagent_deduplication():
    assertions = pd.DataFrame(
        [
            {
                "Gene": "geneA",
                "flybase_gene_id": "FBgn0000001",
                "Allele ID": "FBal0000001",
                "Allele symbol": "geneA[1]",
                "phenotype_name": "abnormal heat stress response",
                "qualifier_names": "",
                "Reference (FBrf)": "FBrf0000001",
                "genotype_symbols": "geneA[1]",
                "keyword_hit": True,
                "keyword_match_terms": "heat",
            },
            {
                "Gene": "geneB",
                "flybase_gene_id": "FBgn0000002",
                "Allele ID": "FBal0000002",
                "Allele symbol": "geneB[1]",
                "phenotype_name": "abnormal thermotaxis",
                "qualifier_names": "",
                "Reference (FBrf)": "FBrf0000002",
                "genotype_symbols": "geneB[1]",
                "keyword_hit": True,
                "keyword_match_terms": "thermo",
            },
        ]
    )
    derived = pd.DataFrame(
        [
            {
                "FBst": "FBst0000100",
                "stock_number": "100",
                "collection": "Bloomington",
                "FB_genotype": "geneA[1]/CyO",
                "derived_stock_component": "FBal0000001",
            },
            {
                "FBst": "FBst0000100",
                "stock_number": "100",
                "collection": "Bloomington",
                "FB_genotype": "geneA[1]/CyO",
                "derived_stock_component": "FBal0000001",
            },
        ]
    )

    attached = attach_reagents(assertions, derived)
    result = deduplicate_reagent_rows(attached)

    assert len(result) == 2
    orderable = result[result["FBst"] == "FBst0000100"].iloc[0]
    custom = result[result["custom_stock"]].iloc[0]
    assert orderable["Stock #"] == "100"
    assert custom["Stock #"] == "geneB[1]"
    assert custom["Collection"] == "Custom phenotype reagent"
    assert custom["reagent_key"] == "custom:FBal0000002"


def test_exclude_genes_removes_w_before_stock_expansion():
    assertions = pd.DataFrame(
        [
            {"Gene": "w", "flybase_gene_id": "FBgn0003996"},
            {"Gene": "TrpA1", "flybase_gene_id": "FBgn0034166"},
        ]
    )

    filtered = exclude_genes(assertions)

    assert filtered["Gene"].tolist() == ["TrpA1"]


def _reference_ready_rows():
    return pd.DataFrame(
        [
            {
                "Gene": "geneMany",
                "flybase_gene_id": "FBgn1",
                "Collection": "Vienna Drosophila Resource Center",
                "Stock #": "v12345",
                "FBst": "FBst1",
                "custom_stock": False,
                "Genotype": "g1",
                "Allele symbol": "a1",
                "Allele ID": "FBal1",
                "Phenotype": "Disrupted Circadian Rhythms",
                "Qualifier": "",
                "Reference (FBrf)": "FBrf1",
                "PMID": "101",
                "Reference Title": "Paper One",
                "Reference Authors": "A Author; B Author",
                "Date of publication": "2020",
                "Reference Journal": "Fly",
                "keyword_hit": True,
                "keyword_match_terms": "thermo",
                "reagent_key": "FBst1",
            },
            {
                "Gene": "geneMany",
                "flybase_gene_id": "FBgn1",
                "Collection": "Vienna Drosophila Resource Center",
                "Stock #": "v12345",
                "FBst": "FBst1",
                "custom_stock": False,
                "Genotype": "g1",
                "Allele symbol": "a1",
                "Allele ID": "FBal1",
                "Phenotype": "abnormal thermotaxis",
                "Qualifier": "",
                "Reference (FBrf)": "FBrf1",
                "PMID": "101",
                "Reference Title": "Paper One",
                "Reference Authors": "A Author; B Author",
                "Date of publication": "2020",
                "Reference Journal": "Fly",
                "keyword_hit": True,
                "keyword_match_terms": "thermo",
                "reagent_key": "FBst1",
            },
            {
                "Gene": "geneMany",
                "flybase_gene_id": "FBgn1",
                "Collection": "Bloomington",
                "Stock #": "200",
                "FBst": "FBst2",
                "custom_stock": False,
                "Genotype": "g2",
                "Allele symbol": "a2",
                "Allele ID": "FBal2",
                "Phenotype": "abnormal cold stress response",
                "Qualifier": "",
                "Reference (FBrf)": "FBrf2",
                "PMID": "102",
                "Reference Title": "Paper Two",
                "Reference Authors": "C Author",
                "Date of publication": "2021",
                "Reference Journal": "Genetics",
                "keyword_hit": True,
                "keyword_match_terms": "cold",
                "reagent_key": "FBst2",
            },
            {
                "Gene": "geneFew",
                "flybase_gene_id": "FBgn2",
                "Collection": "Custom phenotype reagent",
                "Stock #": "geneFew[1]",
                "FBst": "",
                "custom_stock": True,
                "Genotype": "g3",
                "Allele symbol": "a3",
                "Allele ID": "FBal3",
                "Phenotype": "abnormal temperature response",
                "Qualifier": "",
                "Reference (FBrf)": "FBrf3",
                "PMID": "103",
                "Reference Title": "Paper Three",
                "Reference Authors": "D Author",
                "Date of publication": "2022",
                "Reference Journal": "Nature",
                "keyword_hit": True,
                "keyword_match_terms": "temperature",
                "reagent_key": "custom:FBal3",
            },
        ]
    )


def test_gene_sort_and_reference_first_appearance_uniqueness():
    reagents = sort_reagent_rows(_reference_ready_rows())
    assert reagents["Gene"].tolist()[:3] == ["geneMany", "geneMany", "geneMany"]
    assert reagents["n_reagents_for_gene"].tolist()[:3] == [2, 2, 2]

    references = build_references_sheet(reagents)

    assert len(references) == 3
    first = references.iloc[0]
    assert first["PMID"] == "101"  # Alphabetically-first phenotype encountered first.
    vdrc = references[references["PMID"] == "101"].iloc[0]
    assert vdrc["Phenotype"] == (
        "Disrupted Circadian Rhythms; abnormal thermotaxis"
    )
    assert vdrc["Associated reagent"] == (
        "v12345, Disrupted Circadian Rhythms; abnormal thermotaxis, Vienna"
    )
    genes = build_genes_sheet(reagents)
    gene_many = genes[genes["Gene"] == "geneMany"].iloc[0]
    assert "Disrupted Circadian Rhythms (PMID: 101)" in gene_many["Phenotype (PMID)"]
    assert "abnormal cold stress response (PMID: 102)" in gene_many["Phenotype (PMID)"]


def test_workbook_places_references_second_and_uses_arial(tmp_path: Path):
    reagents = sort_reagent_rows(_reference_ready_rows())
    references = build_references_sheet(reagents)
    genes = build_genes_sheet(reagents)
    summary = pd.DataFrame([{"Item": "Inclusion rule", "Value": "keyword hit AND allele/insertion"}])
    output = tmp_path / "temperature.xlsx"

    write_temperature_workbook(output, reagents, references, genes, summary)

    workbook = load_workbook(output, read_only=False)
    assert workbook.sheetnames == ["Reagents", "References", "Genes", "Summary"]
    assert workbook["Reagents"]["A1"].font.name == "Arial"
    assert workbook["References"]["A2"].hyperlink.target.endswith("/101/")
    assert list(pd.read_excel(output, sheet_name="Reagents").columns) == REAGENT_COLUMNS
