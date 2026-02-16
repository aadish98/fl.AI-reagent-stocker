"""
Pipeline 1: Gene → BDSC Stocks → References + Functional Validation

This module provides the AlleleReferencesPipeline class that orchestrates:
1. Loading input genes from CSV files
2. Mapping genes to BDSC stocks
3. Getting references for alleles
4. Running functional validation with OpenAI
5. Saving results to Excel

Usage:
    from fly_stocker_v2 import AlleleReferencesPipeline
    from fly_stocker_v2.config import Settings
    
    settings = Settings()
    pipeline = AlleleReferencesPipeline(settings)
    pipeline.run(input_dir, keywords=["sleep", "circadian"], run_functional_validation=False)
"""

import gc
import subprocess
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Set

import pandas as pd

try:
    from .config import Settings, ValidationStatus, GET_FBGN_IDS_SCRIPT
    from .utils import (
        clean_id,
        parse_semicolon_list,
        find_latest_tsv,
        load_flybase_tsv,
        generate_keyword_column_name,
        get_gpt_derived_columns,
        apply_experimental_prefix,
        EXPERIMENTAL_PREFIX,
    )
    from .external.pubmed import PubMedClient, PubMedCache
    from .external.fulltext import FullTextFetcher, FunctionalValidator
    from .validation_runner import run_functional_validation
except ImportError:
    # Support flat-repo execution (for example `python -m cli ...`).
    from config import Settings, ValidationStatus, GET_FBGN_IDS_SCRIPT
    from utils import (
        clean_id,
        parse_semicolon_list,
        find_latest_tsv,
        load_flybase_tsv,
        generate_keyword_column_name,
        get_gpt_derived_columns,
        apply_experimental_prefix,
        EXPERIMENTAL_PREFIX,
    )
    from external.pubmed import PubMedClient, PubMedCache
    from external.fulltext import FullTextFetcher, FunctionalValidator
    from validation_runner import run_functional_validation


def _apply_grey_fill_openpyxl(ws, df: pd.DataFrame) -> None:
    """
    Apply light grey background fill to GPT-derived column **headers** in an
    openpyxl worksheet.  Only the header row (row 1) is shaded.

    GPT-derived columns are detected by the '[EXPERIMENTAL] ' prefix that
    :func:`apply_experimental_prefix` adds to their header names.
    """
    from openpyxl.styles import PatternFill

    grey_fill = PatternFill(
        start_color='D9D9D9', end_color='D9D9D9', fill_type='solid'
    )
    # openpyxl columns are 1-indexed; row 1 is the header
    for col_idx, col_name in enumerate(df.columns, 1):
        if str(col_name).startswith(EXPERIMENTAL_PREFIX):
            ws.cell(row=1, column=col_idx).fill = grey_fill


class AlleleReferencesPipeline:
    """
    Pipeline for mapping genes to stocks and references with functional validation.
    
    This is Pipeline 1 in the fly-stocker workflow:
    Input: CSV files with gene IDs
    Output: Excel file with stocks, references, and validation results
    """
    
    def __init__(self, settings: Optional[Settings] = None):
        """
        Initialize the pipeline.
        
        Args:
            settings: Configuration settings (uses defaults if None)
        """
        self.settings = settings or Settings()
        
        # Validate settings
        issues = self.settings.validate()
        for issue in issues:
            print(f"    Warning: {issue}")
        
        # Initialize components
        self._pubmed_cache = PubMedCache(self.settings.pubmed_cache_path)
        self._pubmed_client = PubMedClient(
            cache=self._pubmed_cache,
            api_key=self.settings.ncbi_api_key
        )
        self._fulltext_fetcher = FullTextFetcher(
            unpaywall_token=self.settings.unpaywall_token,
            ncbi_api_key=self.settings.ncbi_api_key,
            method_cache_path=self.settings.fulltext_method_cache_path
        )
        self._validator = FunctionalValidator(
            openai_api_key=self.settings.openai_api_key,
            model=self.settings.openai_model,
            log_dir=self.settings.gpt_log_dir / datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            if self.settings.enable_gpt_logging else None
        )
        
        # Persist details of filtered-in references that still have no PMID.
        self._no_pmid_fbrf_records: List[Dict[str, str]] = []
        
        # Data tables (loaded lazily)
        self._tables: Optional[dict] = None
    
    def _load_tables(self) -> dict:
        """Load FlyBase TSV files and BDSC data."""
        if self._tables is not None:
            return self._tables
        
        print("\nLoading data tables...")
        tables = {}
        
        # Load BDSC stockgenes
        print(f"  Loading: {self.settings.bdsc_stockgenes_path.name}")
        tables["bdsc_stockgenes"] = pd.read_csv(
            self.settings.bdsc_stockgenes_path,
            dtype={
                'stknum': str, 'genotype': str, 'component_symbol': str,
                'gene_symbol': str, 'fbgn': str, 'bdsc_symbol_id': str, 'bdsc_gene_id': str
            }
        )
        tables["bdsc_stockgenes"]['stknum'] = tables["bdsc_stockgenes"]['stknum'].apply(clean_id)
        print(f"    Loaded {len(tables['bdsc_stockgenes'])} stock-gene entries")
        
        # Load entity_publication
        entity_pub_path = find_latest_tsv(self.settings.flybase_references_path, "entity_publication")
        print(f"  Loading: {entity_pub_path.name}")
        tables["entity_publication"] = load_flybase_tsv(entity_pub_path)
        print(f"    Loaded {len(tables['entity_publication'])} entity-publication entries")
        
        # Load reference metadata
        fbrf_path = find_latest_tsv(self.settings.flybase_references_path, "fbrf_pmid_pmcid_doi")
        print(f"  Loading: {fbrf_path.name}")
        tables["ref_metadata"] = load_flybase_tsv(fbrf_path)
        
        # Load bloomington.csv for chromosome info (Ch # all)
        if self.settings.bdsc_bloomington_path and self.settings.bdsc_bloomington_path.exists():
            print(f"  Loading: {self.settings.bdsc_bloomington_path.name}")
            # Try different encodings for BDSC files
            for encoding in ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']:
                try:
                    tables["bdsc_bloomington"] = pd.read_csv(
                        self.settings.bdsc_bloomington_path,
                        dtype={'Stk #': str},
                        encoding=encoding
                    )
                    tables["bdsc_bloomington"]['Stk #'] = tables["bdsc_bloomington"]['Stk #'].apply(clean_id)
                    print(f"    Loaded {len(tables['bdsc_bloomington'])} stock entries from bloomington.csv (encoding: {encoding})")
                    break
                except UnicodeDecodeError:
                    continue
            else:
                print(f"  Warning: Could not read bloomington.csv with any encoding, skipping chromosome info")
                tables["bdsc_bloomington"] = None
        else:
            print(f"  Warning: bloomington.csv not found, skipping chromosome info")
            tables["bdsc_bloomington"] = None
        
        # Load stockcomps_map_comments.csv for map statement info
        if self.settings.bdsc_stockcomps_path and self.settings.bdsc_stockcomps_path.exists():
            print(f"  Loading: {self.settings.bdsc_stockcomps_path.name}")
            # Try different encodings for BDSC files which may have special characters
            for encoding in ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']:
                try:
                    tables["bdsc_stockcomps"] = pd.read_csv(
                        self.settings.bdsc_stockcomps_path,
                        dtype={'Stk #': str},
                        encoding=encoding
                    )
                    tables["bdsc_stockcomps"]['Stk #'] = tables["bdsc_stockcomps"]['Stk #'].apply(clean_id)
                    print(f"    Loaded {len(tables['bdsc_stockcomps'])} component entries from stockcomps_map_comments.csv (encoding: {encoding})")
                    break
                except UnicodeDecodeError:
                    continue
            else:
                print(f"  Warning: Could not read stockcomps_map_comments.csv with any encoding, skipping map statement info")
                tables["bdsc_stockcomps"] = None
        else:
            print(f"  Warning: stockcomps_map_comments.csv not found, skipping map statement info")
            tables["bdsc_stockcomps"] = None
        
        # Load FlyBase stocks TSV for flybase_genotype
        try:
            stocks_tsv_path = find_latest_tsv(self.settings.flybase_alleles_stocks_path, "stocks")
            print(f"  Loading: {stocks_tsv_path.name}")
            fb_stocks = load_flybase_tsv(stocks_tsv_path)
            # Filter to Bloomington stocks only
            if 'collection_short_name' in fb_stocks.columns:
                fb_stocks = fb_stocks[
                    fb_stocks['collection_short_name'].str.contains('Bloomington', case=False, na=False)
                ].copy()
            if 'stock_number' in fb_stocks.columns:
                fb_stocks['stock_number'] = fb_stocks['stock_number'].apply(clean_id)
            tables["fb_stocks"] = fb_stocks
            print(f"    Loaded {len(fb_stocks)} Bloomington stock entries from FlyBase stocks TSV")
        except FileNotFoundError:
            print(f"  Warning: FlyBase stocks TSV not found, skipping flybase_genotype")
            tables["fb_stocks"] = None
        
        self._tables = tables
        return tables
    
    def _run_fbgnid_conversion(self, input_dir: Path, gene_col: str = "ext_gene") -> bool:
        """Run GetFBgnIDs.py to convert gene names to FlyBase IDs."""
        print(f"\n{'='*60}")
        print("Step 0: Running FBgnID conversion on input files...")
        print(f"{'='*60}")
        
        if not GET_FBGN_IDS_SCRIPT.exists():
            print(f"ERROR: GetFBgnIDs.py not found at {GET_FBGN_IDS_SCRIPT}")
            return False
        
        cmd = [sys.executable, str(GET_FBGN_IDS_SCRIPT), str(input_dir), gene_col]
        print(f"  Running: python {GET_FBGN_IDS_SCRIPT.name} {input_dir} {gene_col}")
        
        try:
            result = subprocess.run(cmd, capture_output=True, text=True, check=False)
            print(result.stdout)
            if result.stderr:
                print(result.stderr)
        except Exception as e:
            print(f"ERROR: Failed to run GetFBgnIDs.py: {e}")
            return False
        
        return True
    
    def _validate_fbgnids(self, csv_path: Path, gene_col: str = "flybase_gene_id") -> List[str]:
        """Validate that all rows have valid FlyBase gene IDs."""
        df = pd.read_csv(csv_path)
        
        if gene_col not in df.columns:
            return []
        
        invalid_mask = (
            df[gene_col].isna() |
            (df[gene_col].astype(str).str.strip() == "") |
            (df[gene_col].astype(str).str.strip() == "-") |
            ~df[gene_col].astype(str).str.startswith("FBgn")
        )
        
        if invalid_mask.any():
            gene_symbol_col = None
            for col in ["ext_gene", "gene_symbol", "gene", "Gene"]:
                if col in df.columns:
                    gene_symbol_col = col
                    break
            
            if gene_symbol_col:
                return df.loc[invalid_mask, gene_symbol_col].dropna().unique().tolist()
            else:
                return df.loc[invalid_mask, gene_col].astype(str).unique().tolist()
        
        return []
    
    def _aggregate_input_genes(self, input_dir: Path, gene_col: str) -> Tuple[List[str], List[Path]]:
        """Read all CSV files and aggregate unique genes."""
        csv_files = sorted(input_dir.glob("*.csv"))
        
        if not csv_files:
            return [], []
        
        all_genes = set()
        valid_files = []
        
        print(f"\nAggregating genes from {len(csv_files)} CSV file(s)...")
        
        for csv_path in csv_files:
            try:
                gene_series_chunks = []
                for chunk in pd.read_csv(csv_path, usecols=[gene_col], chunksize=10000):
                    gene_series_chunks.append(chunk[gene_col].dropna())
                
                if gene_series_chunks:
                    genes = pd.concat(gene_series_chunks, ignore_index=True).unique()
                    genes = [g for g in genes if isinstance(g, str) and g.startswith("FBgn")]
                    all_genes.update(genes)
                    valid_files.append(csv_path)
                    print(f"  {csv_path.name}: {len(genes)} unique genes")
            except (KeyError, ValueError) as e:
                print(f"  {csv_path.name}: Skipped (no '{gene_col}' column or read error)")
                continue
        
        unique_genes = sorted(all_genes)
        print(f"\nTotal: {len(unique_genes)} unique genes from {len(valid_files)} file(s)")
        
        return unique_genes, valid_files
    
    def _build_stock_mapping(
        self,
        input_genes: List[str],
        bdsc_df: pd.DataFrame
    ) -> pd.DataFrame:
        """
        Build mapping from input genes to BDSC stocks.
        
        Logic:
        1. Match input FBgn gene IDs directly against the BDSC stockgenes 'fbgn'
           column to find all gene-relevant stock components.
        2. Use the BDSC component_symbol values as-is for downstream reference
           lookups (these exist in entity_publication under the same names).
        3. Aggregate stock data to include relevant component symbols, gene symbols,
           and all component/gene symbols.
        4. Return one row per stock.
        """
        if not input_genes:
            return pd.DataFrame()
        
        input_gene_set = set(input_genes)
        
        # 1. Find all BDSC stock-gene rows where fbgn matches an input gene.
        #    Each row in stockgenes.csv is a curated pairing of component_symbol
        #    to a specific fbgn, so this filter directly gives us the gene-relevant
        #    components for each stock.
        matches = bdsc_df[bdsc_df['fbgn'].isin(input_gene_set)].copy()
        
        if len(matches) == 0:
            print(f"    Warning: No BDSC stocks found for {len(input_gene_set)} input genes")
            return pd.DataFrame()
        
        matched_genes = matches['fbgn'].nunique()
        matching_stknums = matches['stknum'].unique()
        print(f"    Found {len(matching_stknums)} stocks for {matched_genes}/{len(input_gene_set)} input genes")
        
        # Get full content for these stocks (all components, not just the gene-relevant ones)
        all_rows = bdsc_df[bdsc_df['stknum'].isin(matching_stknums)].copy()
        
        # 2. Aggregate data
        
        # Aggregate relevant components (from FBgn-matched rows only)
        relevant_aggs = matches.groupby('stknum').agg({
            'component_symbol': lambda x: '; '.join(sorted(set(x.dropna()))),
            'gene_symbol': lambda x: '; '.join(sorted(set(x.dropna()))),
            'fbgn': lambda x: '; '.join(sorted(set(x.dropna())))
        }).reset_index()
        relevant_aggs = relevant_aggs.rename(columns={
            'component_symbol': 'relevant_component_symbols',
            'gene_symbol': 'relevant_gene_symbols',
            'fbgn': 'relevant_flybase_gene_ids'
        })
        
        # Aggregate all components (from all_rows — full stock content)
        full_aggs = all_rows.groupby('stknum').agg({
            'genotype': 'first',
            'component_symbol': lambda x: '; '.join(sorted(set(x.dropna()))),
            'gene_symbol': lambda x: '; '.join(sorted(set(x.dropna()))),
            'fbgn': lambda x: '; '.join(sorted(set(x.dropna())))
        }).reset_index()
        full_aggs = full_aggs.rename(columns={
            'component_symbol': 'all_component_symbols',
            'gene_symbol': 'all_gene_symbols',
            'fbgn': 'all_flybase_gene_ids'
        })
        
        # Split all_gene_symbols into genes (FBgn) vs constructs (non-FBgn)
        # by pairing each gene_symbol with its fbgn from the raw stock rows
        def _split_genes_constructs(stknum):
            rows = all_rows[all_rows['stknum'] == stknum]
            genes = set()
            constructs = set()
            for _, r in rows.iterrows():
                sym = r.get('gene_symbol')
                fbgn_val = str(r.get('fbgn', ''))
                if pd.isna(sym) or not str(sym).strip():
                    continue
                sym = str(sym).strip()
                if fbgn_val.startswith('FBgn'):
                    genes.add(sym)
                else:
                    constructs.add(sym)
            return pd.Series({
                'all_stock_genes': '; '.join(sorted(genes)),
                'all_stock_constructs': '; '.join(sorted(constructs))
            })
        
        gene_construct_split = full_aggs['stknum'].apply(_split_genes_constructs)
        full_aggs = pd.concat([full_aggs, gene_construct_split], axis=1)
        
        # Merge aggregates
        result = pd.merge(full_aggs, relevant_aggs, on='stknum', how='left')
        
        # Rename stock number column
        result = result.rename(columns={'stknum': 'stock_number'})
        
        # Use relevant_flybase_gene_ids as primary 'flybase_gene_id' if single,
        # otherwise keep as aggregated list.
        # Pipeline 2 expects 'flybase_gene_id' column.
        result['flybase_gene_id'] = result['relevant_flybase_gene_ids']
        
        print(f"    Built stock mapping: {len(result)} stocks")
        
        return result
    
    def _join_bdsc_metadata(
        self,
        stocks_df: pd.DataFrame,
        bloomington_df: Optional[pd.DataFrame],
        stockcomps_df: Optional[pd.DataFrame],
        fb_stocks_df: Optional[pd.DataFrame] = None
    ) -> pd.DataFrame:
        """
        Join additional BDSC metadata to stocks DataFrame.
        
        Adds:
        - 'Ch # all': Chromosome info from bloomington.csv
        - 'Stock comments': Stock comments from bloomington.csv
        - 'mapstatement': Map statement from stockcomps_map_comments.csv (aggregated per stock)
        - 'comment1': Comment from stockcomps_map_comments.csv (aggregated per stock)
        - 'flybase_genotype': Genotype from FlyBase stocks TSV (Bloomington only)
        
        Args:
            stocks_df: DataFrame with stock data (must have 'stock_number' column)
            bloomington_df: DataFrame from bloomington.csv (optional)
            stockcomps_df: DataFrame from stockcomps_map_comments.csv (optional)
            fb_stocks_df: DataFrame from FlyBase stocks TSV, pre-filtered to Bloomington (optional)
        
        Returns:
            stocks_df with additional columns
        """
        stocks_df = stocks_df.copy()
        
        # Join Ch # all and Stock comments from bloomington.csv
        if bloomington_df is not None and len(bloomington_df) > 0:
            bloom_cols = ['Stk #', 'Ch # all']
            if 'Stock comments' in bloomington_df.columns:
                bloom_cols.append('Stock comments')
            bloom_subset = bloomington_df[bloom_cols].drop_duplicates(subset=['Stk #'])
            bloom_subset = bloom_subset.rename(columns={'Stk #': 'stock_number'})
            
            stocks_df = pd.merge(
                stocks_df,
                bloom_subset,
                on='stock_number',
                how='left'
            )
            print(f"    Joined 'Ch # all' for {stocks_df['Ch # all'].notna().sum()} stocks")
            if 'Stock comments' in stocks_df.columns:
                print(f"    Joined 'Stock comments' for {stocks_df['Stock comments'].notna().sum()} stocks")
            else:
                stocks_df['Stock comments'] = ''
        else:
            stocks_df['Ch # all'] = ''
            stocks_df['Stock comments'] = ''
        
        # Join mapstatement and comment1 from stockcomps_map_comments.csv
        if stockcomps_df is not None and len(stockcomps_df) > 0:
            agg_dict = {}
            
            # Aggregate mapstatements per stock
            if 'mapstatement' in stockcomps_df.columns:
                agg_dict['mapstatement'] = lambda x: '; '.join(
                    sorted(set(str(m) for m in x.dropna() if str(m).strip()))
                )
            
            # Aggregate comment1 per stock
            if 'comment1' in stockcomps_df.columns:
                agg_dict['comment1'] = lambda x: '; '.join(
                    sorted(set(str(c) for c in x.dropna() if str(c).strip()))
                )
            
            if agg_dict:
                stockcomps_agg = stockcomps_df.groupby('Stk #').agg(agg_dict).reset_index()
                stockcomps_agg = stockcomps_agg.rename(columns={'Stk #': 'stock_number'})
                
                stocks_df = pd.merge(
                    stocks_df,
                    stockcomps_agg,
                    on='stock_number',
                    how='left'
                )
                if 'mapstatement' in agg_dict:
                    print(f"    Joined 'mapstatement' for {stocks_df['mapstatement'].notna().sum()} stocks")
                if 'comment1' in agg_dict:
                    print(f"    Joined 'comment1' for {stocks_df['comment1'].notna().sum()} stocks")
            
            if 'mapstatement' not in stocks_df.columns:
                stocks_df['mapstatement'] = ''
            if 'comment1' not in stocks_df.columns:
                stocks_df['comment1'] = ''
        else:
            stocks_df['mapstatement'] = ''
            stocks_df['comment1'] = ''
        
        # Join flybase_genotype from FlyBase stocks TSV (pre-filtered to Bloomington)
        if fb_stocks_df is not None and len(fb_stocks_df) > 0 and 'FB_genotype' in fb_stocks_df.columns:
            fb_geno = fb_stocks_df[['stock_number', 'FB_genotype']].drop_duplicates(subset=['stock_number'])
            fb_geno = fb_geno.rename(columns={'FB_genotype': 'flybase_genotype'})
            
            stocks_df = pd.merge(
                stocks_df,
                fb_geno,
                on='stock_number',
                how='left'
            )
            print(f"    Joined 'flybase_genotype' for {stocks_df['flybase_genotype'].notna().sum()} stocks")
        else:
            stocks_df['flybase_genotype'] = ''
        
        return stocks_df
    
    def _get_references_for_stocks(
        self,
        stocks_df: pd.DataFrame,
        entity_pub: pd.DataFrame,
        ref_metadata: pd.DataFrame,
        keywords: List[str] = None,
        input_genes: List[str] = None
    ) -> Tuple[pd.DataFrame, pd.DataFrame, Dict[str, Set[str]], Dict[str, dict]]:
        """
        Get references for stocks from entity_publication.
        
        Returns:
            stocks_df: Updated with PMID, PMCID, ref_count, keywords check, keyword_ref_count, total_refs
            references_df: DataFrame of references found with enhanced columns
            component_to_pmids: Dict mapping component_symbol -> set of PMIDs
            pmid_metadata: Dict mapping PMID -> PubMed metadata (title, abstract, etc.)
        """
        # Extract all unique relevant component symbols
        relevant_symbols = set()
        for symbols_str in stocks_df['relevant_component_symbols'].dropna():
            relevant_symbols.update(parse_semicolon_list(symbols_str))
        
        # Use FlyBase TSV metadata only for FBrf -> PMID/PMCID/pub_type mapping.
        effective_ref_metadata = ref_metadata.copy()
        for col in ['FBrf', 'PMID', 'PMCID', 'DOI', 'pub_type', 'miniref']:
            if col not in effective_ref_metadata.columns:
                effective_ref_metadata[col] = ''
            effective_ref_metadata[col] = effective_ref_metadata[col].fillna('').astype(str).str.strip()
        effective_ref_metadata['FBrf'] = effective_ref_metadata['FBrf'].apply(clean_id)
        effective_ref_metadata['PMID'] = effective_ref_metadata['PMID'].apply(clean_id)
        effective_ref_metadata['PMCID'] = effective_ref_metadata['PMCID'].apply(clean_id)
        effective_ref_metadata['DOI'] = effective_ref_metadata['DOI'].apply(clean_id)
        effective_ref_metadata['pub_type'] = effective_ref_metadata['pub_type'].str.lower()
        
        # Pre-filter entity_publication: only keep references that are papers
        # with ≤50 associated genes. This excludes large-scale studies (e.g.,
        # genome-wide screens) that mention many genes but are unlikely to
        # contain stock-specific functional data.
        max_genes_per_pub = 50
        
        # 1. Identify paper-type FBrfs from reference metadata.
        # Unknown publication types are treated as non-paper and excluded.
        if 'pub_type' in effective_ref_metadata.columns:
            paper_fbrfs = set(
                effective_ref_metadata.loc[
                    effective_ref_metadata['pub_type'] == 'paper',
                    'FBrf'
                ].dropna()
            )
        else:
            paper_fbrfs = set(entity_pub['FlyBase_publication_id'].dropna().unique())
        
        # 2. Count unique *genes* per publication, keep those with ≤50.
        #    entity_publication contains all entity types (genes FBgn, alleles
        #    FBal, insertions FBti, constructs FBtp, aberrations FBab, etc.).
        #    Counting all entity types inflates the number for papers that
        #    describe many allelic variants of a few genes (common for newer
        #    TI/CRISPR/split-GAL4 resources), causing legitimate focused papers
        #    to be incorrectly excluded.  Filtering to FBgn-prefixed entity_id
        #    entries counts only the actual genes a paper references.
        gene_entity_pub = entity_pub[entity_pub['entity_id'].str.startswith('FBgn', na=False)]
        gene_counts = gene_entity_pub.groupby('FlyBase_publication_id')['entity_name'].nunique()
        low_gene_pubs = set(gene_counts[gene_counts <= max_genes_per_pub].index)
        # Publications with zero gene entities (only alleles/constructs) should
        # not be excluded by this filter, so include them as well.
        pubs_with_genes = set(gene_counts.index)
        pubs_without_genes = set(entity_pub['FlyBase_publication_id'].dropna().unique()) - pubs_with_genes
        low_gene_pubs = low_gene_pubs | pubs_without_genes
        excluded_pubs = len(pubs_with_genes) - len(gene_counts[gene_counts <= max_genes_per_pub])
        
        # 3. Combine: must be a paper AND have ≤50 associated genes
        valid_pubs = paper_fbrfs & low_gene_pubs
        print(f"    Pre-filtered references: {len(valid_pubs)} papers with ≤{max_genes_per_pub} associated genes (excluded {excluded_pubs} large-scale publications)")
        
        # Apply pre-filter + component symbol match
        entity_pub_filtered = entity_pub[
            entity_pub['FlyBase_publication_id'].isin(valid_pubs) &
            entity_pub['entity_name'].isin(relevant_symbols)
        ].copy()
        
        if len(entity_pub_filtered) == 0:
            stocks_df['PMID'] = ""
            stocks_df['PMCID'] = ""
            stocks_df['FBrf without PMID'] = ""
            stocks_df['references_without_pmid_count'] = 0
            stocks_df['ref_count'] = 0
            stocks_df['keyword_ref_count'] = 0
            stocks_df['keyword_ref_pmids'] = ''
            stocks_df['total_refs'] = 0
            self._no_pmid_fbrf_records = []
            if keywords:
                kw_col_name = generate_keyword_column_name(keywords)
                stocks_df[kw_col_name] = False
            return stocks_df, pd.DataFrame(), {}, {}
        
        # Get unique FBrfs
        fbrfs = [clean_id(x) for x in entity_pub_filtered['FlyBase_publication_id'].dropna().unique()]
        
        # Build refs_df from effective metadata and keep rows even when PMID is empty.
        refs_df = effective_ref_metadata[effective_ref_metadata['FBrf'].isin(fbrfs)].copy()
        
        # Keep paper references only. Unknown publication types are discarded.
        if 'pub_type' in refs_df.columns:
            refs_df = refs_df[refs_df['pub_type'] == 'paper'].copy()
        
        # Build FBrf -> PMID/PMCID maps
        fbrf_to_pmid = dict(zip(refs_df['FBrf'], refs_df['PMID'].apply(clean_id)))
        fbrf_to_pmcid = dict(zip(
            refs_df['FBrf'],
            refs_df['PMCID'].apply(clean_id) if 'PMCID' in refs_df.columns else [''] * len(refs_df)
        ))
        
        # Build component -> FBrfs and component -> PMIDs maps
        comp_fbrfs = entity_pub_filtered.groupby('entity_name')['FlyBase_publication_id'].apply(set).to_dict()
        
        component_to_pmids: Dict[str, Set[str]] = {}
        component_to_fbrfs_without_pmid: Dict[str, Set[str]] = {}
        for comp, fbrf_set in comp_fbrfs.items():
            pmids = set()
            no_pmid_fbrfs = set()
            for fbrf_raw in fbrf_set:
                fbrf = clean_id(fbrf_raw)
                pmid = fbrf_to_pmid.get(fbrf)
                if pmid:
                    pmids.add(pmid)
                else:
                    no_pmid_fbrfs.add(fbrf)
            if pmids:
                component_to_pmids[comp] = pmids
            if no_pmid_fbrfs:
                component_to_fbrfs_without_pmid[comp] = no_pmid_fbrfs
        
        # Build PMID -> stocks mapping for reference sheet
        pmid_to_stocks: Dict[str, Set[str]] = {}
        pmid_to_genes: Dict[str, Set[str]] = {}
        
        for _, row in stocks_df.iterrows():
            stock_num = clean_id(row['stock_number'])
            comps = parse_semicolon_list(row.get('relevant_component_symbols', ''))
            genes = parse_semicolon_list(row.get('relevant_gene_symbols', ''))
            
            for comp in comps:
                if comp in component_to_pmids:
                    for pmid in component_to_pmids[comp]:
                        if pmid not in pmid_to_stocks:
                            pmid_to_stocks[pmid] = set()
                        pmid_to_stocks[pmid].add(stock_num)
                        
                        if pmid not in pmid_to_genes:
                            pmid_to_genes[pmid] = set()
                        pmid_to_genes[pmid].update(genes)
        
        # Add aggregated PMID/PMCID to stocks
        def get_stock_pmids(row):
            comps = parse_semicolon_list(row.get('relevant_component_symbols', ''))
            stock_pmids = set()
            for comp in comps:
                if comp in component_to_pmids:
                    stock_pmids.update(component_to_pmids[comp])
            return '; '.join(sorted(stock_pmids))
        
        def get_stock_pmcids(row):
            comps = parse_semicolon_list(row.get('relevant_component_symbols', ''))
            stock_pmcids = set()
            for comp in comps:
                if comp in comp_fbrfs:
                    for fbrf in comp_fbrfs[comp]:
                        pmcid = fbrf_to_pmcid.get(clean_id(fbrf))
                        if pmcid:
                            stock_pmcids.add(pmcid)
            return '; '.join(sorted(stock_pmcids))
        
        def get_stock_no_pmid_fbrfs(row):
            comps = parse_semicolon_list(row.get('relevant_component_symbols', ''))
            fbrfs_without_pmid = set()
            for comp in comps:
                if comp in component_to_fbrfs_without_pmid:
                    fbrfs_without_pmid.update(component_to_fbrfs_without_pmid[comp])
            return '; '.join(sorted(fbrfs_without_pmid))

        stocks_df['PMID'] = stocks_df.apply(get_stock_pmids, axis=1)
        stocks_df['PMCID'] = stocks_df.apply(get_stock_pmcids, axis=1)
        stocks_df['FBrf without PMID'] = stocks_df.apply(get_stock_no_pmid_fbrfs, axis=1)
        stocks_df['references_without_pmid_count'] = stocks_df['FBrf without PMID'].apply(
            lambda x: len([f for f in str(x).split(';') if f.strip()]) if pd.notna(x) and x else 0
        )
        
        # Total refs count now includes PMIDs + surviving no-PMID FBrfs.
        stocks_df['total_refs'] = stocks_df.apply(
            lambda row: (
                len([p for p in str(row.get('PMID', '')).split(';') if p.strip()]) +
                len([f for f in str(row.get('FBrf without PMID', '')).split(';') if f.strip()])
            ),
            axis=1
        )
        stocks_df['ref_count'] = stocks_df['total_refs']  # Alias for compatibility
        
        # Persist surviving filtered references that have no PMID for downstream reporting.
        no_pmid_refs_df = refs_df[refs_df['PMID'].astype(str).str.strip() == ''].copy()
        no_pmid_records: List[Dict[str, str]] = []
        if len(no_pmid_refs_df) > 0:
            fbrf_to_components: Dict[str, Set[str]] = {}
            for comp, fset in comp_fbrfs.items():
                for f in fset:
                    f_clean = clean_id(f)
                    if f_clean not in fbrf_to_components:
                        fbrf_to_components[f_clean] = set()
                    fbrf_to_components[f_clean].add(comp)
            fbrf_to_stocks: Dict[str, Set[str]] = {}
            for _, row in stocks_df.iterrows():
                stock_num = clean_id(row.get('stock_number', ''))
                for f in parse_semicolon_list(row.get('FBrf without PMID', '')):
                    f_clean = clean_id(f)
                    if not f_clean:
                        continue
                    if f_clean not in fbrf_to_stocks:
                        fbrf_to_stocks[f_clean] = set()
                    fbrf_to_stocks[f_clean].add(stock_num)
            
            for _, row in no_pmid_refs_df.iterrows():
                fbrf = clean_id(row.get('FBrf', ''))
                if not fbrf:
                    continue
                no_pmid_records.append({
                    'FBrf': fbrf,
                    'pub_type': str(row.get('pub_type', '')).strip(),
                    'miniref': str(row.get('miniref', '')).strip(),
                    'associated_stocks': ', '.join(sorted(fbrf_to_stocks.get(fbrf, set()))),
                    'associated_components': ', '.join(sorted(fbrf_to_components.get(fbrf, set()))),
                })
        self._no_pmid_fbrf_records = no_pmid_records
        
        # Get all unique PMIDs referenced
        unique_pmids = set()
        for pmids_set in component_to_pmids.values():
            unique_pmids.update(pmids_set)
        unique_pmids_list = sorted(unique_pmids)
        
        # Fetch PubMed metadata (title, abstract)
        pubmed_metadata = self._pubmed_client.fetch_metadata(unique_pmids_list)
        
        # Check keywords in title/abstract
        kw_col_name = None
        keyword_pmids: Set[str] = set()
        
        if keywords:
            kw_col_name = generate_keyword_column_name(keywords)
            
            # Determine which PMIDs have keyword hits
            for pmid in unique_pmids:
                if self._pubmed_client.check_keywords(pmid, keywords, pubmed_metadata):
                    keyword_pmids.add(pmid)
            
            def check_keyword_refs(pmids_str):
                if pd.isna(pmids_str) or not str(pmids_str).strip():
                    return False
                pmids = [p.strip() for p in str(pmids_str).split(';') if p.strip()]
                return any(pmid in keyword_pmids for pmid in pmids)
            
            def count_keyword_refs(pmids_str):
                if pd.isna(pmids_str) or not str(pmids_str).strip():
                    return 0
                pmids = [p.strip() for p in str(pmids_str).split(';') if p.strip()]
                return sum(1 for pmid in pmids if pmid in keyword_pmids)
            
            def collect_keyword_pmids(pmids_str):
                if pd.isna(pmids_str) or not str(pmids_str).strip():
                    return ''
                pmids = [p.strip() for p in str(pmids_str).split(';') if p.strip()]
                matching = [pmid for pmid in pmids if pmid in keyword_pmids]
                return '; '.join(sorted(matching))
            
            stocks_df[kw_col_name] = stocks_df['PMID'].apply(check_keyword_refs)
            stocks_df['keyword_ref_count'] = stocks_df['PMID'].apply(count_keyword_refs)
            stocks_df['keyword_ref_pmids'] = stocks_df['PMID'].apply(collect_keyword_pmids)
            
            # Add to refs_df
            refs_df[kw_col_name] = refs_df['PMID'].apply(
                lambda x: clean_id(x) in keyword_pmids if pd.notna(x) else False
            )
        else:
            stocks_df['keyword_ref_count'] = 0
            stocks_df['keyword_ref_pmids'] = ''
        
        # Enhance refs_df with new columns
        if len(refs_df) > 0:
            # Add associated_stocks column
            refs_df['associated_stocks'] = refs_df['PMID'].apply(
                lambda x: ', '.join(sorted(pmid_to_stocks.get(clean_id(x), set()))) if pd.notna(x) else ''
            )
            
            # Add associated_genes column (only genes from input list)
            input_gene_symbols = set()
            if input_genes:
                # Get gene symbols for input FBgn IDs from stocks_df
                for _, row in stocks_df.iterrows():
                    input_gene_symbols.update(parse_semicolon_list(row.get('relevant_gene_symbols', '')))
            
            def get_associated_genes(pmid):
                if pd.isna(pmid):
                    return ''
                pmid_clean = clean_id(pmid)
                genes = pmid_to_genes.get(pmid_clean, set())
                # Filter to only input genes
                if input_gene_symbols:
                    genes = genes.intersection(input_gene_symbols)
                return ', '.join(sorted(genes))
            
            refs_df['associated_genes'] = refs_df['PMID'].apply(get_associated_genes)
            
            # Add stock_count column
            refs_df['stock_count'] = refs_df['PMID'].apply(
                lambda x: len(pmid_to_stocks.get(clean_id(x), set())) if pd.notna(x) else 0
            )
            
            # Add publication date, authors, journal from unified PubMed metadata
            def _get_pubmed_meta(pmid):
                if pd.isna(pmid):
                    return {}
                pmid_clean = clean_id(pmid)
                meta = pubmed_metadata.get(pmid_clean, {})
                return meta if isinstance(meta, dict) else {}

            def get_pub_date(pmid):
                meta = _get_pubmed_meta(pmid)
                return str(meta.get('year', '') or '')
            
            def get_authors(pmid):
                meta = _get_pubmed_meta(pmid)
                authors = meta.get('authors', [])
                if isinstance(authors, str):
                    return '; '.join([a.strip() for a in authors.split(';') if a.strip()])
                if isinstance(authors, list):
                    return '; '.join([str(a).strip() for a in authors if str(a).strip()])
                return ''
            
            refs_df['publication_date'] = refs_df['PMID'].apply(get_pub_date)
            refs_df['author(s)'] = refs_df['PMID'].apply(get_authors)
            
            # Get journal from metadata
            def get_journal(pmid):
                meta = _get_pubmed_meta(pmid)
                return str(meta.get('journal', '') or '')
            
            refs_df['journal'] = refs_df['PMID'].apply(get_journal)
            
            # Add title from pubmed_metadata
            refs_df['title'] = refs_df['PMID'].apply(
                lambda x: pubmed_metadata.get(clean_id(x), {}).get('title', '') if pd.notna(x) else ''
            )
        
        return stocks_df, refs_df, component_to_pmids, pubmed_metadata
    
    def _run_functional_validation(
        self,
        stocks_df: pd.DataFrame,
        references_df: pd.DataFrame,
        component_to_pmids: Dict[str, Set[str]],
        keywords: List[str] = None,
        soft_run: bool = False
    ) -> pd.DataFrame:
        """Run functional validation on keyword-hit references."""
        return run_functional_validation(
            stocks_df=stocks_df,
            references_df=references_df,
            keywords=keywords,
            soft_run=soft_run,
            validator=self._validator,
            fulltext_fetcher=self._fulltext_fetcher,
            pubmed_client=self._pubmed_client,
            pubmed_cache=self._pubmed_cache,
            max_gpt_calls_per_stock=self.settings.max_gpt_calls_per_stock,
            component_to_pmids=component_to_pmids,
        )
    
    def run(
        self,
        input_dir: Path,
        keywords: Optional[List[str]] = None,
        gene_col: str = "flybase_gene_id",
        input_gene_col: str = "ext_gene",
        skip_fbgnid_conversion: bool = False,
        run_functional_validation: bool = False,
    ) -> Optional[Path]:
        """
        Run the full pipeline.
        
        Args:
            input_dir: Directory containing CSV files with gene IDs
            keywords: List of keywords for filtering references
            gene_col: Column name for FlyBase gene IDs (after conversion)
            input_gene_col: Column name for gene symbols (before conversion)
            skip_fbgnid_conversion: Skip FBgnID conversion step
            run_functional_validation: Run GPT-based functional validation now
        
        Returns:
            Path to output Excel file, or None if failed
        """
        input_dir = Path(input_dir)
        if not input_dir.is_dir():
            raise ValueError(f"Input path is not a directory: {input_dir}")
        
        if keywords:
            print(f"Keywords: {', '.join(keywords)}")
        
        # Step 0: FBgnID conversion
        if not skip_fbgnid_conversion and not self.settings.skip_fbgnid_conversion:
            self._run_fbgnid_conversion(input_dir, input_gene_col)
            
            # Validate
            print("\nValidating FBgnID conversion...")
            csv_files = sorted(input_dir.glob("*.csv"))
            all_invalid = []
            for csv_path in csv_files:
                invalid = self._validate_fbgnids(csv_path, gene_col)
                if invalid:
                    all_invalid.extend(invalid)
                    print(f"  {csv_path.name}: {len(invalid)} genes without valid FBgnID")
            
            if all_invalid:
                print(f"\nERROR: {len(set(all_invalid))} genes without FBgnIDs")
                return None
            else:
                print("  All genes successfully mapped")
        
        # Create output directory
        output_dir = input_dir / "Stocks"
        output_dir.mkdir(exist_ok=True)
        
        # Aggregate genes
        unique_genes, valid_files = self._aggregate_input_genes(input_dir, gene_col)
        
        if not unique_genes:
            print("No genes found in input files")
            return None
        
        # Load tables
        tables = self._load_tables()
        
        # Build stock mapping
        print(f"\n{'='*60}")
        print("Building stock mapping...")
        stock_mapping = self._build_stock_mapping(
            unique_genes,
            tables["bdsc_stockgenes"]
        )
        
        if len(stock_mapping) == 0:
            print("No stocks found for input genes")
            return None
        
        # Get references
        print(f"\nGetting references...")
        stock_mapping, references_df, component_to_pmids, pubmed_metadata = self._get_references_for_stocks(
            stock_mapping,
            tables["entity_publication"],
            tables["ref_metadata"],
            keywords,
            input_genes=unique_genes
        )
        
        # Join BDSC metadata (Ch # all, Stock comments, mapstatement, comment1, flybase_genotype)
        print(f"\nJoining BDSC metadata...")
        stock_mapping = self._join_bdsc_metadata(
            stock_mapping,
            tables.get("bdsc_bloomington"),
            tables.get("bdsc_stockcomps"),
            tables.get("fb_stocks")
        )
        
        # Add allele symbol for downstream pipelines (before boolean checks so it's available)
        stock_mapping['AlleleSymbol'] = stock_mapping['relevant_component_symbols']
        stock_mapping['collection'] = 'BDSC'
        
        # Compute UAS boolean: case-insensitive substring check across
        # AlleleSymbol, all_stock_constructs, Stock comments, genotype, flybase_genotype, comment1
        def _check_uas(row):
            for col in ['AlleleSymbol', 'all_stock_constructs', 'Stock comments', 'genotype', 'flybase_genotype', 'comment1']:
                val = str(row.get(col, ''))
                if 'uas' in val.lower():
                    return True
            return False
        
        stock_mapping['UAS'] = stock_mapping.apply(_check_uas, axis=1)
        uas_count = stock_mapping['UAS'].sum()
        print(f"    UAS flag: {uas_count} stocks marked as UAS")
        
        # Compute sgRNA boolean: case-insensitive substring check across
        # AlleleSymbol, all_stock_constructs, Stock comments, genotype, flybase_genotype, comment1
        def _check_sgrna(row):
            for col in ['AlleleSymbol', 'all_stock_constructs', 'Stock comments', 'genotype', 'flybase_genotype', 'comment1']:
                val = str(row.get(col, ''))
                if 'sgrna' in val.lower():
                    return True
            return False
        
        stock_mapping['sgRNA'] = stock_mapping.apply(_check_sgrna, axis=1)
        sgrna_count = stock_mapping['sgRNA'].sum()
        print(f"    sgRNA flag: {sgrna_count} stocks marked as sgRNA")
        
        # Rename keyword_ref_count and keyword_ref_pmids to include the actual keywords
        if keywords:
            kw_ref_count_col = f"{' OR '.join(keywords)} references count"
            kw_ref_pmids_col = f"Stock (allele) {' OR '.join(keywords)} references"
            rename_map = {}
            if 'keyword_ref_count' in stock_mapping.columns:
                rename_map['keyword_ref_count'] = kw_ref_count_col
            if 'keyword_ref_pmids' in stock_mapping.columns:
                rename_map['keyword_ref_pmids'] = kw_ref_pmids_col
            if rename_map:
                stock_mapping = stock_mapping.rename(columns=rename_map)
        
        # Run functional validation (now optional; typically executed in split-stocks)
        if run_functional_validation:
            print(f"\n{'='*60}")
            print("Running functional validation...")
            stock_mapping = self._run_functional_validation(
                stock_mapping,
                references_df,
                component_to_pmids,
                keywords,
                soft_run=self.settings.soft_run
            )
        else:
            print(f"\n{'='*60}")
            print("Skipping functional validation in get-allele-refs.")
            print("  Validation will be performed in split-stocks for Ref++ output-sheet stocks.")
        
        # Save fulltext method cache and clean up
        self._fulltext_fetcher.save_cache()
        self._fulltext_fetcher.clear_cache()
        gc.collect()

        # Prepare Stocks sheet formatting
        if keywords is not None:
            old_kw_col = generate_keyword_column_name(keywords)
            if keywords:
                new_kw_col = f"Stock (allele) associated with at least 1 {' OR '.join(keywords)} reference?"
            else:
                new_kw_col = "Stock (allele) associated with at least 1 reference?"
            if old_kw_col in stock_mapping.columns:
                stock_mapping = stock_mapping.rename(columns={old_kw_col: new_kw_col})
        
        # Remove unwanted columns from Stocks sheet
        cols_to_drop = [
            'all_flybase_gene_ids',
            'all_gene_symbols',
            'flybase_gene_id',
            'PMCID',
            'pmcid',
            'ref_count',
            'relevant_flybase_gene_ids',
            'is_custom_stock'
        ]
        stock_mapping = stock_mapping.drop(columns=[c for c in cols_to_drop if c in stock_mapping.columns])
        
        # Reorder columns for Stocks sheet readability
        preferred_order = [
            'stock_number',
            'genotype',
            'flybase_genotype',
            'Ch # all',
            'mapstatement',
            'Stock comments',
            'comment1',
            'relevant_component_symbols',
            'relevant_gene_symbols',
            'all_component_symbols',
            'all_stock_genes',
            'all_stock_constructs',
            'UAS',
            'sgRNA',
        ]
        ordered_cols = [c for c in preferred_order if c in stock_mapping.columns]
        for col in stock_mapping.columns:
            if col not in ordered_cols:
                ordered_cols.append(col)
        stock_mapping = stock_mapping[ordered_cols]
        
        # Process References sheet
        if references_df is not None and len(references_df) > 0:
            # Remove unwanted columns: FBrf, PMCID, miniref, pmid_added
            cols_to_remove = ['FBrf', 'PMCID', 'miniref', 'pmid_added', 'source', 'resolved_at', 'has_pmid', 'has_pubtype']
            for col in cols_to_remove:
                if col in references_df.columns:
                    references_df = references_df.drop(columns=[col])
            
            # Get keyword column name
            kw_col_name = generate_keyword_column_name(keywords) if keywords else None
            
            # Sort references: first by keyword hits (True first), then by stock_count (descending)
            if kw_col_name and kw_col_name in references_df.columns:
                references_df = references_df.sort_values(
                    by=[kw_col_name, 'stock_count'],
                    ascending=[False, False]
                ).reset_index(drop=True)
            elif 'stock_count' in references_df.columns:
                references_df = references_df.sort_values(
                    by=['stock_count'],
                    ascending=[False]
                ).reset_index(drop=True)
            
            # Reorder columns for better readability
            preferred_order = [
                'PMID', 'title', 'associated_stocks', 'associated_genes', 
                'stock_count', 'publication_date', 'author(s)', 'journal'
            ]
            if kw_col_name:
                preferred_order.append(kw_col_name)
            
            # Build final column order
            final_cols = []
            for col in preferred_order:
                if col in references_df.columns:
                    final_cols.append(col)
            # Add any remaining columns
            for col in references_df.columns:
                if col not in final_cols:
                    final_cols.append(col)
            references_df = references_df[final_cols]
        
        # Save output
        output_path = output_dir / "aggregated_bdsc_stock_refs.xlsx"
        print(f"\nSaving output to: {output_path}")
        
        if self.settings.soft_run:
            # Soft run: drop GPT-derived columns entirely
            gpt_cols_to_drop = get_gpt_derived_columns(stock_mapping)
            stock_mapping_out = stock_mapping.drop(
                columns=[c for c in gpt_cols_to_drop if c in stock_mapping.columns]
            )
        else:
            # Apply [EXPERIMENTAL] prefix to GPT-derived column headers
            stock_mapping_out = apply_experimental_prefix(stock_mapping)
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            stock_mapping_out.to_excel(writer, sheet_name='Stocks', index=False)
            if references_df is not None and len(references_df) > 0:
                references_df.to_excel(writer, sheet_name='References', index=False)
            
            if not self.settings.soft_run:
                # Apply light grey fill to GPT-derived column headers only
                _apply_grey_fill_openpyxl(
                    writer.sheets['Stocks'], stock_mapping_out
                )
        
        # Write surviving filtered FBrfs without PMID to a sidecar report.
        no_pmid_report = output_dir / "references_without_pmid_fbrf.txt"
        no_pmid_rows = sorted(
            self._no_pmid_fbrf_records,
            key=lambda r: (r.get('associated_stocks', ''), r.get('FBrf', ''))
        )
        with open(no_pmid_report, 'w', encoding='utf-8') as f:
            f.write("Filtered-in FlyBase references without PMID\n")
            f.write(f"Generated: {datetime.now().isoformat(timespec='seconds')}\n")
            f.write(f"Count: {len(no_pmid_rows)}\n")
            f.write("\n")
            for row in no_pmid_rows:
                f.write(f"FBrf: {row.get('FBrf', '')}\n")
                f.write(f"  Stocks: {row.get('associated_stocks', '')}\n")
                f.write(f"  Components: {row.get('associated_components', '')}\n")
                f.write(f"  pub_type: {row.get('pub_type', '')}\n")
                if row.get('miniref', ''):
                    f.write(f"  miniref: {row.get('miniref', '')}\n")
                f.write("\n")
        print(f"  Saved no-PMID FBrf report: {no_pmid_report}")
        
        print(f"\n{'='*60}")
        print("COMPLETE")
        print(f"  Input files: {len(valid_files)}")
        print(f"  Total genes: {len(unique_genes)}")
        print(f"  Output: {output_path}")
        
        return output_path
